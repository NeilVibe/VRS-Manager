"""
Update Excel process guides with Super Group Word Analysis information.

Adds information about the new "Super Group Word Analysis" sheet to:
- VRS_Manager_Process_Guide_EN.xlsx
- VRS_Manager_Process_Guide_KR.xlsx
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def update_english_guide():
    """Update English guide with Super Group Word Analysis info."""

    wb = openpyxl.load_workbook('VRS_Manager_Process_Guide_EN.xlsx')

    # Find or create "Technical Reference" sheet
    if "Technical Reference" in wb.sheetnames:
        ws = wb["Technical Reference"]
    else:
        ws = wb.create_sheet("Technical Reference")

    # Find next empty row
    last_row = ws.max_row
    start_row = last_row + 3

    # Add Super Group Word Analysis section
    ws[f"A{start_row}"] = "Super Group Word Analysis Sheet"
    ws[f"A{start_row}"].font = Font(bold=True, size=14)
    ws[f"A{start_row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{start_row}"].font = Font(bold=True, size=14, color="FFFFFF")

    start_row += 2

    content = [
        ["Purpose:", "Track word count changes and translation progress at super group level"],
        ["Location:", "Available in RAW VRS CHECK output (Raw Process)"],
        ["", "Sheet name: 'Super Group Word Analysis'"],
        ["", ""],
        ["Super Groups:", ""],
        ["  1. Main Chapters", "Group contains 'chapter' OR Group in ['intro', 'prolog']"],
        ["  2. Faction 1", "Group = 'faction_01'"],
        ["  3. Faction 2", "Group = 'faction_02'"],
        ["  4. Faction 3", "Group = 'faction_03'"],
        ["  5. Faction ETC", "Group = 'faction_etc'"],
        ["  6. Quest Dialog", "DialogType contains 'questdialog' (case-insensitive)"],
        ["  7. AI Dialog", "DialogType contains 'aidialog' (case-insensitive)"],
        ["  8. Others", "DialogType contains 'stageclosedialog' (case-insensitive)"],
        ["  9. Other", "Default for rows that don't match any super group"],
        ["", ""],
        ["Columns:", ""],
        ["  • Super Group Name", "Name of the super group"],
        ["  • Total Words (Previous)", "Total word count in previous file"],
        ["  • Total Words (Current)", "Total word count in current file"],
        ["  • Words Added", "Words from new rows in this super group"],
        ["  • Words Deleted", "Words from deleted rows in this super group"],
        ["  • Words Changed", "Words from rows where StrOrigin changed"],
        ["  • Words Unchanged", "Words from rows with no StrOrigin changes"],
        ["  • Words Migrated In", "Words that moved INTO this super group from another"],
        ["  • Words Migrated Out", "Words that moved OUT of this super group to another"],
        ["  • Net Change", "Total Words (Current) - Total Words (Previous)"],
        ["  • % Change", "Percentage change in total words"],
        ["  • Translated Words", "Words where Text does NOT contain 'NO TRANSLATION'"],
        ["  • Untranslated Words", "Words where Text contains 'NO TRANSLATION'"],
        ["  • Translation %", "Percentage of words that are translated"],
        ["", ""],
        ["Use Cases:", ""],
        ["  • Voice Actors", "See which super groups need translation work"],
        ["  • Project Managers", "Track translation progress by super group"],
        ["  • Stakeholders", "High-level view of content changes and translation status"],
        ["", ""],
        ["Important Notes:", ""],
        ["  • Word count based on StrOrigin column (space-separated words)"],
        ["  • DialogType rules have PRIORITY over Group rules"],
        ["  • All matching is case-insensitive"],
        ["  • Translated + Untranslated = Total Words (Current)"],
        ["  • Migration tracking: Words moving between super groups"],
    ]

    for row_data in content:
        ws[f"A{start_row}"] = row_data[0]
        if len(row_data) > 1:
            ws[f"B{start_row}"] = row_data[1]
        start_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 70

    wb.save('VRS_Manager_Process_Guide_EN.xlsx')
    print("✓ Updated VRS_Manager_Process_Guide_EN.xlsx")


def update_korean_guide():
    """Update Korean guide with Super Group Word Analysis info."""

    wb = openpyxl.load_workbook('VRS_Manager_Process_Guide_KR.xlsx')

    # Find or create "Technical Reference" sheet
    if "Technical Reference" in wb.sheetnames:
        ws = wb["Technical Reference"]
    elif "기술 참조" in wb.sheetnames:
        ws = wb["기술 참조"]
    else:
        ws = wb.create_sheet("Technical Reference")

    # Find next empty row
    last_row = ws.max_row
    start_row = last_row + 3

    # Add Super Group Word Analysis section (Korean)
    ws[f"A{start_row}"] = "슈퍼 그룹 단어 분석 시트"
    ws[f"A{start_row}"].font = Font(bold=True, size=14)
    ws[f"A{start_row}"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f"A{start_row}"].font = Font(bold=True, size=14, color="FFFFFF")

    start_row += 2

    content = [
        ["목적:", "슈퍼 그룹 수준에서 단어 수 변경 및 번역 진행 상황 추적"],
        ["위치:", "RAW VRS CHECK 출력에서 사용 가능 (Raw Process)"],
        ["", "시트 이름: 'Super Group Word Analysis'"],
        ["", ""],
        ["슈퍼 그룹:", ""],
        ["  1. Main Chapters", "Group에 'chapter' 포함 OR Group이 ['intro', 'prolog']에 해당"],
        ["  2. Faction 1", "Group = 'faction_01'"],
        ["  3. Faction 2", "Group = 'faction_02'"],
        ["  4. Faction 3", "Group = 'faction_03'"],
        ["  5. Faction ETC", "Group = 'faction_etc'"],
        ["  6. Quest Dialog", "DialogType에 'questdialog' 포함 (대소문자 무시)"],
        ["  7. AI Dialog", "DialogType에 'aidialog' 포함 (대소문자 무시)"],
        ["  8. Others", "DialogType에 'stageclosedialog' 포함 (대소문자 무시)"],
        ["  9. Other", "어떤 슈퍼 그룹과도 일치하지 않는 행의 기본값"],
        ["", ""],
        ["열:", ""],
        ["  • Super Group Name", "슈퍼 그룹 이름"],
        ["  • Total Words (Previous)", "이전 파일의 총 단어 수"],
        ["  • Total Words (Current)", "현재 파일의 총 단어 수"],
        ["  • Words Added", "이 슈퍼 그룹의 새 행에서 나온 단어"],
        ["  • Words Deleted", "이 슈퍼 그룹의 삭제된 행에서 나온 단어"],
        ["  • Words Changed", "StrOrigin이 변경된 행의 단어"],
        ["  • Words Unchanged", "StrOrigin 변경 없는 행의 단어"],
        ["  • Words Migrated In", "다른 슈퍼 그룹에서 이 슈퍼 그룹으로 이동한 단어"],
        ["  • Words Migrated Out", "이 슈퍼 그룹에서 다른 슈퍼 그룹으로 이동한 단어"],
        ["  • Net Change", "Total Words (Current) - Total Words (Previous)"],
        ["  • % Change", "총 단어의 백분율 변경"],
        ["  • Translated Words", "Text에 'NO TRANSLATION'이 포함되지 않은 단어"],
        ["  • Untranslated Words", "Text에 'NO TRANSLATION'이 포함된 단어"],
        ["  • Translation %", "번역된 단어의 백분율"],
        ["", ""],
        ["사용 사례:", ""],
        ["  • 성우", "번역 작업이 필요한 슈퍼 그룹 확인"],
        ["  • 프로젝트 관리자", "슈퍼 그룹별 번역 진행 상황 추적"],
        ["  • 이해 관계자", "콘텐츠 변경 및 번역 상태에 대한 높은 수준의 보기"],
        ["", ""],
        ["중요 사항:", ""],
        ["  • 단어 수는 StrOrigin 열 기반 (공백으로 구분된 단어)"],
        ["  • DialogType 규칙이 Group 규칙보다 우선"],
        ["  • 모든 매칭은 대소문자를 무시합니다"],
        ["  • 번역됨 + 번역되지 않음 = 총 단어 (현재)"],
        ["  • 마이그레이션 추적: 슈퍼 그룹 간 이동하는 단어"],
    ]

    for row_data in content:
        ws[f"A{start_row}"] = row_data[0]
        if len(row_data) > 1:
            ws[f"B{start_row}"] = row_data[1]
        start_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 70

    wb.save('VRS_Manager_Process_Guide_KR.xlsx')
    print("✓ Updated VRS_Manager_Process_Guide_KR.xlsx")


if __name__ == "__main__":
    print("Updating Excel process guides with Super Group Word Analysis information...")
    print()
    update_english_guide()
    update_korean_guide()
    print()
    print("✓ All guides updated successfully!")
