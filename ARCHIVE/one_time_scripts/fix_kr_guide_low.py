#!/usr/bin/env python3
"""
Fix Korean Excel guide LOW importance documentation
"""

import openpyxl
from openpyxl.styles import Font

def fix_kr_guide():
    """Fix Korean Excel guide LOW importance documentation"""
    print("Fixing VRS_Manager_Process_Guide_KR.xlsx LOW importance section...")

    wb = openpyxl.load_workbook('VRS_Manager_Process_Guide_KR.xlsx')

    if '4. Master Update' in wb.sheetnames:
        ws = wb['4. Master Update']

        # Find row 88-89 and update
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws[f'A{row_idx}'].value
            if cell_value and '6단계' in str(cell_value) and 'Low Importance' in str(cell_value):
                print(f"Found at row {row_idx}: {cell_value}")

                # Update the description
                ws[f'A{row_idx + 1}'].value = '  중요: Low importance 행은 High importance와 다른 로직을 사용합니다:'
                ws[f'A{row_idx + 1}'].font = Font(bold=True, size=11, color='FF0000')

                ws[f'A{row_idx + 2}'].value = '    • 기존 행 ("New Row" 아님): TARGET 데이터 유지 (원본 Master 데이터 보존)'
                ws[f'A{row_idx + 3}'].value = '    • 신규 행: 출력에서 제외 (Low importance 신규 콘텐츠는 삭제됨)'
                ws[f'A{row_idx + 4}'].value = '    • 삭제된 행: "Deleted Rows" 시트에 추적'
                ws[f'A{row_idx + 5}'].value = '  EventName 변경을 고려하기 위해 듀얼 룩업을 사용합니다.'

                ws[f'A{row_idx + 7}'].value = '  왜 이런 규칙을 사용하나요?'
                ws[f'A{row_idx + 7}'].font = Font(bold=True, size=11)

                ws[f'A{row_idx + 8}'].value = '    • LOW = 중요하지 않은 콘텐츠 → Master에 이미 있는 내용 유지'
                ws[f'A{row_idx + 9}'].value = '    • HIGH = 중요한 콘텐츠 → 항상 SOURCE 데이터로 업데이트'

                print(f"Updated rows {row_idx + 1} through {row_idx + 9}")
                break

        # Also update Sheet 2 description if exists
        for row_idx in range(130, ws.max_row + 1):
            cell_value = ws[f'A{row_idx}'].value
            if cell_value and 'Low Importance' in str(cell_value) and '시트' in str(cell_value):
                ws[f'A{row_idx + 1}'].value = '  TARGET 데이터가 보존된(기존) 또는 제외된(신규) Low importance 행.'
                print(f"Updated sheet description at row {row_idx + 1}")
                break

    wb.save('VRS_Manager_Process_Guide_KR.xlsx')
    print("✅ Korean guide fixed!")

if __name__ == '__main__':
    fix_kr_guide()
    print("\n🎉 Korean Excel guide updated with correct LOW importance documentation!")
