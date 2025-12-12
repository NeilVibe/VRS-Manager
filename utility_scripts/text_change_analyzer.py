"""
Text Change Analyzer - Standalone Utility Script

Compares two Excel files (PREVIOUS and CURRENT) and detects text changes
based on StrOrigin+EventName matching.

Usage:
    python text_change_analyzer.py
    (Opens file dialogs to select PREVIOUS and CURRENT Excel files)

Output:
    Creates output file in same folder as CURRENT file with "_TextChanges" suffix.

Author: Neil Schmitt
"""

import os
import sys
from difflib import SequenceMatcher

try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas not installed. Run: pip install pandas openpyxl")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# DIFF FUNCTIONS (based on strorigin_analysis.py pattern)
# =============================================================================

def extract_text_differences(text1: str, text2: str, max_length: int = 150) -> str:
    """
    Extract WORD-LEVEL differences between two texts using difflib.

    Shows exactly what changed in WinMerge style:
    - [old words->new words] for replacements
    - [-deleted words] for deletions
    - [+added words] for additions

    Args:
        text1: Previous text
        text2: Current text
        max_length: Maximum length for diff output (truncate if longer)

    Returns:
        Diff string showing changes, or empty string if identical
    """
    if not text1 or not text2:
        if not text1 and text2:
            return f"[+{text2[:max_length]}]" if len(str(text2)) <= max_length else f"[+{text2[:max_length-3]}...]"
        if text1 and not text2:
            return f"[-{text1[:max_length]}]" if len(str(text1)) <= max_length else f"[-{text1[:max_length-3]}...]"
        return ""

    text1 = str(text1)
    text2 = str(text2)

    # Split into words (preserves Korean and English spacing)
    words1 = text1.split()
    words2 = text2.split()

    # Word-level diff (automatic chunking of consecutive changes)
    matcher = SequenceMatcher(None, words1, words2)
    changes = []

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'replace':
            old_words = ' '.join(words1[i1:i2])
            new_words = ' '.join(words2[j1:j2])
            changes.append(f"[{old_words}->{new_words}]")
        elif tag == 'delete':
            deleted_words = ' '.join(words1[i1:i2])
            changes.append(f"[-{deleted_words}]")
        elif tag == 'insert':
            added_words = ' '.join(words2[j1:j2])
            changes.append(f"[+{added_words}]")

    if not changes:
        return ""

    diff_str = ' '.join(changes)

    # Truncate if too long (avoid huge Excel cells)
    if len(diff_str) > max_length:
        diff_str = diff_str[:max_length - 3] + "..."

    return diff_str


def safe_str(value) -> str:
    """Safely convert value to string, handling NaN/None."""
    if pd.isna(value) or value is None:
        return ""
    return str(value).strip()


# =============================================================================
# MAIN LOGIC
# =============================================================================

def select_file(title: str) -> str:
    """Open file dialog and return selected path."""
    from tkinter import Tk, filedialog
    root = Tk()
    root.withdraw()  # Hide main window
    root.attributes('-topmost', True)  # Bring dialog to front

    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
    )
    root.destroy()
    return file_path


def show_error(title: str, message: str):
    """Show error message box."""
    from tkinter import Tk, messagebox
    root = Tk()
    root.withdraw()
    messagebox.showerror(title, message)
    root.destroy()


def show_info(title: str, message: str):
    """Show info message box."""
    from tkinter import Tk, messagebox
    root = Tk()
    root.withdraw()
    messagebox.showinfo(title, message)
    root.destroy()


def validate_columns(df: pd.DataFrame, file_name: str) -> bool:
    """Check if DataFrame has required columns."""
    required = ["StrOrigin", "Text", "EventName"]
    missing = [col for col in required if col not in df.columns]

    if missing:
        show_error(
            "Missing Columns",
            f"File '{file_name}' is missing required columns:\n{', '.join(missing)}\n\n"
            f"Required: StrOrigin, Text, EventName"
        )
        return False
    return True


def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reorder columns: StrOrigin, Text, Text_Change, Previous_Text, Text_Diff, then rest.
    """
    priority_cols = ["StrOrigin", "Text", "Text_Change", "Previous_Text", "Text_Diff"]

    # Get columns that exist in priority order
    ordered = [col for col in priority_cols if col in df.columns]

    # Add remaining columns
    remaining = [col for col in df.columns if col not in priority_cols]

    return df[ordered + remaining]


def apply_styling(ws):
    """
    Apply Excel styling: light blue header, orange for Text_Change cells, filter.
    """
    # Style definitions
    header_fill = PatternFill(start_color="ADD8E6", fill_type="solid")  # Light blue
    header_font = Font(bold=True)
    change_fill = PatternFill(start_color="FFD580", fill_type="solid")  # Orange

    # Find Text_Change column index
    text_change_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        # Apply header styling
        cell.fill = header_fill
        cell.font = header_font

        if cell.value == "Text_Change":
            text_change_col_idx = idx

    # Apply orange fill to Text_Change cells with "Text Change" value
    if text_change_col_idx:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=text_change_col_idx, max_col=text_change_col_idx):
            cell = row[0]
            if cell.value == "Text Change":
                cell.fill = change_fill

    # Set column widths
    col_widths = {
        "StrOrigin": 25,
        "Text": 50,
        "Text_Change": 15,
        "Previous_Text": 50,
        "Text_Diff": 60,
        "EventName": 25,
        "Group": 20,
    }

    for idx, cell in enumerate(ws[1], start=1):
        col_name = cell.value
        if col_name in col_widths:
            ws.column_dimensions[get_column_letter(idx)].width = col_widths[col_name]
        else:
            ws.column_dimensions[get_column_letter(idx)].width = 15

    # Add filter to header row
    ws.auto_filter.ref = ws.dimensions


def analyze_text_changes(prev_df: pd.DataFrame, curr_df: pd.DataFrame) -> pd.DataFrame:
    """
    Compare PREVIOUS and CURRENT DataFrames, detect text changes.

    Match key: StrOrigin + EventName
    Output: CURRENT with added columns for text change analysis
    """
    # Build lookup from PREVIOUS: key = (StrOrigin, EventName) -> Text
    prev_lookup = {}
    for _, row in prev_df.iterrows():
        str_origin = safe_str(row.get("StrOrigin", ""))
        event_name = safe_str(row.get("EventName", ""))
        text = safe_str(row.get("Text", ""))

        if str_origin and event_name:  # Only add if key fields exist
            key = (str_origin, event_name)
            prev_lookup[key] = text

    # Process CURRENT and add analysis columns
    result_df = curr_df.copy()

    text_change_col = []
    prev_text_col = []
    text_diff_col = []

    for _, row in curr_df.iterrows():
        str_origin = safe_str(row.get("StrOrigin", ""))
        event_name = safe_str(row.get("EventName", ""))
        curr_text = safe_str(row.get("Text", ""))

        key = (str_origin, event_name)

        if key in prev_lookup:
            prev_text = prev_lookup[key]

            if prev_text != curr_text:
                # Text changed!
                text_change_col.append("Text Change")
                prev_text_col.append(prev_text)
                text_diff_col.append(extract_text_differences(prev_text, curr_text))
            else:
                # Same text
                text_change_col.append("")
                prev_text_col.append("")
                text_diff_col.append("")
        else:
            # Key not found in PREVIOUS (new row)
            text_change_col.append("")
            prev_text_col.append("")
            text_diff_col.append("")

    # Add columns to result
    result_df["Text_Change"] = text_change_col
    result_df["Previous_Text"] = prev_text_col
    result_df["Text_Diff"] = text_diff_col

    return result_df


def main():
    """Main entry point."""
    print("=" * 60)
    print("Text Change Analyzer")
    print("=" * 60)
    print()

    # Select PREVIOUS file
    print("Step 1: Select PREVIOUS Excel file...")
    prev_path = select_file("Select PREVIOUS Excel File")
    if not prev_path:
        print("Cancelled - no PREVIOUS file selected.")
        return
    print(f"  -> {os.path.basename(prev_path)}")

    # Select CURRENT file
    print("\nStep 2: Select CURRENT Excel file...")
    curr_path = select_file("Select CURRENT Excel File")
    if not curr_path:
        print("Cancelled - no CURRENT file selected.")
        return
    print(f"  -> {os.path.basename(curr_path)}")

    # Load files
    print("\nStep 3: Loading Excel files...")
    try:
        prev_df = pd.read_excel(prev_path)
        print(f"  PREVIOUS: {len(prev_df)} rows")
    except Exception as e:
        show_error("Error", f"Failed to load PREVIOUS file:\n{e}")
        return

    try:
        curr_df = pd.read_excel(curr_path)
        print(f"  CURRENT:  {len(curr_df)} rows")
    except Exception as e:
        show_error("Error", f"Failed to load CURRENT file:\n{e}")
        return

    # Validate columns
    print("\nStep 4: Validating columns...")
    if not validate_columns(prev_df, os.path.basename(prev_path)):
        return
    if not validate_columns(curr_df, os.path.basename(curr_path)):
        return
    print("  All required columns found!")

    # Analyze
    print("\nStep 5: Analyzing text changes...")
    result_df = analyze_text_changes(prev_df, curr_df)

    # Count changes
    change_count = (result_df["Text_Change"] == "Text Change").sum()
    print(f"  Found {change_count} text changes")

    # Reorder columns
    print("\nStep 6: Reordering columns...")
    result_df = reorder_columns(result_df)
    print("  StrOrigin -> Text -> Text_Change -> Previous_Text -> Text_Diff -> [rest]")

    # Save output
    print("\nStep 7: Saving output with styling...")
    curr_dir = os.path.dirname(curr_path)
    curr_name = os.path.splitext(os.path.basename(curr_path))[0]
    output_path = os.path.join(curr_dir, f"{curr_name}_TextChanges.xlsx")

    try:
        result_df.to_excel(output_path, index=False)

        # Apply styling
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        apply_styling(ws)
        wb.save(output_path)

        print(f"  -> {output_path}")
    except Exception as e:
        show_error("Error", f"Failed to save output:\n{e}")
        return

    # Done
    print("\n" + "=" * 60)
    print("COMPLETE!")
    print(f"  Text Changes Found: {change_count}")
    print(f"  Output: {output_path}")
    print("=" * 60)

    show_info(
        "Complete",
        f"Text Change Analysis Complete!\n\n"
        f"Text Changes Found: {change_count}\n\n"
        f"Output saved to:\n{output_path}"
    )


if __name__ == "__main__":
    main()
