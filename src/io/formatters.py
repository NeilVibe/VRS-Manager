"""
Excel formatting module.

This module provides functionality to apply colors, styles, and formatting
to Excel worksheets for VRS Manager output files.
"""

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from src.config import CHAR_GROUP_COLS


def generate_color_for_value(value):
    """
    Generate a unique, consistent color for a given value using deterministic HSL-based generation.

    This function generates visually distinct colors for different values by:
    1. Using a stable hash of the value to generate a hue (0-360 degrees)
    2. Using golden ratio distribution for better hue spread
    3. Keeping saturation and lightness in visually pleasant ranges

    Args:
        value: Any value to generate a color for

    Returns:
        str: Hex color code (without #)
    """
    import hashlib

    # Use hashlib for consistent hashing across Python sessions
    value_str = str(value)
    hash_bytes = hashlib.md5(value_str.encode()).digest()

    # Extract 3 bytes for H, S, L components
    h_raw = hash_bytes[0]  # 0-255
    s_raw = hash_bytes[1]  # 0-255
    l_raw = hash_bytes[2]  # 0-255

    # Golden ratio for better hue distribution
    golden_ratio = 0.618033988749895
    hue = (h_raw / 255.0 + golden_ratio * (hash_bytes[3] % 20)) % 1.0  # 0-1

    # Saturation: 40-70% for pastel colors (readable on white background)
    saturation = 0.40 + (s_raw / 255.0) * 0.30  # 0.40-0.70

    # Lightness: 70-85% for light pastel colors (readable text on top)
    lightness = 0.70 + (l_raw / 255.0) * 0.15  # 0.70-0.85

    # Convert HSL to RGB
    def hsl_to_rgb(h, s, l):
        if s == 0:
            r = g = b = l
        else:
            def hue_to_rgb(p, q, t):
                if t < 0:
                    t += 1
                if t > 1:
                    t -= 1
                if t < 1/6:
                    return p + (q - p) * 6 * t
                if t < 1/2:
                    return q
                if t < 2/3:
                    return p + (q - p) * (2/3 - t) * 6
                return p

            q = l * (1 + s) if l < 0.5 else l + s - l * s
            p = 2 * l - q
            r = hue_to_rgb(p, q, h + 1/3)
            g = hue_to_rgb(p, q, h)
            b = hue_to_rgb(p, q, h - 1/3)

        return int(r * 255), int(g * 255), int(b * 255)

    r, g, b = hsl_to_rgb(hue, saturation, lightness)
    return f"{r:02X}{g:02X}{b:02X}"


def apply_direct_coloring(ws, is_master=False, changed_columns_map=None):
    """
    Apply direct coloring to worksheet cells based on change types and status values.

    This function colors:
    - Header row
    - CHANGES column based on change type
    - STATUS columns based on status value
    - Character group columns when they've changed

    Args:
        ws: openpyxl worksheet to format
        is_master: If True, looks for STATUS_KR/EN/CN columns (default: False)
        changed_columns_map: Dictionary mapping row indices to lists of changed column names
    """
    changes_col_idx = None
    status_col_indices = {}
    char_group_col_indices = {}

    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "CHANGES":
            changes_col_idx = idx
        elif is_master:
            if cell.value in ["STATUS_KR", "STATUS_EN", "STATUS_CN"]:
                status_col_indices[cell.value] = idx
        else:
            if cell.value and cell.value.upper() == "STATUS":
                status_col_indices["STATUS"] = idx

        if cell.value in CHAR_GROUP_COLS:
            char_group_col_indices[cell.value] = idx

    changes_fills = {
        # Pure changes
        "StrOrigin Change": PatternFill(start_color="FFD580", fill_type="solid"),
        "Desc Change": PatternFill(start_color="E1D5FF", fill_type="solid"),
        "TimeFrame Change": PatternFill(start_color="FF9999", fill_type="solid"),
        "EventName Change": PatternFill(start_color="FFFF99", fill_type="solid"),
        "SequenceName Change": PatternFill(start_color="B3E5FC", fill_type="solid"),
        "CastingKey Change": PatternFill(start_color="FFB347", fill_type="solid"),  # Orange

        # Stage 1 composites (without EventName/SequenceName)
        "StrOrigin+Desc Change": PatternFill(start_color="FFA07A", fill_type="solid"),
        "StrOrigin+TimeFrame Change": PatternFill(start_color="FFB6C1", fill_type="solid"),
        "Desc+TimeFrame Change": PatternFill(start_color="DDA0DD", fill_type="solid"),
        "StrOrigin+Desc+TimeFrame Change": PatternFill(start_color="F08080", fill_type="solid"),
        "CastingKey+StrOrigin Change": PatternFill(start_color="FFAA7F", fill_type="solid"),
        "CastingKey+Desc Change": PatternFill(start_color="F0C8FF", fill_type="solid"),
        "CastingKey+TimeFrame Change": PatternFill(start_color="FFB3CC", fill_type="solid"),
        "CastingKey+StrOrigin+Desc Change": PatternFill(start_color="FF9F8F", fill_type="solid"),
        "CastingKey+StrOrigin+TimeFrame Change": PatternFill(start_color="FFC0CB", fill_type="solid"),
        "CastingKey+Desc+TimeFrame Change": PatternFill(start_color="E6C3E6", fill_type="solid"),
        "CastingKey+StrOrigin+Desc+TimeFrame Change": PatternFill(start_color="FF9999", fill_type="solid"),

        # EventName composites (Stage 2)
        "EventName+Desc Change": PatternFill(start_color="F0E68C", fill_type="solid"),
        "EventName+TimeFrame Change": PatternFill(start_color="FFDAB9", fill_type="solid"),
        "EventName+Desc+TimeFrame Change": PatternFill(start_color="FFD8A8", fill_type="solid"),
        "EventName+CastingKey Change": PatternFill(start_color="FFD966", fill_type="solid"),  # Yellow-orange
        "EventName+CastingKey+Desc Change": PatternFill(start_color="FFCC66", fill_type="solid"),
        "EventName+CastingKey+TimeFrame Change": PatternFill(start_color="FFC966", fill_type="solid"),
        "EventName+CastingKey+Desc+TimeFrame Change": PatternFill(start_color="FFB84D", fill_type="solid"),

        # SequenceName composites (Stage 3)
        "SequenceName+CastingKey Change": PatternFill(start_color="A0D9FF", fill_type="solid"),  # Light blue
        "SequenceName+Desc Change": PatternFill(start_color="C4E5FF", fill_type="solid"),
        "SequenceName+TimeFrame Change": PatternFill(start_color="B3D9FF", fill_type="solid"),
        "SequenceName+CastingKey+Desc Change": PatternFill(start_color="99D6FF", fill_type="solid"),
        "SequenceName+CastingKey+TimeFrame Change": PatternFill(start_color="8CD3FF", fill_type="solid"),
        "SequenceName+Desc+TimeFrame Change": PatternFill(start_color="B8DBFF", fill_type="solid"),
        "SequenceName+CastingKey+Desc+TimeFrame Change": PatternFill(start_color="7FCCFF", fill_type="solid"),

        # Special cases
        "CharacterGroup Change": PatternFill(start_color="87CEFA", fill_type="solid"),
        "New Row": PatternFill(start_color="90EE90", fill_type="solid"),
        "No Relevant Change": PatternFill(start_color="D3D3D3", fill_type="solid"),
        "No Change": PatternFill(start_color="E8E8E8", fill_type="solid"),
    }

    char_group_change_fill = PatternFill(start_color="FFD700", fill_type="solid")

    status_fills = {
        "RECORDED": PatternFill(start_color="90EE90", fill_type="solid"),
        "POLISHED": PatternFill(start_color="E6D5FF", fill_type="solid"),
        "RE-RECORD": PatternFill(start_color="FFB3B3", fill_type="solid"),
        "RERECORD": PatternFill(start_color="FFB3B3", fill_type="solid"),
        "RE-RECORDED": PatternFill(start_color="C5E8C5", fill_type="solid"),
        "RERECORDED": PatternFill(start_color="C5E8C5", fill_type="solid"),
        "PREVIOUSLY RECORDED": PatternFill(start_color="FFFFE0", fill_type="solid"),
        "FINAL": PatternFill(start_color="87CEEB", fill_type="solid"),
        "SHIPPED": PatternFill(start_color="87CEEB", fill_type="solid"),
        "SPEC-OUT": PatternFill(start_color="FFC0CB", fill_type="solid"),
        "CHECK": PatternFill(start_color="FFE4B5", fill_type="solid"),
        "전달 완료": PatternFill(start_color="FFFFE0", fill_type="solid"),
        "녹음 완료": PatternFill(start_color="90EE90", fill_type="solid"),
        "재녹음 필요": PatternFill(start_color="FFB3B3", fill_type="solid"),
        "재녹음 완료": PatternFill(start_color="C5E8C5", fill_type="solid"),
        "준비 중": PatternFill(start_color="E6D5FF", fill_type="solid"),
        "확인 필요": PatternFill(start_color="FFE4B5", fill_type="solid"),
        "已传达": PatternFill(start_color="FFFFE0", fill_type="solid"),
        "已录音": PatternFill(start_color="90EE90", fill_type="solid"),
        "需补录": PatternFill(start_color="FFB3B3", fill_type="solid"),
        "已补录": PatternFill(start_color="C5E8C5", fill_type="solid"),
        "准备中": PatternFill(start_color="E6D5FF", fill_type="solid"),
        "需要确认": PatternFill(start_color="FFE4B5", fill_type="solid"),
    }

    header_fill = PatternFill(start_color="ADD8E6", fill_type="solid")

    for cell in ws[1]:
        cell.fill = header_fill

    if changes_col_idx:
        colored_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=changes_col_idx, max_col=changes_col_idx), start=0):
            cell = row[0]
            cell_value = cell.value
            if cell_value in changes_fills:
                cell.fill = changes_fills[cell_value]
                colored_count += 1
            elif cell_value and str(cell_value).strip() and "Change" in str(cell_value):
                # Fallback: Auto-generate color for unlisted composite changes
                cell.fill = PatternFill(
                    start_color=generate_color_for_value(cell_value),
                    fill_type="solid"
                )
                colored_count += 1

            # Highlight specific CharacterGroup columns (works for standalone and composite)
            # This is OUTSIDE if/elif to work for ALL change types that include CharacterGroup
            if "CharacterGroup" in str(cell_value) and changed_columns_map and row_idx in changed_columns_map:
                changed_cols = changed_columns_map[row_idx]
                for col_name in changed_cols:
                    if col_name in char_group_col_indices:
                        col_idx = char_group_col_indices[col_name]
                        char_cell = ws.cell(row=row_idx + 2, column=col_idx)
                        char_cell.fill = char_group_change_fill

        ws.column_dimensions[get_column_letter(changes_col_idx)].width = 40

    for status_col_name, status_col_idx in status_col_indices.items():
        colored_count = 0
        unknown_statuses = {}

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=status_col_idx, max_col=status_col_idx):
            cell = row[0]
            cell_value = cell.value

            if cell_value and str(cell_value).strip():
                if cell_value in status_fills:
                    cell.fill = status_fills[cell_value]
                    colored_count += 1
                else:
                    if cell_value not in unknown_statuses:
                        unknown_statuses[cell_value] = PatternFill(
                            start_color=generate_color_for_value(cell_value),
                            fill_type="solid"
                        )
                    cell.fill = unknown_statuses[cell_value]
                    colored_count += 1

        ws.column_dimensions[get_column_letter(status_col_idx)].width = 25

    ws.sheet_view.showGridLines = True
    ws.auto_filter.ref = ws.dimensions


def widen_summary_columns(ws):
    """
    Set appropriate column widths for summary sheets.

    Args:
        ws: openpyxl worksheet to format
    """
    ws.column_dimensions[get_column_letter(1)].width = 50
    ws.column_dimensions[get_column_letter(2)].width = 25
    ws.column_dimensions[get_column_letter(3)].width = 25
    ws.column_dimensions[get_column_letter(4)].width = 25
    ws.column_dimensions[get_column_letter(5)].width = 25
    ws.column_dimensions[get_column_letter(6)].width = 25


def format_update_history_sheet(ws):
    """
    Format the update history sheet with appropriate colors and styling.

    Args:
        ws: openpyxl worksheet to format
    """
    ws.column_dimensions[get_column_letter(1)].width = 80

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            if cell.value:
                content = str(cell.value)

                if "UPDATE HISTORY" in content or "LATEST UPDATE" in content:
                    cell.font = Font(bold=True, size=14, color="000080")
                    cell.fill = PatternFill(start_color="FFFFCC", fill_type="solid")

                elif "✓ UPDATED" in content:
                    cell.font = Font(bold=True, color="006600")
                    cell.fill = PatternFill(start_color="CCFFCC", fill_type="solid")

                elif "○ Preserved" in content:
                    cell.font = Font(italic=True, color="666666")
                    cell.fill = PatternFill(start_color="F0F0F0", fill_type="solid")

    ws.sheet_view.showGridLines = False
