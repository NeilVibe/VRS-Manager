"""
Excel file output formatting and styling.

This module provides functions for writing formatted Excel output files,
including conditional formatting, column widths, and color coding.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def write_group_word_analysis(writer, df_group_analysis, sheet_name="Group Word Analysis"):
    """
    Write group-level word count analysis to Excel sheet.

    Creates a formatted sheet showing word count statistics for each group,
    with conditional formatting and summary rows.

    Args:
        writer: ExcelWriter object
        df_group_analysis: DataFrame with group-level statistics
        sheet_name: Name of sheet to create (default: "Group Word Analysis")
    """
    if df_group_analysis is None or df_group_analysis.empty:
        # No group data to write
        return

    # Write to Excel
    df_group_analysis.to_excel(writer, sheet_name=sheet_name, index=False)

    # Get worksheet for formatting
    worksheet = writer.sheets[sheet_name]

    # Set column widths
    column_widths = {
        'A': 22,  # Group Name
        'B': 18,  # Total Words (Previous)
        'C': 18,  # Total Words (Current)
        'D': 14,  # Words Added (New Rows)
        'E': 14,  # Words Deleted (Removed Rows)
        'F': 14,  # Words Changed (StrOrigin modifications)
        'G': 14,  # Words Unchanged
        'H': 16,  # Words Migrated Out
        'I': 16,  # Words Migrated In
        'J': 14,  # Net Change (Current - Previous)
        'K': 12   # % Change
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # Apply header formatting
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Apply data formatting
    for row_num in range(2, len(df_group_analysis) + 2):
        # Number formatting for all numeric columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = worksheet[f"{col}{row_num}"]
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")

        # Net Change with color coding and signed format
        net_cell = worksheet[f"J{row_num}"]
        net_cell.number_format = '+#,##0;[Red]-#,##0;0'
        net_cell.alignment = Alignment(horizontal="right")

        # Apply color to Net Change
        if net_cell.value and isinstance(net_cell.value, (int, float)):
            if net_cell.value > 0:
                net_cell.font = Font(color="00B050", bold=True)  # Green
            elif net_cell.value < 0:
                net_cell.font = Font(color="FF0000", bold=True)  # Red

        # % Change alignment
        pct_cell = worksheet[f"K{row_num}"]
        pct_cell.alignment = Alignment(horizontal="right")

    # Add summary row
    summary_row = len(df_group_analysis) + 3
    worksheet[f"A{summary_row}"] = "TOTAL"
    worksheet[f"A{summary_row}"].font = Font(bold=True, size=11)

    # Sum columns B through J
    for col_letter, col_idx in [('B', 2), ('C', 3), ('D', 4), ('E', 5),
                                  ('F', 6), ('G', 7), ('H', 8), ('I', 9), ('J', 10)]:
        cell = worksheet[f"{col_letter}{summary_row}"]
        cell.value = f"=SUM({col_letter}2:{col_letter}{len(df_group_analysis)+1})"
        cell.font = Font(bold=True, size=11)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right")

    # Add border to summary row
    from openpyxl.styles import Border, Side
    thin_border = Border(top=Side(style='thin'))
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        worksheet[f"{col}{summary_row}"].border = thin_border

    # Freeze header row
    worksheet.freeze_panes = worksheet['A2']


def write_super_group_word_analysis(writer, super_group_analysis, migration_details=None, sheet_name="Super Group Word Analysis"):
    """
    Write super group-level word count analysis to Excel sheet with translation tracking.

    Creates a formatted sheet showing word count statistics for each super group,
    including translation progress metrics.

    Args:
        writer: ExcelWriter object
        super_group_analysis: Dictionary with super group statistics (including translation metrics)
        migration_details: List of tuples (source_group, dest_group, word_count) or None
        sheet_name: Name of sheet to create (default: "Super Group Word Analysis")
    """
    if not super_group_analysis:
        # No super group data to write
        return

    # Convert super_group_analysis dict to DataFrame
    rows = []
    for super_group_name, stats in sorted(super_group_analysis.items()):
        prev_total = stats["total_words_prev"]
        curr_total = stats["total_words_curr"]
        net_change = curr_total - prev_total

        # Calculate percentage change
        if prev_total > 0:
            pct_change = (net_change / prev_total) * 100
            pct_change_str = f"{pct_change:+.2f}%"
        else:
            pct_change_str = "N/A" if curr_total > 0 else "0.00%"

        # Calculate current translation percentage
        translated_words = stats.get("translated_words", 0)
        untranslated_words = stats.get("untranslated_words", 0)
        if curr_total > 0:
            translation_pct_curr = (translated_words / curr_total) * 100
            translation_pct_curr_str = f"{translation_pct_curr:.1f}%"
        else:
            translation_pct_curr_str = "N/A"

        # Calculate previous translation percentage
        translated_words_prev = stats.get("translated_words_prev", 0)
        untranslated_words_prev = stats.get("untranslated_words_prev", 0)
        if prev_total > 0:
            translation_pct_prev = (translated_words_prev / prev_total) * 100
            translation_pct_prev_str = f"{translation_pct_prev:.1f}%"
        else:
            translation_pct_prev_str = "N/A"

        # Calculate translation percentage change
        if prev_total > 0 and curr_total > 0:
            translation_pct_change = translation_pct_curr - translation_pct_prev
            translation_pct_change_str = f"{translation_pct_change:+.1f}%"
        else:
            translation_pct_change_str = "N/A"

        rows.append({
            "Super Group Name": super_group_name,
            "Total Words (Current)": curr_total,
            "Total Words (Previous)": prev_total,
            "Net Change": net_change,
            "% Change": pct_change_str,
            "Translated Words": translated_words,
            "Not Translated": untranslated_words,
            "Translation % (Current)": translation_pct_curr_str,
            "Translation % (Previous)": translation_pct_prev_str,
            "Translation % Change": translation_pct_change_str,
            "Words Added": stats["added_words"],
            "Words Deleted": stats["deleted_words"],
            "Words Changed": stats["changed_words"],
            "Words Unchanged": stats["unchanged_words"]
        })

    df_super_group_analysis = pd.DataFrame(rows)

    # Write to Excel
    df_super_group_analysis.to_excel(writer, sheet_name=sheet_name, index=False)

    # Get worksheet for formatting
    worksheet = writer.sheets[sheet_name]

    # Set column widths (new column order)
    column_widths = {
        'A': 22,  # Super Group Name
        'B': 18,  # Total Words (Current)
        'C': 18,  # Total Words (Previous)
        'D': 14,  # Net Change
        'E': 12,  # % Change
        'F': 16,  # Translated Words
        'G': 16,  # Not Translated
        'H': 18,  # Translation % (Current)
        'I': 18,  # Translation % (Previous)
        'J': 18,  # Translation % Change
        'K': 14,  # Words Added
        'L': 14,  # Words Deleted
        'M': 14,  # Words Changed
        'N': 14   # Words Unchanged
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # Apply header formatting
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Apply data formatting
    for row_num in range(2, len(df_super_group_analysis) + 2):
        # Number formatting for numeric columns (B, C, F, G, K-N)
        for col in ['B', 'C', 'F', 'G', 'K', 'L', 'M', 'N']:
            cell = worksheet[f"{col}{row_num}"]
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")

        # Net Change (D) with color coding and signed format
        net_cell = worksheet[f"D{row_num}"]
        net_cell.number_format = '+#,##0;[Red]-#,##0;0'
        net_cell.alignment = Alignment(horizontal="right")

        # Apply color to Net Change
        if net_cell.value and isinstance(net_cell.value, (int, float)):
            if net_cell.value > 0:
                net_cell.font = Font(color="00B050", bold=True)  # Green
            elif net_cell.value < 0:
                net_cell.font = Font(color="FF0000", bold=True)  # Red

    # Add summary row
    summary_row = len(df_super_group_analysis) + 3
    worksheet[f"A{summary_row}"] = "TOTAL"
    worksheet[f"A{summary_row}"].font = Font(bold=True, size=11)

    # Sum columns: B, C, D, F, G, K, L, M, N (numeric columns only)
    for col_letter in ['B', 'C', 'D', 'F', 'G', 'K', 'L', 'M', 'N']:
        cell = worksheet[f"{col_letter}{summary_row}"]
        cell.value = f"=SUM({col_letter}2:{col_letter}{len(df_super_group_analysis)+1})"
        cell.font = Font(bold=True, size=11)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right")

    # Add border to summary row
    from openpyxl.styles import Border, Side
    thin_border = Border(top=Side(style='thin'))
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        worksheet[f"{col}{summary_row}"].border = thin_border

    # Add explanatory notes below the table
    note_row = summary_row + 2
    worksheet[f"A{note_row}"] = "Notes:"
    worksheet[f"A{note_row}"].font = Font(bold=True, size=10, italic=True)

    # Note 1: NET CHANGE explanation
    note_text_row_1 = note_row + 1
    note_text_1 = (
        'NET CHANGE = Total Words (Current) - Total Words (Previous). '
        'Positive (+) = words added to the project. Negative (-) = words removed from the project.'
    )
    worksheet[f"A{note_text_row_1}"] = note_text_1
    worksheet[f"A{note_text_row_1}"].font = Font(size=9, italic=True, color="666666")
    worksheet[f"A{note_text_row_1}"].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.merge_cells(f"A{note_text_row_1}:G{note_text_row_1}")
    worksheet.row_dimensions[note_text_row_1].height = 30

    # Note 2: Main Chapters
    note_text_row_2 = note_text_row_1 + 1
    note_text_2 = (
        'Main Chapters: Groups containing "chapter", "intro", "prolog", or "epilog" (keyword-based). '
        'This super group aggregates all main story dialogue.'
    )
    worksheet[f"A{note_text_row_2}"] = note_text_2
    worksheet[f"A{note_text_row_2}"].font = Font(size=9, italic=True, color="666666")
    worksheet[f"A{note_text_row_2}"].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.merge_cells(f"A{note_text_row_2}:G{note_text_row_2}")
    worksheet.row_dimensions[note_text_row_2].height = 30

    # Note 3: Other
    note_text_row_3 = note_text_row_2 + 1
    note_text_3 = (
        'Other: Specific game systems - Police, Minigame, Trade, Church, Shop, Contribution, RoyalSupply, Research, '
        'Quest Groups (Hernand, Demeniss, Delesyia), faction_etc, Item. This super group aggregates system/feature dialogue.'
    )
    worksheet[f"A{note_text_row_3}"] = note_text_3
    worksheet[f"A{note_text_row_3}"].font = Font(size=9, italic=True, color="666666")
    worksheet[f"A{note_text_row_3}"].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.merge_cells(f"A{note_text_row_3}:G{note_text_row_3}")
    worksheet.row_dimensions[note_text_row_3].height = 30

    # Note 4: Everything Else
    note_text_row_4 = note_text_row_3 + 1
    note_text_4 = (
        'Everything Else: All groups that do not match any other Super Group classification. '
        'This includes groups not specifically categorized under Main Chapters, Factions, Quest Dialog, '
        'AI Dialog, NarrationDialog, or Other.'
    )
    worksheet[f"A{note_text_row_4}"] = note_text_4
    worksheet[f"A{note_text_row_4}"].font = Font(size=9, italic=True, color="666666")
    worksheet[f"A{note_text_row_4}"].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.merge_cells(f"A{note_text_row_4}:G{note_text_row_4}")
    worksheet.row_dimensions[note_text_row_4].height = 30

    # Add migration details table if provided
    if migration_details:
        migration_start_row = note_text_row_4 + 3
        write_super_group_migration_details(worksheet, migration_details, migration_start_row)

    # Freeze header row
    worksheet.freeze_panes = worksheet['A2']


def write_super_group_migration_details(worksheet, migration_details, start_row):
    """
    Write detailed super group migration table below main analysis.

    Args:
        worksheet: Excel worksheet object
        migration_details: List of tuples (source_group, dest_group, word_count)
        start_row: Row number to start writing (below main table)
    """
    if not migration_details:
        # No migrations detected
        return

    # Aggregate migrations by source → destination pairs
    migration_summary = {}
    for source, dest, words in migration_details:
        key = (source, dest)
        migration_summary[key] = migration_summary.get(key, 0) + words

    # Create migration table
    # Header row
    worksheet[f"A{start_row}"] = "Super Group Migrations"
    worksheet[f"A{start_row}"].font = Font(bold=True, size=12)

    # Column headers
    header_row = start_row + 2
    worksheet[f"A{header_row}"] = "Source Group"
    worksheet[f"B{header_row}"] = "Destination Group"
    worksheet[f"C{header_row}"] = "Words Migrated"

    # Apply header formatting
    from openpyxl.styles import Border, Side
    for col in ['A', 'B', 'C']:
        cell = worksheet[f"{col}{header_row}"]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    data_row = header_row + 1
    for (source, dest), word_count in sorted(migration_summary.items(),
                                               key=lambda x: x[1], reverse=True):
        worksheet[f"A{data_row}"] = source
        worksheet[f"B{data_row}"] = dest
        worksheet[f"C{data_row}"] = word_count
        worksheet[f"C{data_row}"].number_format = '#,##0'
        worksheet[f"C{data_row}"].alignment = Alignment(horizontal="right")
        data_row += 1

    # Set column widths
    worksheet.column_dimensions['A'].width = 22
    worksheet.column_dimensions['B'].width = 22
    worksheet.column_dimensions['C'].width = 16

    # Add total row
    total_row = data_row + 1
    worksheet[f"A{total_row}"] = "TOTAL MIGRATIONS"
    worksheet[f"A{total_row}"].font = Font(bold=True)
    worksheet[f"C{total_row}"] = f"=SUM(C{header_row+1}:C{data_row-1})"
    worksheet[f"C{total_row}"].font = Font(bold=True)
    worksheet[f"C{total_row}"].number_format = '#,##0'
    worksheet[f"C{total_row}"].alignment = Alignment(horizontal="right")

    # Add border to total row
    thin_border = Border(top=Side(style='thin'))
    for col in ['A', 'B', 'C']:
        worksheet[f"{col}{total_row}"].border = thin_border
