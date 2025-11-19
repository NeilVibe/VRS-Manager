"""
Excel reading and normalization module.

This module provides functionality to safely read Excel files and
normalize dataframe content, particularly status columns.
"""

import pandas as pd
from openpyxl import load_workbook

from src.config import AFTER_RECORDING_STATUSES
from src.utils.helpers import safe_str


def find_status_column(columns):
    """
    Find the STATUS column in a dataframe's columns.

    Args:
        columns: List of column names to search

    Returns:
        str: The column name if found, None otherwise
    """
    for col in columns:
        if col.upper() == "STATUS":
            return col
    return None


def normalize_status(status_value):
    """
    Normalize a status value to uppercase.

    Args:
        status_value: The status value to normalize

    Returns:
        str: The normalized status value in uppercase, or empty string if invalid
    """
    clean_val = safe_str(status_value)
    if not clean_val:
        return ""
    return clean_val.upper()


def is_after_recording_status(status_value):
    """
    Check if a status value indicates post-recording state.

    Args:
        status_value: The status value to check

    Returns:
        bool: True if status is after recording, False otherwise
    """
    if not status_value:
        return False
    normalized = normalize_status(status_value)
    return normalized in AFTER_RECORDING_STATUSES


def clean_numeric_columns(df):
    """
    Clean numeric columns by removing trailing zeros and decimal points.

    Args:
        df: DataFrame to clean

    Returns:
        DataFrame: Cleaned dataframe
    """
    from src.config import COL_STARTFRAME, COL_ENDFRAME

    numeric_cols = [COL_STARTFRAME, COL_ENDFRAME]

    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: safe_str(x).rstrip('0').rstrip('.') if '.' in safe_str(x) else safe_str(x))

    return df


def clean_dataframe_none_values(df):
    """
    Clean dataframe by replacing None/NaN/empty values with empty strings.

    Args:
        df: DataFrame to clean

    Returns:
        DataFrame: Cleaned dataframe
    """
    for col in df.columns:
        df[col] = df[col].apply(lambda x: "" if safe_str(x).upper() in ["NONE", "NAN", ""] else safe_str(x))
    return df


def safe_read_excel(filepath, header=0, dtype=str):
    """
    Safely read an Excel file, converting all values to strings and cleaning data.

    This function uses openpyxl to read Excel files with data_only=True to get
    formula values, then converts to pandas DataFrame with all string types.
    It also cleans numeric columns and None values.

    Args:
        filepath: Path to the Excel file
        header: Row number to use as header (default: 0)
        dtype: Data type for columns (default: str)

    Returns:
        DataFrame: Cleaned pandas DataFrame

    Raises:
        ValueError: If the Excel file is empty
    """
    wb = load_workbook(filepath, data_only=True, read_only=False)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    wb.close()

    if len(data) > 0:
        df = pd.DataFrame(data[1:], columns=data[0])
        for col in df.columns:
            df[col] = df[col].apply(safe_str)
        df = clean_numeric_columns(df)
        df = clean_dataframe_none_values(df)
        return df
    else:
        raise ValueError("Excel file is empty")


def normalize_dataframe_status(df):
    """
    Normalize the STATUS column in a dataframe.

    Finds the STATUS column (case-insensitive), normalizes all values
    to uppercase, and renames the column to "STATUS" if needed.

    Args:
        df: DataFrame to normalize

    Returns:
        DataFrame: DataFrame with normalized STATUS column
    """
    status_col = find_status_column(df.columns)
    if status_col:
        df[status_col] = df[status_col].apply(normalize_status)
        if status_col != "STATUS":
            df.rename(columns={status_col: "STATUS"}, inplace=True)
    return df
