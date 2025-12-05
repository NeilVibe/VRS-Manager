"""
Master File Update processor.

This processor updates a Master File by importing data from a Working Process output,
using simplified EventName matching with high/low importance logic and TimeFrame preservation.

Version: v12051348
- Enhanced Super Group Word Analysis with reorganized columns and translation tracking
- TimeFrame (StartFrame) preservation when StrOrigin unchanged
"""

import os
import pandas as pd
from datetime import datetime

# Conditional tkinter import for headless testing
if os.environ.get('HEADLESS', '').lower() not in ('1', 'true', 'yes'):
    from tkinter import messagebox
else:
    messagebox = None

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

from src.processors.base_processor import BaseProcessor
from src.io.excel_reader import safe_read_excel
from src.io.formatters import apply_direct_coloring, widen_summary_columns, format_update_history_sheet
from src.utils.data_processing import normalize_dataframe_status, remove_full_duplicates
from src.utils.helpers import log, get_script_dir, safe_str
from src.config import (
    COL_SEQUENCE, COL_EVENTNAME, COL_STRORIGIN, COL_CASTINGKEY,
    COL_IMPORTANCE, COL_STARTFRAME, COL_ENDFRAME,
    COL_CHARACTERKEY, COL_DIALOGVOICE, COL_SPEAKER_GROUPKEY
)
from src.core.casting import generate_casting_key
from src.io.summary import create_master_file_update_history_sheet
from src.history.history_manager import add_master_file_update_record


class MasterProcessor(BaseProcessor):
    """
    Processor for Master File Update.

    Updates a master file with data from a working process output,
    using simplified EventName matching with high/low importance logic.
    """

    def __init__(self):
        """Initialize the master processor."""
        super().__init__()
        self.df_source = None
        self.df_target = None
        self.df_high = None
        self.df_low = None
        self.df_high_output = None
        self.df_low_output = None
        self.target_wb = None
        self.output_structure = []
        self.high_counter = {}
        self.low_counter = {}
        self.total_counter = {}
        self.marked_high_indices = set()
        self.marked_low_indices = set()

    def get_process_name(self):
        """Get the process name."""
        return "PROCESS MASTER FILE UPDATE"

    def select_files(self):
        """Prompt user to select source and target files."""
        self.source_file = self._select_single_file(
            "MASTER FILE UPDATE: Select SOURCE (Working Process output)"
        )
        if not self.source_file:
            return False

        self.target_file = self._select_single_file(
            "MASTER FILE UPDATE: Select TARGET (Master File to update)"
        )
        if not self.target_file:
            return False

        return True

    def read_files(self):
        """Read and normalize the source and target files."""
        try:
            log(f"Reading SOURCE: {os.path.basename(self.source_file)}")
            self.df_source = safe_read_excel(self.source_file, header=0, dtype=str)
            log(f"  → {len(self.df_source):,} rows, {self.df_source.shape[1]} columns")

            log(f"Reading TARGET: {os.path.basename(self.target_file)}")
            self.df_target = safe_read_excel(self.target_file, header=0, dtype=str)
            log(f"  → {len(self.df_target):,} rows, {self.df_target.shape[1]} columns")

            # Identify target structure
            target_columns = list(self.df_target.columns)
            log(f"\nTARGET structure: {len(target_columns)} columns")

            # Define output structure (TARGET + CHANGES + Importance)
            self.output_structure = target_columns.copy()
            if "CHANGES" not in self.output_structure:
                self.output_structure.append("CHANGES")
            if COL_IMPORTANCE not in self.output_structure:
                self.output_structure.append(COL_IMPORTANCE)

            log(f"Output structure: {len(self.output_structure)} columns (all sheets will use this)")

            log("\nNormalizing STATUS columns...")
            self.df_source = normalize_dataframe_status(self.df_source)
            self.df_target = normalize_dataframe_status(self.df_target)

            log("Removing full duplicate rows...")
            self.df_source = remove_full_duplicates(self.df_source, "SOURCE")
            self.df_target = remove_full_duplicates(self.df_target, "TARGET")

            if COL_IMPORTANCE not in self.df_source.columns:
                log("Warning: SOURCE has no 'Importance' column - treating all rows as High")
                self.df_source[COL_IMPORTANCE] = "High"

            log("Generating CastingKey column for SOURCE...")
            casting_keys_source = []
            for idx, row in self.df_source.iterrows():
                casting_key = generate_casting_key(
                    row.get(COL_CHARACTERKEY, ""),
                    row.get(COL_DIALOGVOICE, ""),
                    row.get(COL_SPEAKER_GROUPKEY, ""),
                    row.get("DialogType", "")
                )
                casting_keys_source.append(casting_key)
            self.df_source[COL_CASTINGKEY] = casting_keys_source
            log(f"  → Generated CastingKey for {len(casting_keys_source):,} source rows")

            log("Generating CastingKey column for TARGET...")
            casting_keys_target = []
            for idx, row in self.df_target.iterrows():
                casting_key = generate_casting_key(
                    row.get(COL_CHARACTERKEY, ""),
                    row.get(COL_DIALOGVOICE, ""),
                    row.get(COL_SPEAKER_GROUPKEY, ""),
                    row.get("DialogType", "")
                )
                casting_keys_target.append(casting_key)
            self.df_target[COL_CASTINGKEY] = casting_keys_target
            log(f"  → Generated CastingKey for {len(casting_keys_target):,} target rows")

            return True

        except Exception as e:
            log(f"Error reading files: {e}")
            import traceback
            traceback.print_exc()
            return False

    def process_data(self):
        """Process the data using simplified EventName matching."""
        try:
            # Separate high and low importance from SOURCE
            log("\nSeparating SOURCE by Importance...")
            self.df_high = self.df_source[self.df_source[COL_IMPORTANCE].str.lower() == "high"].copy()
            self.df_low = self.df_source[self.df_source[COL_IMPORTANCE].str.lower() != "high"].copy()
            log(f"  → High importance: {len(self.df_high):,} rows")
            log(f"  → Low importance: {len(self.df_low):,} rows (will be skipped)")

            # Build simple EventName lookups
            log("\nBuilding EventName lookups...")
            source_high_lookup = {}
            for idx, row in self.df_high.iterrows():
                event_name = safe_str(row.get(COL_EVENTNAME, ""))
                source_high_lookup[event_name] = row

            target_lookup = {}
            for idx, row in self.df_target.iterrows():
                event_name = safe_str(row.get(COL_EVENTNAME, ""))
                target_lookup[event_name] = row

            log(f"  → SOURCE HIGH: {len(source_high_lookup):,} EventNames")
            log(f"  → TARGET: {len(target_lookup):,} EventNames")

            # Process HIGH importance rows
            log("\nProcessing HIGH importance rows...")
            self.df_high_output, self.high_counter = self._process_high_importance(
                self.df_high, target_lookup
            )
            log(f"  → Processed {len(self.df_high_output):,} HIGH rows")

            # Skip LOW importance entirely
            log("\nSkipping LOW importance rows (no processing)...")
            self.df_low_output = pd.DataFrame()  # Empty
            self.low_counter = {"Skipped (LOW importance)": len(self.df_low)}

            # Find deleted rows
            log("\nMarking deleted rows...")
            deleted_rows, deleted_count = self._find_deleted_rows(
                target_lookup, source_high_lookup
            )
            log(f"  → Marked {deleted_count:,} deleted rows")

            # Combine HIGH + DELETED into single output
            if deleted_rows:
                self.df_high_output = pd.concat([self.df_high_output, pd.DataFrame(deleted_rows)], ignore_index=True)
                log(f"  → Total output rows: {len(self.df_high_output):,} (HIGH + DELETED)")

            # No separate deleted sheet needed
            self.df_deleted = pd.DataFrame()

            # Create summary
            log("\nCreating summary report...")
            self.total_counter = {}
            for k, v in self.high_counter.items():
                self.total_counter[k] = self.total_counter.get(k, 0) + v
            for k, v in self.low_counter.items():
                self.total_counter[k] = self.total_counter.get(k, 0) + v
            if deleted_count > 0:
                self.total_counter["Deleted Rows"] = deleted_count

            self.df_summary = self._create_summary()
            self.df_history = create_master_file_update_history_sheet()

            return True

        except Exception as e:
            log(f"Error processing data: {e}")
            import traceback
            traceback.print_exc()
            return False

    def write_output(self):
        """Write results to Excel file with formatting."""
        try:
            script_dir = get_script_dir()
            out_filename = "MasterFile_Updated_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
            self.output_path = os.path.join(script_dir, out_filename)
            log(f"\nWriting results to: {out_filename}")

            # Load target workbook for formatting
            log("Loading TARGET workbook for formatting...")
            wb_target = load_workbook(self.target_file)
            ws_target = wb_target.active

            with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
                self.df_high_output.to_excel(writer, sheet_name="Main Sheet", index=False)
                self.df_history.to_excel(writer, sheet_name="📅 Update History", index=False, header=False)
                self.df_summary.to_excel(writer, sheet_name="Summary Report", index=False, header=True)

                wb = writer.book

                # Apply formatting to main sheet
                ws = wb["Main Sheet"]
                log("Applying formatting to Main Sheet...")

                # Copy column widths
                for col_idx in range(1, min(ws_target.max_column + 1, ws.max_column + 1)):
                    col_letter = get_column_letter(col_idx)
                    if col_letter in ws_target.column_dimensions:
                        ws.column_dimensions[col_letter].width = ws_target.column_dimensions[col_letter].width

                # Copy header row formatting
                for col_idx in range(1, min(ws_target.max_column + 1, ws.max_column + 1)):
                    source_cell = ws_target.cell(row=1, column=col_idx)
                    target_cell = ws.cell(row=1, column=col_idx)

                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.alignment = copy(source_cell.alignment)

                # Apply CHANGES column coloring
                apply_direct_coloring(ws, is_master=False)

                format_update_history_sheet(wb["📅 Update History"])
                widen_summary_columns(wb["Summary Report"])

            wb_target.close()

            # Add to history
            add_master_file_update_record(
                out_filename, self.source_file, self.target_file,
                self.total_counter, len(self.df_high_output)
            )

            log(f"✓ File saved: {self.output_path}")
            return True

        except Exception as e:
            log(f"Error writing output: {e}")
            import traceback
            traceback.print_exc()
            return False

    def show_summary(self):
        """Display completion message with summary."""
        summary_msg = "Master File Update completed successfully!\n\n"
        summary_msg += f"Output file:\n{self.output_path}\n\n"
        summary_msg += "Results:\n"
        summary_msg += f"  Main Sheet: {len(self.df_high_output):,} rows (HIGH + DELETED)\n\n"
        summary_msg += "Change Summary:\n"
        for change_type, count in sorted(self.total_counter.items()):
            summary_msg += f"  {change_type}: {count:,}\n"

        messagebox.showinfo("MASTER FILE UPDATE Complete", summary_msg)

    # Helper methods
    def _build_lookups(self, df, label):
        """
        Build 10-key lookup dictionaries (TWO-PASS algorithm).

        Lookups store DataFrame INDEX only (not full rows).
        Use df.loc[index] to retrieve row when needed.
        """
        lookup_se = {}
        lookup_so = {}
        lookup_sc = {}
        lookup_eo = {}
        lookup_ec = {}
        lookup_oc = {}
        lookup_seo = {}
        lookup_sec = {}
        lookup_soc = {}
        lookup_eoc = {}

        for df_idx, row in df.iterrows():
            S = row[COL_SEQUENCE]
            E = row[COL_EVENTNAME]
            O = row[COL_STRORIGIN]
            C = row[COL_CASTINGKEY]

            # 2-key combinations
            key_se = (S, E)
            key_so = (S, O)
            key_sc = (S, C)
            key_eo = (E, O)
            key_ec = (E, C)
            key_oc = (O, C)
            # 3-key combinations
            key_seo = (S, E, O)
            key_sec = (S, E, C)
            key_soc = (S, O, C)
            key_eoc = (E, O, C)

            # Store DataFrame INDEX in each lookup (first occurrence only)
            if key_se not in lookup_se:
                lookup_se[key_se] = df_idx
            if key_so not in lookup_so:
                lookup_so[key_so] = df_idx
            if key_sc not in lookup_sc:
                lookup_sc[key_sc] = df_idx
            if key_eo not in lookup_eo:
                lookup_eo[key_eo] = df_idx
            if key_ec not in lookup_ec:
                lookup_ec[key_ec] = df_idx
            if key_oc not in lookup_oc:
                lookup_oc[key_oc] = df_idx
            if key_seo not in lookup_seo:
                lookup_seo[key_seo] = df_idx
            if key_sec not in lookup_sec:
                lookup_sec[key_sec] = df_idx
            if key_soc not in lookup_soc:
                lookup_soc[key_soc] = df_idx
            if key_eoc not in lookup_eoc:
                lookup_eoc[key_eoc] = df_idx

        log(f"  → {label} indexed: {len(lookup_se):,} rows")
        return (lookup_se, lookup_so, lookup_sc, lookup_eo, lookup_ec,
                lookup_oc, lookup_seo, lookup_sec, lookup_soc, lookup_eoc)

    def _process_high_importance(self, df_high, target_lookup):
        """
        Process high importance rows using simple EventName matching.

        Args:
            df_high: DataFrame of HIGH importance rows from SOURCE
            target_lookup: Dict mapping EventName -> row data from TARGET

        Returns: (df_output, counter)
        """
        output_rows = []
        counter = {}
        target_columns = set(self.df_target.columns)

        for _, source_row in df_high.iterrows():
            event_name = safe_str(source_row.get(COL_EVENTNAME, ""))

            if event_name in target_lookup:
                # UPDATE: EventName exists in TARGET
                target_row = target_lookup[event_name]
                output_row = target_row.to_dict()  # Start with TARGET data

                # Update only columns that exist in TARGET schema
                for col in target_columns:
                    if col in source_row.index and pd.notna(source_row[col]):
                        output_row[col] = source_row[col]  # Overwrite with SOURCE

                # TimeFrame preservation logic (v1118.3)
                # Only preserve StartFrame when timing changed but dialogue didn't
                source_startframe = safe_str(source_row.get(COL_STARTFRAME, ""))
                target_startframe = safe_str(target_row.get(COL_STARTFRAME, ""))
                source_strorigin = safe_str(source_row.get(COL_STRORIGIN, ""))
                target_strorigin = safe_str(target_row.get(COL_STRORIGIN, ""))

                startframe_changed = (source_startframe != target_startframe)
                strorigin_changed = (source_strorigin != target_strorigin)

                # Preserve TARGET StartFrame when only timing changed
                if startframe_changed and not strorigin_changed:
                    output_row[COL_STARTFRAME] = target_startframe

                # ALWAYS set CHANGES from SOURCE (regardless of TARGET schema)
                change_type = safe_str(source_row.get("CHANGES", ""))
                output_row["CHANGES"] = change_type
                counter[change_type] = counter.get(change_type, 0) + 1
            else:
                # ADD: New row (EventName not in TARGET)
                output_row = {}
                # Only use columns from TARGET schema
                for col in target_columns:
                    output_row[col] = source_row.get(col, "")

                # ALWAYS set CHANGES from SOURCE
                change_type = safe_str(source_row.get("CHANGES", "New Row"))
                output_row["CHANGES"] = change_type
                counter[change_type] = counter.get(change_type, 0) + 1

            # Ensure importance is set
            output_row[COL_IMPORTANCE] = "High"
            output_rows.append(output_row)

        df_output = pd.DataFrame(output_rows, columns=self.output_structure)
        return df_output, counter

    def _process_low_importance(self, df_low,
                                 t_lookup_se, t_lookup_so, t_lookup_sc, t_lookup_eo, t_lookup_ec,
                                 t_lookup_oc, t_lookup_seo, t_lookup_sec, t_lookup_soc, t_lookup_eoc):
        """
        Process low importance rows using TWO-PASS algorithm.

        Returns: (df_output, counter)
        """
        marked_target_indices = set()
        pass1_results = {}
        counter = {}

        # PASS 1: Detect No Change (perfect 4-key match) and New rows
        for src_idx, source_row in df_low.iterrows():
            S = source_row[COL_SEQUENCE]
            E = source_row[COL_EVENTNAME]
            O = source_row[COL_STRORIGIN]
            C = source_row[COL_CASTINGKEY]

            # Generate all 10 keys
            key_se = (S, E)
            key_so = (S, O)
            key_sc = (S, C)
            key_eo = (E, O)
            key_ec = (E, C)
            key_oc = (O, C)
            key_seo = (S, E, O)
            key_sec = (S, E, C)
            key_soc = (S, O, C)
            key_eoc = (E, O, C)

            # Check for perfect 4-key match (No Change)
            if key_sec in t_lookup_sec:
                target_idx = t_lookup_sec[key_sec]
                if target_idx not in marked_target_indices:
                    target_row = self.df_target.loc[target_idx]
                    target_S = target_row[COL_SEQUENCE]
                    target_E = target_row[COL_EVENTNAME]
                    target_O = target_row[COL_STRORIGIN]
                    target_C = target_row[COL_CASTINGKEY]

                    # Verify all 4 keys match
                    if S == target_S and E == target_E and O == target_O and C == target_C:
                        marked_target_indices.add(target_idx)
                        change_type = safe_str(source_row.get("CHANGES", "No Change"))
                        pass1_results[src_idx] = (change_type, target_idx)
                        counter[change_type] = counter.get(change_type, 0) + 1
                        continue

            # Check if NEW row (all 10 keys missing)
            if (key_se not in t_lookup_se) and \
               (key_so not in t_lookup_so) and \
               (key_sc not in t_lookup_sc) and \
               (key_eo not in t_lookup_eo) and \
               (key_ec not in t_lookup_ec) and \
               (key_oc not in t_lookup_oc) and \
               (key_seo not in t_lookup_seo) and \
               (key_sec not in t_lookup_sec) and \
               (key_soc not in t_lookup_soc) and \
               (key_eoc not in t_lookup_eoc):
                pass1_results[src_idx] = ("New Row", None)
                counter["New Row"] = counter.get("New Row", 0) + 1

        # PASS 2: Pattern matching using UNMARKED TARGET rows
        for src_idx, source_row in df_low.iterrows():
            if src_idx in pass1_results:
                continue  # Already processed in PASS 1

            S = source_row[COL_SEQUENCE]
            E = source_row[COL_EVENTNAME]
            O = source_row[COL_STRORIGIN]
            C = source_row[COL_CASTINGKEY]

            # Generate all 10 keys
            key_se = (S, E)
            key_so = (S, O)
            key_sc = (S, C)
            key_eo = (E, O)
            key_ec = (E, C)
            key_oc = (O, C)
            key_seo = (S, E, O)
            key_sec = (S, E, C)
            key_soc = (S, O, C)
            key_eoc = (E, O, C)

            change_type = None
            target_idx = None
            matched = False

            # LEVEL 1: 3-Key Matches (check if unmarked)
            if not matched and key_seo in t_lookup_seo:
                candidate_idx = t_lookup_seo[key_seo]
                if candidate_idx not in marked_target_indices:
                    change_type = "CastingKey Change"
                    target_idx = candidate_idx
                    marked_target_indices.add(target_idx)
                    matched = True

            if not matched and key_sec in t_lookup_sec:
                candidate_idx = t_lookup_sec[key_sec]
                if candidate_idx not in marked_target_indices:
                    target_row = self.df_target.loc[candidate_idx]
                    if O != target_row[COL_STRORIGIN]:
                        change_type = "StrOrigin Change"
                    else:
                        change_type = safe_str(source_row.get("CHANGES", "No Change"))
                    target_idx = candidate_idx
                    marked_target_indices.add(target_idx)
                    matched = True

            if not matched and key_soc in t_lookup_soc:
                candidate_idx = t_lookup_soc[key_soc]
                if candidate_idx not in marked_target_indices:
                    change_type = "EventName Change"
                    target_idx = candidate_idx
                    marked_target_indices.add(target_idx)
                    matched = True

            if not matched and key_eoc in t_lookup_eoc:
                candidate_idx = t_lookup_eoc[key_eoc]
                if candidate_idx not in marked_target_indices:
                    change_type = "SequenceName Change"
                    target_idx = candidate_idx
                    marked_target_indices.add(target_idx)
                    matched = True

            # LEVEL 2: 2-Key Matches (check if unmarked)
            if not matched and key_se in t_lookup_se:
                candidate_idx = t_lookup_se[key_se]
                if candidate_idx not in marked_target_indices:
                    target_row = self.df_target.loc[candidate_idx]
                    changes = []
                    if O != target_row[COL_STRORIGIN]:
                        changes.append("StrOrigin")
                    if C != target_row[COL_CASTINGKEY]:
                        changes.append("CastingKey")

                    if changes:
                        change_type = "+".join(changes) + " Change"
                    else:
                        change_type = safe_str(source_row.get("CHANGES", "No Change"))
                    target_idx = candidate_idx
                    marked_target_indices.add(target_idx)
                    matched = True

            if not matched and key_oc in t_lookup_oc:
                candidate_idx = t_lookup_oc[key_oc]
                if candidate_idx not in marked_target_indices:
                    target_row = self.df_target.loc[candidate_idx]
                    changes = []
                    if S != target_row[COL_SEQUENCE]:
                        changes.append("SequenceName")
                    if E != target_row[COL_EVENTNAME]:
                        changes.append("EventName")
                    change_type = "+".join(changes) + " Change" if changes else "No Relevant Change"
                    target_idx = candidate_idx
                    marked_target_indices.add(target_idx)
                    matched = True

            if not matched and key_ec in t_lookup_ec:
                candidate_idx = t_lookup_ec[key_ec]
                if candidate_idx not in marked_target_indices:
                    target_row = self.df_target.loc[candidate_idx]
                    changes = []
                    if S != target_row[COL_SEQUENCE]:
                        changes.append("SequenceName")
                    if O != target_row[COL_STRORIGIN]:
                        changes.append("StrOrigin")
                    change_type = "+".join(changes) + " Change" if changes else "No Relevant Change"
                    target_idx = candidate_idx
                    marked_target_indices.add(target_idx)
                    matched = True

            if not matched and key_sc in t_lookup_sc:
                candidate_idx = t_lookup_sc[key_sc]
                if candidate_idx not in marked_target_indices:
                    change_type = "EventName+StrOrigin Change"
                    target_idx = candidate_idx
                    marked_target_indices.add(target_idx)
                    matched = True

            if not matched and key_so in t_lookup_so:
                candidate_idx = t_lookup_so[key_so]
                if candidate_idx not in marked_target_indices:
                    target_row = self.df_target.loc[candidate_idx]
                    old_eventname = target_row[COL_EVENTNAME]

                    # Check if (S, old_eventname) exists
                    if (S, old_eventname) in t_lookup_se:
                        se_idx = t_lookup_se[(S, old_eventname)]
                        if se_idx not in marked_target_indices:
                            target_row = self.df_target.loc[se_idx]
                            changes = []
                            if E != old_eventname:
                                changes.append("EventName")
                            if C != target_row[COL_CASTINGKEY]:
                                changes.append("CastingKey")
                            change_type = "+".join(changes) + " Change" if changes else "No Relevant Change"
                            target_idx = se_idx
                            marked_target_indices.add(target_idx)
                            matched = True
                    else:
                        change_type = "EventName+CastingKey Change"
                        target_idx = candidate_idx
                        marked_target_indices.add(target_idx)
                        matched = True

            if not matched and key_eo in t_lookup_eo:
                candidate_idx = t_lookup_eo[key_eo]
                if candidate_idx not in marked_target_indices:
                    change_type = "SequenceName Change"
                    target_idx = candidate_idx
                    marked_target_indices.add(target_idx)
                    matched = True

            if not matched:
                change_type = "ERROR: Logic Fault"
                target_idx = None

            pass1_results[src_idx] = (change_type, target_idx)
            counter[change_type] = counter.get(change_type, 0) + 1

        # Generate output rows
        low_rows = []
        for src_idx, source_row in df_low.iterrows():
            change_type, target_idx = pass1_results[src_idx]

            # LOW IMPORTANCE LOGIC: For existing rows, preserve TARGET data
            output_row = {}
            if change_type != "New Row" and target_idx is not None:
                # Existing row: Use TARGET data (preserve original)
                target_row = self.df_target.loc[target_idx]
                for col in self.output_structure:
                    if col in target_row.index:
                        output_row[col] = safe_str(target_row[col])
                    elif col in source_row.index:
                        output_row[col] = safe_str(source_row[col])
                    else:
                        output_row[col] = ""
            else:
                # New row: Use SOURCE data (will be deleted in post-process)
                for col in self.output_structure:
                    if col in source_row.index:
                        output_row[col] = safe_str(source_row[col])
                    else:
                        output_row[col] = ""

            # Always set CHANGES and Importance
            output_row["CHANGES"] = change_type
            output_row[COL_IMPORTANCE] = "Low"
            low_rows.append(output_row)

        df_output = pd.DataFrame(low_rows, columns=self.output_structure) if low_rows else pd.DataFrame(columns=self.output_structure)

        # Store marked indices for deleted row detection
        self.marked_low_indices = marked_target_indices

        return df_output, counter

    def _find_deleted_rows(self, target_lookup, source_high_lookup):
        """
        Find deleted rows using simple EventName matching.

        Args:
            target_lookup: Dict mapping EventName -> row data from TARGET
            source_high_lookup: Dict mapping EventName -> row data from SOURCE HIGH

        Returns: (deleted_rows list, deleted_count int)
        """
        deleted_rows = []
        target_columns = set(self.df_target.columns)

        for event_name, target_row in target_lookup.items():
            if event_name not in source_high_lookup:
                # DELETED: EventName not in SOURCE
                output_row = target_row.to_dict()
                output_row["CHANGES"] = "Deleted"
                deleted_rows.append(output_row)

        return deleted_rows, len(deleted_rows)

    def _create_summary(self):
        """Create summary DataFrame."""
        summary_rows = [
            ["MASTER FILE UPDATE SUMMARY", "", "", ""],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", ""],
            ["Source file", os.path.basename(self.source_file), "", ""],
            ["Target file", os.path.basename(self.target_file), "", ""],
            ["", "", "", ""],
            ["RESULT COUNTS", "Count", "", ""],
            ["Main Sheet", len(self.df_high_output), "", ""],
            ["", "", "", ""],
            ["CHANGE BREAKDOWN", "Count", "", ""],
        ]

        for change_type in sorted(self.total_counter.keys()):
            summary_rows.append([f"  {change_type}", self.total_counter[change_type], "", ""])

        return pd.DataFrame(summary_rows, columns=["Metric", "Value", "Col3", "Col4"])
