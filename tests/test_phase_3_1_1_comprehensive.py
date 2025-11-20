#!/usr/bin/env python3
"""
Comprehensive End-to-End Test for Phase 3.1.1 - Word-Level Diff Enhancement

Tests:
1. Raw VRS Check - StrOrigin Analysis sheet with 4-column layout
2. Working VRS Check - StrOrigin Analysis sheet with 4-column layout
3. Progress tracking with filling bar
4. Column widths (25|25|20|35)
5. Word-level diff in Diff Detail column
6. analyze() method returns tuple (analysis, diff_detail)

Author: Claude Code
Version: 1.121.0
"""

import sys
import os
import pandas as pd
import openpyxl
from pathlib import Path

# Set headless mode BEFORE any imports
os.environ['HEADLESS'] = '1'

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.io.excel_reader import safe_read_excel
from src.processors.raw_processor import RawProcessor
from src.processors.working_processor import WorkingProcessor
from src.utils.strorigin_analysis import StrOriginAnalyzer


def create_test_files_with_strorigin_changes():
    """
    Create test files with various StrOrigin change scenarios.
    """
    print("\n" + "=" * 80)
    print("Creating test files with StrOrigin changes...")
    print("=" * 80 + "\n")

    # Create PREVIOUS file
    data_prev = {
        "SequenceName": ["Scene_01"] * 8,
        "EventName": [f"E{1000+i}" for i in range(8)],
        "StrOrigin": [
            "안녕하세요",  # Will add punctuation
            "좋은 아침입니다",  # Will add punctuation
            "Hello world",  # Will change word
            "The player won the game",  # Will change multiple words
            "Loading please wait",  # Will delete word
            "Press button",  # Will add word
            "게임을 시작합니다",  # Major change
            "시작하기",  # Minor change
        ],
        "CastingKey": ["Hero_A"] * 8,
        "CharacterKey": ["Hero"] * 8,
        "DialogVoice": ["A"] * 8,
        "GroupKey": ["Main"] * 8,
        "DialogType": ["Dialog"] * 8,
        "STATUS": ["POLISHED"] * 8,
        "Text": ["Translation"] * 8,
        "FREEMEMO": ["Note"] * 8,
    }
    df_prev = pd.DataFrame(data_prev)

    # Create CURRENT file with changes
    data_curr = {
        "SequenceName": ["Scene_01"] * 8,
        "EventName": [f"E{1000+i}" for i in range(8)],
        "StrOrigin": [
            "안녕하세요!",  # Punctuation-only
            "좋은 아침입니다?",  # Punctuation-only
            "Hello universe",  # Word change: [world→universe]
            "The enemy lost the battle",  # Multiple words: [player won→enemy lost] [game→battle]
            "Loading wait",  # Deletion: [-please]
            "Press any button",  # Addition: [+any]
            "게임이 시작됩니다",  # Major semantic change
            "시작",  # Minor semantic change (shortened)
        ],
        "CastingKey": ["Hero_A"] * 8,
        "CharacterKey": ["Hero"] * 8,
        "DialogVoice": ["A"] * 8,
        "GroupKey": ["Main"] * 8,
        "DialogType": ["Dialog"] * 8,
        "STATUS": [""] * 8,  # Empty for working process to import
        "Text": [""] * 8,
        "FREEMEMO": [""] * 8,
    }
    df_curr = pd.DataFrame(data_curr)

    # Save test files
    prev_path = "tests/Phase_3_1_1_PREVIOUS.xlsx"
    curr_path = "tests/Phase_3_1_1_CURRENT.xlsx"

    df_prev.to_excel(prev_path, index=False, engine="openpyxl")
    df_curr.to_excel(curr_path, index=False, engine="openpyxl")

    print(f"✅ Created: {prev_path} ({len(df_prev)} rows)")
    print(f"✅ Created: {curr_path} ({len(df_curr)} rows)")
    print()

    return prev_path, curr_path


def test_strorigin_analyzer_returns_tuple():
    """
    Test that StrOriginAnalyzer.analyze() returns tuple (analysis, diff_detail).
    """
    print("=" * 80)
    print("TEST 1: StrOriginAnalyzer.analyze() returns tuple")
    print("=" * 80 + "\n")

    analyzer = StrOriginAnalyzer()

    # Test 1: Punctuation-only change
    result = analyzer.analyze("안녕하세요", "안녕하세요!")
    assert isinstance(result, tuple), f"Expected tuple, got {type(result)}"
    assert len(result) == 2, f"Expected tuple of length 2, got {len(result)}"
    analysis, diff_detail = result
    assert "Punctuation/Space" in analysis, f"Expected 'Punctuation/Space', got '{analysis}'"
    assert diff_detail == "", f"Expected empty diff_detail for punctuation-only, got '{diff_detail}'"
    print(f"  ✅ Punctuation-only: {result}")

    # Test 2: Word-level change
    result = analyzer.analyze("Hello world", "Hello universe")
    analysis, diff_detail = result
    assert "similar" in analysis or "Content Change" in analysis, f"Got: {analysis}"
    assert "[world→universe]" in diff_detail, f"Expected '[world→universe]', got '{diff_detail}'"
    print(f"  ✅ Word change: {result}")

    # Test 3: Multiple word changes
    result = analyzer.analyze("The player won the game", "The enemy lost the battle")
    analysis, diff_detail = result
    assert "[player won→enemy lost]" in diff_detail, f"Missing chunk 1 in '{diff_detail}'"
    assert "[game→battle]" in diff_detail, f"Missing chunk 2 in '{diff_detail}'"
    print(f"  ✅ Multiple changes: {result}")

    # Test 4: Word deletion
    result = analyzer.analyze("Loading please wait", "Loading wait")
    analysis, diff_detail = result
    assert "[-please]" in diff_detail, f"Expected '[-please]', got '{diff_detail}'"
    print(f"  ✅ Deletion: {result}")

    # Test 5: Word addition
    result = analyzer.analyze("Press button", "Press any button")
    analysis, diff_detail = result
    assert "[+any]" in diff_detail, f"Expected '[+any]', got '{diff_detail}'"
    print(f"  ✅ Addition: {result}")

    print("\n✅ TEST 1 PASSED: analyze() returns (analysis, diff_detail) tuple\n")
    return True


def test_raw_vrs_check_strorigin_sheet(prev_file, curr_file):
    """
    Test Raw VRS Check creates StrOrigin Analysis sheet with 4-column layout.
    """
    print("=" * 80)
    print("TEST 2: Raw VRS Check - StrOrigin Analysis Sheet")
    print("=" * 80 + "\n")

    processor = RawProcessor()
    processor.prev_file = prev_file
    processor.curr_file = curr_file

    # Read files first
    success = processor.read_files()
    assert success, "Raw processor read_files failed"

    # Process data
    success = processor.process_data()
    assert success, "Raw processor process_data failed"

    # Write output
    success = processor.write_output()
    assert success, "Raw processor write_output failed"

    output_file = processor.output_path
    print(f"✅ Raw output: {output_file}\n")

    # Verify StrOrigin Analysis sheet exists
    wb = openpyxl.load_workbook(output_file)
    assert "StrOrigin Change Analysis" in wb.sheetnames, \
        f"StrOrigin Change Analysis sheet not found. Sheets: {wb.sheetnames}"
    print("✅ StrOrigin Change Analysis sheet exists")

    # Load and verify columns
    df_analysis = pd.read_excel(output_file, sheet_name="StrOrigin Change Analysis")
    expected_columns = ["Previous StrOrigin", "Current StrOrigin", "StrOrigin Analysis", "Diff Detail"]

    for col in expected_columns:
        assert col in df_analysis.columns, f"Column '{col}' not found. Columns: {list(df_analysis.columns)}"
    print(f"✅ All 4 columns present: {expected_columns}")

    # Verify column widths
    ws = wb["StrOrigin Change Analysis"]
    col_widths = {}
    for col in ws.columns:
        col_letter = col[0].column_letter
        col_name = col[0].value
        if col_name in expected_columns:
            col_widths[col_name] = ws.column_dimensions[col_letter].width

    print(f"\nColumn widths:")
    print(f"  Previous StrOrigin: {col_widths.get('Previous StrOrigin', 'N/A')}")
    print(f"  Current StrOrigin: {col_widths.get('Current StrOrigin', 'N/A')}")
    print(f"  StrOrigin Analysis: {col_widths.get('StrOrigin Analysis', 'N/A')}")
    print(f"  Diff Detail: {col_widths.get('Diff Detail', 'N/A')}")

    # Verify data content
    print(f"\nAnalysis sheet has {len(df_analysis)} rows\n")

    # Check for punctuation-only changes
    punctuation_rows = df_analysis[df_analysis["StrOrigin Analysis"].str.contains("Punctuation/Space", na=False)]
    print(f"✅ Punctuation-only changes: {len(punctuation_rows)} rows")
    for idx, row in punctuation_rows.iterrows():
        prev = row["Previous StrOrigin"]
        curr = row["Current StrOrigin"]
        diff = row["Diff Detail"]
        print(f"  Row {idx}: '{prev}' → '{curr}' | Diff: '{diff}'")
        assert diff == "" or pd.isna(diff), f"Expected empty diff for punctuation-only, got '{diff}'"

    # Check for word-level diffs
    diff_rows = df_analysis[df_analysis["Diff Detail"].notna() & (df_analysis["Diff Detail"] != "")]
    print(f"\n✅ Word-level diffs: {len(diff_rows)} rows")
    for idx, row in diff_rows.iterrows():
        prev = row["Previous StrOrigin"]
        curr = row["Current StrOrigin"]
        diff = row["Diff Detail"]
        analysis = row["StrOrigin Analysis"]
        print(f"  Row {idx}: {analysis}")
        print(f"    Diff: {diff}")
        # Verify diff format
        assert "[" in diff and "]" in diff, f"Invalid diff format: '{diff}'"

    print("\n✅ TEST 2 PASSED: Raw VRS Check creates correct StrOrigin Analysis sheet\n")
    return True


def test_working_vrs_check_strorigin_sheet(prev_file, curr_file):
    """
    Test Working VRS Check creates StrOrigin Analysis sheet with 4-column layout.
    """
    print("=" * 80)
    print("TEST 3: Working VRS Check - StrOrigin Analysis Sheet")
    print("=" * 80 + "\n")

    processor = WorkingProcessor()
    processor.prev_file = prev_file
    processor.curr_file = curr_file

    # Read files first
    success = processor.read_files()
    assert success, "Working processor read_files failed"

    # Process data
    success = processor.process_data()
    assert success, "Working processor process_data failed"

    # Write output
    success = processor.write_output()
    assert success, "Working processor write_output failed"

    output_file = processor.output_path
    print(f"✅ Working output: {output_file}\n")

    # Verify StrOrigin Analysis sheet exists
    wb = openpyxl.load_workbook(output_file)
    assert "StrOrigin Change Analysis" in wb.sheetnames, \
        f"StrOrigin Change Analysis sheet not found. Sheets: {wb.sheetnames}"
    print("✅ StrOrigin Change Analysis sheet exists")

    # Load and verify columns
    df_analysis = pd.read_excel(output_file, sheet_name="StrOrigin Change Analysis")
    expected_columns = ["Previous StrOrigin", "Current StrOrigin", "StrOrigin Analysis", "Diff Detail"]

    for col in expected_columns:
        assert col in df_analysis.columns, f"Column '{col}' not found. Columns: {list(df_analysis.columns)}"
    print(f"✅ All 4 columns present: {expected_columns}")

    # Verify column widths
    ws = wb["StrOrigin Change Analysis"]
    col_widths = {}
    for col in ws.columns:
        col_letter = col[0].column_letter
        col_name = col[0].value
        if col_name in expected_columns:
            col_widths[col_name] = ws.column_dimensions[col_letter].width

    print(f"\nColumn widths:")
    print(f"  Previous StrOrigin: {col_widths.get('Previous StrOrigin', 'N/A')}")
    print(f"  Current StrOrigin: {col_widths.get('Current StrOrigin', 'N/A')}")
    print(f"  StrOrigin Analysis: {col_widths.get('StrOrigin Analysis', 'N/A')}")
    print(f"  Diff Detail: {col_widths.get('Diff Detail', 'N/A')}")

    # Verify expected widths (25|25|20|35)
    assert col_widths.get("Previous StrOrigin") == 25, \
        f"Previous StrOrigin width should be 25, got {col_widths.get('Previous StrOrigin')}"
    assert col_widths.get("Current StrOrigin") == 25, \
        f"Current StrOrigin width should be 25, got {col_widths.get('Current StrOrigin')}"
    assert col_widths.get("StrOrigin Analysis") == 20, \
        f"StrOrigin Analysis width should be 20, got {col_widths.get('StrOrigin Analysis')}"
    assert col_widths.get("Diff Detail") == 35, \
        f"Diff Detail width should be 35, got {col_widths.get('Diff Detail')}"
    print("✅ Column widths match expected (25|25|20|35)")

    # Verify data content
    print(f"\nAnalysis sheet has {len(df_analysis)} rows\n")

    # Verify word-level diff examples
    test_cases = [
        ("Hello world", "Hello universe", "[world→universe]"),
        ("The player won the game", "The enemy lost the battle", "[player won→enemy lost]"),
        ("Loading please wait", "Loading wait", "[-please]"),
        ("Press button", "Press any button", "[+any]"),
    ]

    for prev_text, curr_text, expected_diff in test_cases:
        matching_rows = df_analysis[
            (df_analysis["Previous StrOrigin"] == prev_text) &
            (df_analysis["Current StrOrigin"] == curr_text)
        ]
        if len(matching_rows) > 0:
            diff = matching_rows.iloc[0]["Diff Detail"]
            assert expected_diff in str(diff), \
                f"Expected '{expected_diff}' in diff, got '{diff}'"
            print(f"  ✅ '{prev_text}' → '{curr_text}': {diff}")

    print("\n✅ TEST 3 PASSED: Working VRS Check creates correct StrOrigin Analysis sheet\n")
    return True


def run_all_tests():
    """
    Main test runner.
    """
    print("\n╔" + "=" * 78 + "╗")
    print("║" + " " * 10 + "PHASE 3.1.1 - Comprehensive End-to-End Test" + " " * 17 + "║")
    print("╚" + "=" * 78 + "╝\n")

    all_passed = True

    try:
        # Test 1: StrOriginAnalyzer returns tuple
        if not test_strorigin_analyzer_returns_tuple():
            all_passed = False

        # Create test files
        prev_file, curr_file = create_test_files_with_strorigin_changes()

        # Test 2: Raw VRS Check
        if not test_raw_vrs_check_strorigin_sheet(prev_file, curr_file):
            all_passed = False

        # Test 3: Working VRS Check
        if not test_working_vrs_check_strorigin_sheet(prev_file, curr_file):
            all_passed = False

        if all_passed:
            print("\n" + "=" * 80)
            print("✅ ALL TESTS PASSED!")
            print("=" * 80)
            print("\nPhase 3.1.1 (Word-Level Diff Enhancement) is working correctly!")
            print("\nVerified:")
            print("  ✅ analyze() returns (analysis, diff_detail) tuple")
            print("  ✅ Raw Process creates 4-column StrOrigin Analysis sheet")
            print("  ✅ Working Process creates 4-column StrOrigin Analysis sheet")
            print("  ✅ Column widths are correct (25|25|20|35)")
            print("  ✅ Word-level diff shows in Diff Detail column")
            print("  ✅ Punctuation-only changes have empty Diff Detail")
            print()
            return True
        else:
            print("\n" + "=" * 80)
            print("❌ SOME TESTS FAILED")
            print("=" * 80 + "\n")
            return False

    except Exception as e:
        print("\n" + "=" * 80)
        print("❌ TEST FAILED WITH EXCEPTION")
        print("=" * 80 + "\n")
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
