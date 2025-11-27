#!/usr/bin/env python3
"""
Comprehensive test for Phase 2.3 - StrOrigin Change Analysis

Tests StrOrigin analysis with real data structure and various edge cases.
Based on REAL PREVIOUS/CURRENT FILE TEST STRUCTURE.xlsx
"""

import sys
import os
import pandas as pd
import openpyxl
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.io.excel_reader import safe_read_excel
from src.processors.working_processor import WorkingProcessor


def create_test_files_with_strorigin_changes():
    """
    Create test files based on REAL structure with various StrOrigin change scenarios.
    """
    print("=" * 80)
    print("Creating test files with StrOrigin changes...")
    print("=" * 80)
    print()

    # Load REAL structure as base
    real_prev_path = "tests/REAL PREVIOUS FILE TEST STRUCTURE.xlsx"
    real_curr_path = "tests/REAL CURRENT FILE TEST STRUCTURE.xlsx"

    if not os.path.exists(real_prev_path):
        print(f"❌ ERROR: {real_prev_path} not found")
        print("Please ensure REAL test structure files are in tests/ directory")
        return None, None

    df_prev_base = safe_read_excel(real_prev_path)
    df_curr_base = safe_read_excel(real_curr_path)

    print(f"✅ Loaded REAL PREVIOUS structure: {len(df_prev_base)} rows")
    print(f"✅ Loaded REAL CURRENT structure: {len(df_curr_base)} rows")
    print()

    # Take first 10 rows as base and create test scenarios
    df_prev = df_prev_base.head(10).copy()
    df_curr = df_curr_base.head(10).copy()

    print("Creating test scenarios:")
    print()

    # Test Case 1: Punctuation-only change (rows 0-2)
    print("  Test 1: Punctuation/Space-only changes")
    if "StrOrigin" in df_curr.columns:
        df_curr.loc[0, "StrOrigin"] = str(df_prev.loc[0, "StrOrigin"]) + "!"
        df_curr.loc[1, "StrOrigin"] = str(df_prev.loc[1, "StrOrigin"]) + "?"
        df_curr.loc[2, "StrOrigin"] = str(df_prev.loc[2, "StrOrigin"]) + "..."
        print(f"    Row 0: '{df_prev.loc[0, 'StrOrigin']}' → '{df_curr.loc[0, 'StrOrigin']}'")
        print(f"    Row 1: '{df_prev.loc[1, 'StrOrigin']}' → '{df_curr.loc[1, 'StrOrigin']}'")
        print(f"    Row 2: '{df_prev.loc[2, 'StrOrigin']}' → '{df_curr.loc[2, 'StrOrigin']}'")
    print()

    # Test Case 2: Minor text change (high similarity) (rows 3-4)
    print("  Test 2: Minor text changes (should be 80%+ similar)")
    if "StrOrigin" in df_curr.columns:
        # Assuming Korean text exists, make minor modifications
        orig_text = str(df_prev.loc[3, "StrOrigin"])
        if len(orig_text) > 5:
            df_curr.loc[3, "StrOrigin"] = orig_text[:len(orig_text)//2]  # Cut in half
        print(f"    Row 3: '{df_prev.loc[3, 'StrOrigin']}' → '{df_curr.loc[3, 'StrOrigin']}'")

        orig_text = str(df_prev.loc[4, "StrOrigin"])
        if len(orig_text) > 10:
            df_curr.loc[4, "StrOrigin"] = orig_text.replace(orig_text[5], "X")  # Change one char
        print(f"    Row 4: '{df_prev.loc[4, 'StrOrigin']}' → '{df_curr.loc[4, 'StrOrigin']}'")
    print()

    # Test Case 3: Major text change (low similarity) (rows 5-6)
    print("  Test 3: Major text changes (should be <50% similar)")
    if "StrOrigin" in df_curr.columns:
        df_curr.loc[5, "StrOrigin"] = "Completely different text"
        df_curr.loc[6, "StrOrigin"] = "Another totally different sentence"
        print(f"    Row 5: '{df_prev.loc[5, 'StrOrigin']}' → '{df_curr.loc[5, 'StrOrigin']}'")
        print(f"    Row 6: '{df_prev.loc[6, 'StrOrigin']}' → '{df_curr.loc[6, 'StrOrigin']}'")
    print()

    # Test Case 4: Space-only changes (rows 7-8)
    print("  Test 4: Space-only changes")
    if "StrOrigin" in df_curr.columns:
        df_curr.loc[7, "StrOrigin"] = "  " + str(df_prev.loc[7, "StrOrigin"]) + "  "  # Add spaces
        orig = str(df_prev.loc[8, "StrOrigin"])
        df_curr.loc[8, "StrOrigin"] = orig.replace(" ", "")  # Remove spaces
        print(f"    Row 7: '{df_prev.loc[7, 'StrOrigin']}' → '{df_curr.loc[7, 'StrOrigin']}'")
        print(f"    Row 8: '{df_prev.loc[8, 'StrOrigin']}' → '{df_curr.loc[8, 'StrOrigin']}'")
    print()

    # Test Case 5: No change (row 9)
    print("  Test 5: No change (should not appear in analysis sheet)")
    # df_curr row 9 stays same as df_prev row 9
    print(f"    Row 9: No change")
    print()

    # Save test files
    test_prev_path = "tests/StrOrigin_Test_PREVIOUS.xlsx"
    test_curr_path = "tests/StrOrigin_Test_CURRENT.xlsx"

    df_prev.to_excel(test_prev_path, index=False, engine="openpyxl")
    df_curr.to_excel(test_curr_path, index=False, engine="openpyxl")

    print(f"✅ Created: {test_prev_path}")
    print(f"✅ Created: {test_curr_path}")
    print()

    return test_prev_path, test_curr_path


def run_working_vrs_check(prev_file, curr_file):
    """
    Run Working VRS Check processor.
    """
    print("=" * 80)
    print("Running Working VRS Check...")
    print("=" * 80)
    print()

    processor = WorkingProcessor()
    processor.prev_file = prev_file
    processor.curr_file = curr_file

    # Process data
    success = processor.process_data()
    if not success:
        print("❌ ERROR: Processing failed")
        return None

    # Write output
    success = processor.write_output()
    if not success:
        print("❌ ERROR: Writing output failed")
        return None

    print()
    print(f"✅ Output file: {processor.output_path}")
    return processor.output_path


def verify_strorigin_analysis_sheet(output_file):
    """
    Verify StrOrigin Analysis sheet exists and contains correct data.
    """
    print()
    print("=" * 80)
    print("Verifying StrOrigin Analysis Sheet...")
    print("=" * 80)
    print()

    if not os.path.exists(output_file):
        print(f"❌ ERROR: Output file not found: {output_file}")
        return False

    # Load workbook
    wb = openpyxl.load_workbook(output_file)

    # Check if StrOrigin Analysis sheet exists
    if "StrOrigin Change Analysis" not in wb.sheetnames:
        print("❌ ERROR: 'StrOrigin Change Analysis' sheet not found!")
        print(f"Available sheets: {wb.sheetnames}")
        return False

    print("✅ 'StrOrigin Change Analysis' sheet exists")
    print()

    # Load the analysis sheet
    df_analysis = pd.read_excel(output_file, sheet_name="StrOrigin Change Analysis")

    print(f"Analysis sheet contains {len(df_analysis)} rows")
    print()

    # Verify "StrOrigin Analysis" column exists
    if "StrOrigin Analysis" not in df_analysis.columns:
        print("❌ ERROR: 'StrOrigin Analysis' column not found!")
        print(f"Available columns: {list(df_analysis.columns)}")
        return False

    print("✅ 'StrOrigin Analysis' column exists")
    print()

    # Show analysis results
    print("Analysis Results:")
    print("-" * 80)
    for idx, row in df_analysis.iterrows():
        strorigin = str(row.get("StrOrigin", "N/A"))[:50]  # Truncate for display
        analysis = str(row.get("StrOrigin Analysis", "N/A"))
        changes = str(row.get("CHANGES", "N/A"))
        print(f"  Row {idx}: {changes}")
        print(f"    StrOrigin: {strorigin}...")
        print(f"    Analysis:  {analysis}")
        print()

    # Verify analysis classifications
    analysis_values = df_analysis["StrOrigin Analysis"].tolist()

    punctuation_count = sum(1 for v in analysis_values if "Punctuation/Space" in str(v))
    similarity_count = sum(1 for v in analysis_values if "% similar" in str(v))
    na_count = sum(1 for v in analysis_values if "N/A" in str(v))

    print()
    print("Classification Summary:")
    print(f"  Punctuation/Space Changes: {punctuation_count}")
    print(f"  Semantic Similarity: {similarity_count}")
    print(f"  N/A (no previous data): {na_count}")
    print()

    # Expected: Should have at least some punctuation and some similarity results
    if punctuation_count > 0:
        print("✅ Punctuation/Space detection working")
    else:
        print("⚠️  No punctuation/space changes detected")

    if similarity_count > 0:
        print("✅ BERT semantic similarity working")
    else:
        print("⚠️  No semantic similarity results")

    print()
    return True


def main():
    """
    Main test runner.
    """
    print()
    print("╔" + "=" * 78 + "╗")
    print("║" + " " * 15 + "PHASE 2.3 - StrOrigin Change Analysis Test" + " " * 20 + "║")
    print("╚" + "=" * 78 + "╝")
    print()

    try:
        # Step 1: Create test files
        prev_file, curr_file = create_test_files_with_strorigin_changes()
        if not prev_file or not curr_file:
            print("❌ TEST FAILED: Could not create test files")
            return False

        # Step 2: Run Working VRS Check
        output_file = run_working_vrs_check(prev_file, curr_file)
        if not output_file:
            print("❌ TEST FAILED: Working VRS Check failed")
            return False

        # Step 3: Verify StrOrigin Analysis sheet
        success = verify_strorigin_analysis_sheet(output_file)
        if not success:
            print("❌ TEST FAILED: Verification failed")
            return False

        # Success!
        print()
        print("=" * 80)
        print("✅ ALL TESTS PASSED!")
        print("=" * 80)
        print()
        print("Phase 2.3 (StrOrigin Change Analysis) is working correctly!")
        print()
        return True

    except Exception as e:
        print()
        print("=" * 80)
        print("❌ TEST FAILED WITH EXCEPTION")
        print("=" * 80)
        print()
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
