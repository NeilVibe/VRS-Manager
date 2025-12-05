#!/usr/bin/env python3
"""
Generic Excel Process Guides Updater

This script updates both EN and KR Excel guides with new version information.
Edit the version and content sections below to add new updates.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ============================================================================
# CONFIGURATION - EDIT THIS SECTION FOR EACH UPDATE
# ============================================================================

VERSION = "12051348"
VERSION_TEXT_EN = "Version 12051348 (Smarter Change Detection + Enhanced Tracking)"
VERSION_TEXT_KR = "버전 12051348 (스마트 변경 감지 + 향상된 추적)"

# English content to add
EN_HEADER = "WHAT'S NEW IN v12051348?"
EN_CONTENT = [
    {
        "title": "🎯 Smarter Change Classification",
        "items": [
            "CHANGES column now shows the MOST IMPORTANT change when multiple fields change",
            "Priority order: StrOrigin → Desc → CastingKey → TimeFrame → Group → EventName → SequenceName → DialogType → CharacterGroup",
            "Example: If EventName AND StrOrigin both changed → Shows 'StrOrigin Change' (higher priority)",
            "Makes it easier to quickly identify what needs attention first",
        ]
    },
    {
        "title": "📋 New DETAILED_CHANGES Column",
        "items": [
            "Shows the FULL list of all changes when multiple fields changed",
            "Example: 'EventName+StrOrigin+Desc Change' - all 3 fields changed",
            "Located at the far right of the output for detailed review",
            "CHANGES = quick view, DETAILED_CHANGES = complete picture",
        ]
    },
    {
        "title": "🔄 New PreviousEventName Column",
        "items": [
            "When EventName changes, you can now see what the OLD EventName was",
            "Helps track row reorganization and event renaming",
            "Only populated when EventName actually changed",
            "Empty for New Rows and No Change rows",
        ]
    },
    {
        "title": "📝 New PreviousText Column",
        "items": [
            "Shows the previous Text/Translation for ALL matched rows",
            "Instantly see what the old translation was without searching",
            "Helpful for reviewing what needs re-translation",
            "Empty only for New Rows (no previous data exists)",
        ]
    },
    {
        "title": "🔧 Improved CastingKey Accuracy",
        "items": [
            "CastingKey comparison is now more reliable across files",
            "Consistent handling even when files have different column structures",
            "Warnings displayed when source data is incomplete",
        ]
    },
]

# Korean content to add
KR_HEADER = "v12051348의 새로운 기능"
KR_CONTENT = [
    {
        "title": "🎯 스마트 변경 분류",
        "items": [
            "CHANGES 컬럼이 이제 여러 필드가 변경되었을 때 가장 중요한 변경만 표시",
            "우선순위: StrOrigin → Desc → CastingKey → TimeFrame → Group → EventName → SequenceName → DialogType → CharacterGroup",
            "예시: EventName과 StrOrigin 모두 변경됨 → 'StrOrigin Change' 표시 (더 높은 우선순위)",
            "어떤 작업을 먼저 해야 하는지 빠르게 파악 가능",
        ]
    },
    {
        "title": "📋 새로운 DETAILED_CHANGES 컬럼",
        "items": [
            "여러 필드가 변경되었을 때 모든 변경사항의 전체 목록 표시",
            "예시: 'EventName+StrOrigin+Desc Change' - 3개 필드 모두 변경됨",
            "상세 검토를 위해 출력의 맨 오른쪽에 위치",
            "CHANGES = 빠른 보기, DETAILED_CHANGES = 전체 그림",
        ]
    },
    {
        "title": "🔄 새로운 PreviousEventName 컬럼",
        "items": [
            "EventName이 변경되면 이전 EventName이 무엇이었는지 확인 가능",
            "행 재구성 및 이벤트 이름 변경 추적에 도움",
            "EventName이 실제로 변경된 경우에만 채워짐",
            "New Row 및 No Change 행에서는 비어 있음",
        ]
    },
    {
        "title": "📝 새로운 PreviousText 컬럼",
        "items": [
            "모든 매칭된 행에 대해 이전 Text/번역을 표시",
            "검색 없이 즉시 이전 번역이 무엇이었는지 확인",
            "재번역이 필요한 내용 검토에 유용",
            "New Row에서만 비어 있음 (이전 데이터 없음)",
        ]
    },
    {
        "title": "🔧 CastingKey 정확도 개선",
        "items": [
            "파일 간 CastingKey 비교가 이제 더 안정적",
            "파일 컬럼 구조가 다를 때도 일관된 처리",
            "소스 데이터가 불완전할 때 경고 표시",
        ]
    },
]

# ============================================================================
# SCRIPT LOGIC - NO NEED TO EDIT BELOW THIS LINE
# ============================================================================

def widen_columns(wb):
    """Widen columns A, B, C for better text visibility"""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.column_dimensions['A'].width = 100  # Main content column - wider
        ws.column_dimensions['B'].width = 30   # Secondary column
        ws.column_dimensions['C'].width = 20   # Tertiary column
    return True


def add_content_to_sheet(ws, content_sections):
    """Add content sections to worksheet"""
    # Find last row with content
    last_row = ws.max_row
    while last_row > 0 and not any(ws[f'{col}{last_row}'].value for col in ['A', 'B', 'C']):
        last_row -= 1

    new_row = last_row + 1

    for section in content_sections:
        new_row += 1

        # Add section title
        ws[f'A{new_row}'] = section['title']
        ws[f'A{new_row}'].font = Font(bold=True, size=11)

        # Add section items
        for item in section['items']:
            new_row += 1
            ws[f'A{new_row}'] = f"  • {item}"

    return new_row


def fix_pass1_to_pass2(wb):
    """Fix 'after PASS 1' deleted rows references to 'after PASS 2'"""
    corrections_made = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original = cell.value

                    # Fix various patterns
                    fixed = original.replace("After PASS 1: Any unmarked", "After PASS 2: Any unmarked")
                    fixed = fixed.replace("after PASS 1", "after PASS 2")
                    fixed = fixed.replace("After all CURRENT processed: Unmarked", "After PASS 2: Unmarked")

                    if fixed != original:
                        cell.value = fixed
                        corrections_made += 1

    return corrections_made


def update_en_guide():
    """Update English Process Guide"""
    print("Updating docs/VRS_Manager_Process_Guide_EN.xlsx...")

    wb = openpyxl.load_workbook('docs/VRS_Manager_Process_Guide_EN.xlsx')

    # Update Overview sheet
    if 'Overview' in wb.sheetnames:
        ws = wb['Overview']

        # Update version (row 2, column A)
        ws['A2'] = VERSION_TEXT_EN
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Find last row with content
        last_row = ws.max_row
        while last_row > 0 and not any(ws[f'{col}{last_row}'].value for col in ['A', 'B', 'C']):
            last_row -= 1

        # Add new section header
        new_row = last_row + 3

        ws[f'A{new_row}'] = EN_HEADER
        ws[f'A{new_row}'].font = Font(bold=True, size=14)
        ws[f'A{new_row}'].alignment = Alignment(horizontal='left')

        new_row += 1

        # Add content sections
        for section in EN_CONTENT:
            new_row += 1

            # Add section title
            ws[f'A{new_row}'] = section['title']
            ws[f'A{new_row}'].font = Font(bold=True, size=11)

            # Add section items
            for item in section['items']:
                new_row += 1
                ws[f'A{new_row}'] = f"  • {item}"

    # Update Master Update sheet with TimeFrame logic
    if '4. Master Update' in wb.sheetnames:
        ws = wb['4. Master Update']

        # Find last row
        last_row = ws.max_row
        while last_row > 0 and not any(ws[f'{col}{last_row}'].value for col in ['A', 'B', 'C']):
            last_row -= 1

        new_row = last_row + 3

        ws[f'A{new_row}'] = 'TimeFrame Preservation Logic (v1117 - High Importance Only)'
        ws[f'A{new_row}'].font = Font(bold=True, size=13)
        ws[f'A{new_row}'].fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')

        new_row += 2
        ws[f'A{new_row}'] = 'Rule:'
        ws[f'A{new_row}'].font = Font(bold=True, size=11)

        new_row += 1
        ws[f'A{new_row}'] = '  • If TimeFrame changed AND StrOrigin changed → Update TimeFrame (use SOURCE)'

        new_row += 1
        ws[f'A{new_row}'] = '  • If TimeFrame changed BUT StrOrigin did NOT change → Preserve TimeFrame (keep TARGET)'

        new_row += 2
        ws[f'A{new_row}'] = 'This ensures TimeFrame updates only apply when accompanied by dialogue content changes.'
        ws[f'A{new_row}'].font = Font(italic=True)

    # Fix PASS 1 → PASS 2 for deleted rows
    corrections = fix_pass1_to_pass2(wb)
    if corrections > 0:
        print(f"  ✅ Fixed {corrections} 'PASS 1' → 'PASS 2' references for deleted rows")

    # Widen columns for better UX
    widen_columns(wb)
    print("  ✅ Widened columns A, B, C for better text visibility")

    wb.save('docs/VRS_Manager_Process_Guide_EN.xlsx')
    print("✅ English guide updated successfully!")


def update_kr_guide():
    """Update Korean Process Guide"""
    print("\nUpdating docs/VRS_Manager_Process_Guide_KR.xlsx...")

    wb = openpyxl.load_workbook('docs/VRS_Manager_Process_Guide_KR.xlsx')

    # Update Overview sheet (개요)
    if '개요' in wb.sheetnames:
        ws = wb['개요']

        # Update version (row 2, column A)
        ws['A2'] = VERSION_TEXT_KR
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Find last row with content
        last_row = ws.max_row
        while last_row > 0 and not any(ws[f'{col}{last_row}'].value for col in ['A', 'B', 'C']):
            last_row -= 1

        # Add new section header
        new_row = last_row + 3

        ws[f'A{new_row}'] = KR_HEADER
        ws[f'A{new_row}'].font = Font(bold=True, size=14)
        ws[f'A{new_row}'].alignment = Alignment(horizontal='left')

        new_row += 1

        # Add content sections
        for section in KR_CONTENT:
            new_row += 1

            # Add section title
            ws[f'A{new_row}'] = section['title']
            ws[f'A{new_row}'].font = Font(bold=True, size=11)

            # Add section items
            for item in section['items']:
                new_row += 1
                ws[f'A{new_row}'] = f"  • {item}"

    # Update Master Update sheet with TimeFrame logic
    if '4. Master Update' in wb.sheetnames:
        ws = wb['4. Master Update']

        # Find last row
        last_row = ws.max_row
        while last_row > 0 and not any(ws[f'{col}{last_row}'].value for col in ['A', 'B', 'C']):
            last_row -= 1

        new_row = last_row + 3

        ws[f'A{new_row}'] = 'TimeFrame 보존 로직 (v1117 - High Importance만 해당)'
        ws[f'A{new_row}'].font = Font(bold=True, size=13)
        ws[f'A{new_row}'].fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')

        new_row += 2
        ws[f'A{new_row}'] = '규칙:'
        ws[f'A{new_row}'].font = Font(bold=True, size=11)

        new_row += 1
        ws[f'A{new_row}'] = '  • TimeFrame 변경 AND StrOrigin 변경 → TimeFrame 업데이트 (SOURCE 사용)'

        new_row += 1
        ws[f'A{new_row}'] = '  • TimeFrame 변경 BUT StrOrigin 변경 안 됨 → TimeFrame 보존 (TARGET 유지)'

        new_row += 2
        ws[f'A{new_row}'] = 'TimeFrame 업데이트는 대사 내용 변경과 함께 발생할 때만 적용됩니다.'
        ws[f'A{new_row}'].font = Font(italic=True)

    # Fix PASS 1 → PASS 2 for deleted rows
    corrections = fix_pass1_to_pass2(wb)
    if corrections > 0:
        print(f"  ✅ Fixed {corrections} 'PASS 1' → 'PASS 2' references for deleted rows")

    # Widen columns for better UX
    widen_columns(wb)
    print("  ✅ Widened columns A, B, C for better text visibility")

    wb.save('docs/VRS_Manager_Process_Guide_KR.xlsx')
    print("✅ Korean guide updated successfully!")


if __name__ == '__main__':
    print(f"Updating Excel guides to version {VERSION}...")
    print("=" * 60)
    update_en_guide()
    update_kr_guide()
    print("\n🎉 All Excel guides updated successfully!")
    print("=" * 60)
    print("\nNext steps:")
    print("1. Open the Excel files to verify the changes")
    print("2. git add docs/VRS_Manager_Process_Guide_EN.xlsx docs/VRS_Manager_Process_Guide_KR.xlsx")
    print(f'3. git commit -m "Update Excel guides to v{VERSION}"')
    print("4. git push")
