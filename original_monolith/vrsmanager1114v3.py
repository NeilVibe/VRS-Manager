print("- VRS Manager - Version : 1114v3 - By Neil Schmitt -")
print("- VRS Manager - Version : 1114v3 - By Neil Schmitt -")
print("- VRS Manager - Version : 1114v3 - By Neil Schmitt -")
print("- VRS Manager - Version : 1114v3 - By Neil Schmitt -")
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
import json
import threading

# =========================================================================== #
# CONFIGURATION
# =========================================================================== #
COL_SEQUENCE = "SequenceName"
COL_EVENTNAME = "EventName"
COL_STRORIGIN = "StrOrigin"
COL_DESC = "Desc"
COL_STARTFRAME = "StartFrame"
COL_ENDFRAME = "EndFrame"
COL_TEXT = "Text"
COL_STATUS = "STATUS"
COL_FREEMEMO = "FREEMEMO"
COL_CHARACTERKEY = "CharacterKey"
COL_CHARACTERNAME = "CharacterName"
COL_DIALOGVOICE = "DialogVoice"
COL_SPEAKER_GROUPKEY = "Speaker|CharacterGroupKey"
COL_CASTINGKEY = "CastingKey"
COL_PREVIOUSDATA = "PreviousData"
COL_MAINLINE_TRANSLATION = "Mainline Translation"
COL_PREVIOUS_STRORIGIN = "Previous StrOrigin"
COL_IMPORTANCE = "Importance"

CHAR_GROUP_COLS = ["Tribe", "Age", "Gender", "Job", "Region"]

AFTER_RECORDING_STATUSES = {
    "RECORDED", "PREVIOUSLY RECORDED", "FINAL", "RE-RECORD", "RE-RECORDED",
    "RERECORD", "RERECORDED", "SHIPPED",
    "전달 완료", "녹음 완료", "재녹음 필요", "재녹음 완료",
    "已传达", "已录音", "需补录", "已补录",
}

PRE_RECORDING_STATUSES = {
    "", "POLISHED", "SPEC-OUT", "CHECK",
    "준비 중", "확인 필요",
    "准备中", "需要确认",
}

OUTPUT_COLUMNS = [
    "DialogType", "Group", "SequenceName", "CharacterName", "CharacterKey",
    "DialogVoice", "CastingKey", "StrOrigin", "STATUS", "Text", "Desc",
    "FREEMEMO", "SubTimelineName", "CHANGES", "EventName", "StartFrame",
    "EndFrame", "Tribe", "Age", "Gender", "Job", "Region", "UpdateTime",
    "PreviousData", "Mainline Translation"
]

OUTPUT_COLUMNS_RAW = [
    "DialogType", "Group", "SequenceName", "CharacterName", "CharacterKey",
    "DialogVoice", "CastingKey", "StrOrigin", "STATUS", "Text", "Desc",
    "FREEMEMO", "SubTimelineName", "CHANGES", "EventName", "StartFrame",
    "EndFrame", "Tribe", "Age", "Gender", "Job", "Region", "UpdateTime",
    "Previous StrOrigin"
]

OUTPUT_COLUMNS_MASTER = [
    "DialogType", "Group", "SequenceName",
    "CharacterName_KR", "CharacterName_EN", "CharacterName_CN",
    "CharacterKey", "DialogVoice", "CastingKey", "StrOrigin",
    "STATUS_KR", "STATUS_EN", "STATUS_CN",
    "Text_KR", "Text_EN", "Text_CN",
    "FREEMEMO_KR", "FREEMEMO_EN", "FREEMEMO_CN",
    "SubTimelineName", "CHANGES", "EventName", "StartFrame", "EndFrame",
    "Tribe", "Age", "Gender", "Job", "Region", "UpdateTime",
    "PreviousData_KR", "PreviousData_EN", "PreviousData_CN",
    "Mainline Translation_KR", "Mainline Translation_EN", "Mainline Translation_CN"
]

WORKING_HISTORY_FILE = "working_update_history.json"
MASTER_HISTORY_FILE = "master_update_history.json"
ALLLANG_HISTORY_FILE = "alllang_update_history.json"

# =========================================================================== #
# UPDATE HISTORY MANAGEMENT
# =========================================================================== #
def get_history_file_path(process_type="master"):
    script_dir = get_script_dir()
    if process_type == "working":
        return os.path.join(script_dir, WORKING_HISTORY_FILE)
    elif process_type == "alllang":
        return os.path.join(script_dir, ALLLANG_HISTORY_FILE)
    else:
        return os.path.join(script_dir, MASTER_HISTORY_FILE)

def load_update_history(process_type="master"):
    history_path = get_history_file_path(process_type)
    
    if not os.path.exists(history_path):
        return {"process_type": process_type, "updates": []}
    
    try:
        with open(history_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        log(f"Warning: Could not load {process_type} history file: {e}")
        return {"process_type": process_type, "updates": []}

def save_update_history(history, process_type="master"):
    history_path = get_history_file_path(process_type)
    
    try:
        with open(history_path, 'w', encoding='utf-8') as f:
            json.dump(history, f, indent=2, ensure_ascii=False)
        log(f"✓ Update history saved")
    except Exception as e:
        log(f"Warning: Could not save {process_type} history file: {e}")

def add_working_update_record(output_filename, prev_path, curr_path, counter, total_rows):
    history = load_update_history("working")
    
    record = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "process_type": "Working",
        "output_file": output_filename,
        "previous_file": os.path.basename(prev_path),
        "current_file": os.path.basename(curr_path),
        "statistics": {
            "total_rows": total_rows,
            **{k: v for k, v in counter.items()}
        }
    }
    
    history["updates"].append(record)
    save_update_history(history, "working")
    return record

def add_alllang_update_record(output_filename, prev_kr, prev_en, prev_cn, curr_kr, curr_en, curr_cn, 
                              counter, total_rows):
    history = load_update_history("alllang")
    
    record = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "process_type": "AllLanguage",
        "output_file": output_filename,
        "languages_updated": {
            "KR": prev_kr is not None,
            "EN": prev_en is not None,
            "CN": prev_cn is not None
        },
        "previous_files": {
            "KR": os.path.basename(prev_kr) if prev_kr else None,
            "EN": os.path.basename(prev_en) if prev_en else None,
            "CN": os.path.basename(prev_cn) if prev_cn else None
        },
        "current_files": {
            "KR": os.path.basename(curr_kr),
            "EN": os.path.basename(curr_en),
            "CN": os.path.basename(curr_cn)
        },
        "statistics": {
            "total_rows": total_rows,
            **{k: v for k, v in counter.items()}
        }
    }
    
    history["updates"].append(record)
    save_update_history(history, "alllang")
    return record

def add_master_file_update_record(output_filename, source_path, target_path, counter, total_rows):
    history = load_update_history("master")
    
    record = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "process_type": "MasterFileUpdate",
        "output_file": output_filename,
        "source_file": os.path.basename(source_path),
        "target_file": os.path.basename(target_path),
        "statistics": {
            "total_rows": total_rows,
            **{k: v for k, v in counter.items()}
        }
    }
    
    history["updates"].append(record)
    save_update_history(history, "master")
    return record

def clear_update_history(process_type="master"):
    process_name = process_type.upper()
    
    if messagebox.askyesno("Clear History", 
                           f"Are you sure you want to clear ALL {process_name} update history?\n\nThis cannot be undone."):
        history = {"process_type": process_type, "updates": []}
        save_update_history(history, process_type)
        messagebox.showinfo("Success", f"{process_name} update history cleared successfully!")
        return True
    return False

def delete_specific_update(index, process_type="master"):
    history = load_update_history(process_type)
    
    if 0 <= index < len(history["updates"]):
        deleted = history["updates"].pop(index)
        save_update_history(history, process_type)
        return True, deleted
    return False, None

# =========================================================================== #
# UPDATE HISTORY VIEWER GUI
# =========================================================================== #
def show_update_history_viewer():
    viewer = tk.Toplevel()
    viewer.title("Update History Viewer")
    viewer.geometry("900x700")
    
    control_frame = tk.Frame(viewer, bg="#f0f0f0", pady=10)
    control_frame.pack(fill=tk.X)
    
    tk.Label(control_frame, text="Select Process:", font=("Arial", 10, "bold"), bg="#f0f0f0").pack(side=tk.LEFT, padx=10)
    
    process_var = tk.StringVar(value="master")
    
    def refresh_history():
        process_type = process_var.get()
        history = load_update_history(process_type)
        
        text_widget.delete(1.0, tk.END)
        
        if not history["updates"]:
            text_widget.insert(tk.END, f"No {process_type.upper()} process updates recorded yet.\n\n")
            text_widget.insert(tk.END, f"Updates will appear here after you run the {process_type.upper()} process.")
            return
        
        process_name = "MASTER FILE UPDATE" if process_type == "master" else ("ALL LANGUAGE PROCESS" if process_type == "alllang" else "WORKING PROCESS")
        text_widget.insert(tk.END, "═" * 100 + "\n", "header")
        text_widget.insert(tk.END, f"   {process_name} - UPDATE HISTORY\n", "header")
        text_widget.insert(tk.END, "═" * 100 + "\n", "header")
        text_widget.insert(tk.END, f"\n   Total Updates: {len(history['updates'])}\n\n", "subheader")
        
        for idx, update in enumerate(reversed(history["updates"]), 1):
            actual_idx = len(history["updates"]) - idx
            
            text_widget.insert(tk.END, "─" * 100 + "\n", "divider")
            text_widget.insert(tk.END, f"UPDATE #{idx}", "update_num")
            text_widget.insert(tk.END, f" (Index: {actual_idx})\n", "small")
            text_widget.insert(tk.END, "─" * 100 + "\n", "divider")
            
            text_widget.insert(tk.END, f"📅 Timestamp: ", "label")
            text_widget.insert(tk.END, f"{update['timestamp']}\n", "value")
            
            text_widget.insert(tk.END, f"📄 Output File: ", "label")
            text_widget.insert(tk.END, f"{update['output_file']}\n\n", "value")
            
            if process_type == "alllang":
                text_widget.insert(tk.END, "🌐 LANGUAGES UPDATED:\n", "section")
                for lang in ["KR", "EN", "CN"]:
                    status = "✓ UPDATED" if update["languages_updated"][lang] else "○ Preserved"
                    color_tag = "updated" if update["languages_updated"][lang] else "preserved"
                    previous = update["previous_files"][lang] if update["previous_files"][lang] else "N/A"
                    text_widget.insert(tk.END, f"   {lang}: ", "label")
                    text_widget.insert(tk.END, f"{status}", color_tag)
                    text_widget.insert(tk.END, f" | Previous: {previous}\n", "small")
                
                text_widget.insert(tk.END, "\n📁 CURRENT FILES (Complete Base):\n", "section")
                for lang in ["KR", "EN", "CN"]:
                    text_widget.insert(tk.END, f"   {lang}: ", "label")
                    text_widget.insert(tk.END, f"{update['current_files'][lang]}\n", "value")
            
            elif process_type == "master":
                text_widget.insert(tk.END, "📁 FILES:\n", "section")
                text_widget.insert(tk.END, f"   Source: ", "label")
                text_widget.insert(tk.END, f"{update['source_file']}\n", "value")
                text_widget.insert(tk.END, f"   Target: ", "label")
                text_widget.insert(tk.END, f"{update['target_file']}\n", "value")
            
            else:
                text_widget.insert(tk.END, "📁 FILES:\n", "section")
                text_widget.insert(tk.END, f"   Previous: ", "label")
                text_widget.insert(tk.END, f"{update['previous_file']}\n", "value")
                text_widget.insert(tk.END, f"   Current: ", "label")
                text_widget.insert(tk.END, f"{update['current_file']}\n", "value")
            
            text_widget.insert(tk.END, "\n📊 STATISTICS:\n", "section")
            stats = update["statistics"]
            text_widget.insert(tk.END, f"   Total Rows: ", "label")
            text_widget.insert(tk.END, f"{stats['total_rows']:,}\n", "value")
            
            text_widget.insert(tk.END, "\n   Change Breakdown:\n", "subsection")
            for key, value in stats.items():
                if key != "total_rows":
                    text_widget.insert(tk.END, f"      • {key}: ", "change_label")
                    text_widget.insert(tk.END, f"{value:,}\n", "change_value")
            
            text_widget.insert(tk.END, "\n")
        
        text_widget.insert(tk.END, "═" * 100 + "\n", "header")
        text_widget.insert(tk.END, f"   End of History ({len(history['updates'])} total updates)\n", "footer")
        text_widget.insert(tk.END, "═" * 100 + "\n", "header")
    
    def on_process_change():
        refresh_history()
    
    tk.Radiobutton(control_frame, text="Master File Update", variable=process_var, value="master",
                   command=on_process_change, font=("Arial", 10), bg="#f0f0f0").pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(control_frame, text="All Language Process", variable=process_var, value="alllang",
                   command=on_process_change, font=("Arial", 10), bg="#f0f0f0").pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(control_frame, text="Working Process", variable=process_var, value="working",
                   command=on_process_change, font=("Arial", 10), bg="#f0f0f0").pack(side=tk.LEFT, padx=5)
    
    tk.Button(control_frame, text="🔄 Refresh", command=refresh_history,
              font=("Arial", 9), bg="#4CAF50", fg="white", padx=10).pack(side=tk.LEFT, padx=20)
    
    def clear_current_history():
        process_type = process_var.get()
        if clear_update_history(process_type):
            refresh_history()
    
    tk.Button(control_frame, text="🗑️ Clear History", command=clear_current_history,
              font=("Arial", 9), bg="#f44336", fg="white", padx=10).pack(side=tk.LEFT, padx=5)
    
    def delete_update():
        process_type = process_var.get()
        index_str = tk.simpledialog.askstring("Delete Update", 
                                                "Enter the index number of the update to delete:")
        if index_str:
            try:
                index = int(index_str)
                success, deleted = delete_specific_update(index, process_type)
                if success:
                    messagebox.showinfo("Success", f"Update #{index} deleted successfully!")
                    refresh_history()
                else:
                    messagebox.showerror("Error", f"Invalid index: {index}")
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid number")
    
    tk.Button(control_frame, text="🗑️ Delete Update", command=delete_update,
              font=("Arial", 9), bg="#FF9800", fg="white", padx=10).pack(side=tk.LEFT, padx=5)
    
    text_frame = tk.Frame(viewer)
    text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    scrollbar = tk.Scrollbar(text_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    text_widget = tk.Text(text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set,
                         font=("Courier New", 9), bg="#ffffff", fg="#000000")
    text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=text_widget.yview)
    
    text_widget.tag_config("header", foreground="#000080", font=("Courier New", 10, "bold"))
    text_widget.tag_config("subheader", foreground="#0066cc", font=("Courier New", 9, "bold"))
    text_widget.tag_config("update_num", foreground="#006600", font=("Courier New", 10, "bold"))
    text_widget.tag_config("divider", foreground="#999999")
    text_widget.tag_config("label", foreground="#000000", font=("Courier New", 9, "bold"))
    text_widget.tag_config("value", foreground="#0066cc")
    text_widget.tag_config("section", foreground="#cc6600", font=("Courier New", 9, "bold"))
    text_widget.tag_config("subsection", foreground="#666666", font=("Courier New", 9, "bold"))
    text_widget.tag_config("small", foreground="#666666", font=("Courier New", 8))
    text_widget.tag_config("updated", foreground="#006600", font=("Courier New", 9, "bold"))
    text_widget.tag_config("preserved", foreground="#999999", font=("Courier New", 9))
    text_widget.tag_config("change_label", foreground="#333333")
    text_widget.tag_config("change_value", foreground="#0066cc", font=("Courier New", 9, "bold"))
    text_widget.tag_config("footer", foreground="#666666", font=("Courier New", 9, "italic"))
    
    refresh_history()
    
    viewer.update_idletasks()
    width = viewer.winfo_width()
    height = viewer.winfo_height()
    x = (viewer.winfo_screenwidth() // 2) - (width // 2)
    y = (viewer.winfo_screenheight() // 2) - (height // 2)
    viewer.geometry(f'{width}x{height}+{x}+{y}')

# =========================================================================== #
# UTILITIES
# =========================================================================== #
def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def get_script_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

def print_progress(current, total, label="Progress"):
    pct = int((current / total) * 100)
    bar_length = 30
    filled = int(bar_length * current / total)
    bar = "█" * filled + "░" * (bar_length - filled)
    print(f"\r   {label}: {bar} {pct}%", end="", flush=True)

def finalize_progress():
    print()

def find_status_column(columns):
    for col in columns:
        if col.upper() == "STATUS":
            return col
    return None

def safe_str(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if str(value).strip().upper() == "NAN":
        return ""
    return str(value).strip()

def normalize_status(status_value):
    clean_val = safe_str(status_value)
    if not clean_val:
        return ""
    return clean_val.upper()

def is_after_recording_status(status_value):
    if not status_value:
        return False
    normalized = normalize_status(status_value)
    return normalized in AFTER_RECORDING_STATUSES

def contains_korean(text):
    if not text:
        return False
    text_str = safe_str(text)
    for char in text_str:
        code = ord(char)
        if (0xAC00 <= code <= 0xD7A3) or (0x1100 <= code <= 0x11FF) or (0x3130 <= code <= 0x318F):
            return True
    return False

def generate_casting_key(character_key, dialog_voice, speaker_groupkey, dialog_type=""):
    char_key = safe_str(character_key)
    dialog_v = safe_str(dialog_voice)
    speaker_gk = safe_str(speaker_groupkey)
    dtype = safe_str(dialog_type)
    
    if dtype.lower() in ["aidialog", "questdialog"]:
        return dialog_v if dialog_v else "Not Found"
    
    if dialog_v and "unique_" in dialog_v.lower():
        return dialog_v
    
    if char_key and speaker_gk and char_key.lower() in speaker_gk.lower():
        return speaker_gk.lower()
    
    if char_key:
        return char_key.lower()
    
    return "Not Found"

def generate_previous_data(prev_row, text_col, status_col, freememo_col):
    if not prev_row:
        return ""
    
    parts = [
        safe_str(prev_row.get(COL_STRORIGIN, "")),
        safe_str(prev_row.get(text_col, "")),
        safe_str(prev_row.get(status_col, "")),
        safe_str(prev_row.get(freememo_col, "")),
        safe_str(prev_row.get(COL_STARTFRAME, ""))
    ]
    
    return "|".join(parts)

def generate_color_for_value(value):
    fallback_colors = [
        "FFB3BA", "BAFFC9", "BAE1FF", "FFFFBA", "FFD9BA",
        "E0BBE4", "FFDFD3", "D4F1F4", "C9E4DE", "F7D9C4",
    ]
    hash_val = hash(str(value))
    return fallback_colors[hash_val % len(fallback_colors)]

def clean_numeric_columns(df):
    numeric_cols = [COL_STARTFRAME, COL_ENDFRAME]
    
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: safe_str(x).rstrip('0').rstrip('.') if '.' in safe_str(x) else safe_str(x))
    
    return df

def clean_dataframe_none_values(df):
    for col in df.columns:
        df[col] = df[col].apply(lambda x: "" if safe_str(x).upper() in ["NONE", "NAN", ""] else safe_str(x))
    return df

def safe_read_excel(filepath, header=0, dtype=str):
    wb = load_workbook(filepath, data_only=True, read_only=False)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    wb.close()
    
    if len(data) > 0:
        df = pd.DataFrame(data[1:], columns=data[0])
        for col in df.columns:
            df[col] = df[col].apply(safe_str)
        df = clean_numeric_columns(df)
        df = clean_dataframe_none_values(df)
        return df
    else:
        raise ValueError("Excel file is empty")

def normalize_dataframe_status(df):
    status_col = find_status_column(df.columns)
    if status_col:
        df[status_col] = df[status_col].apply(normalize_status)
        if status_col != "STATUS":
            df.rename(columns={status_col: "STATUS"}, inplace=True)
    return df

def filter_output_columns(df, column_list=OUTPUT_COLUMNS):
    available_cols = [col for col in column_list if col in df.columns]
    return df[available_cols]

# =========================================================================== #
# RAW VRS CHECK FUNCTIONS - WITH 3 KEY SYSTEM
# =========================================================================== #
def build_lookups(df):
    """Build 4 lookup dictionaries for matching (4-Tier Key System)"""
    lookup_cw = {}  # (SequenceName, EventName)
    lookup_cg = {}  # (SequenceName, StrOrigin)
    lookup_es = {}  # (EventName, StrOrigin)
    lookup_cs = {}  # (CastingKey, SequenceName) - NEW 4th key
    
    col_idx = {col: i for i, col in enumerate(df.columns)}
    
    total = len(df)
    for idx, row in enumerate(df.itertuples(index=False, name=None)):
        key_cw = (row[col_idx[COL_SEQUENCE]], row[col_idx[COL_EVENTNAME]])
        key_cg = (row[col_idx[COL_SEQUENCE]], row[col_idx[COL_STRORIGIN]])
        key_es = (row[col_idx[COL_EVENTNAME]], row[col_idx[COL_STRORIGIN]])
        key_cs = (row[col_idx[COL_CASTINGKEY]], row[col_idx[COL_SEQUENCE]])  # NEW
        
        if key_cw not in lookup_cw:
            lookup_cw[key_cw] = row
        if key_cg not in lookup_cg:
            lookup_cg[key_cg] = row[col_idx[COL_EVENTNAME]]
        if key_es not in lookup_es:
            lookup_es[key_es] = row
        if key_cs not in lookup_cs:
            lookup_cs[key_cs] = row  # NEW
        
        if (idx + 1) % 500 == 0 or idx == total - 1:
            print_progress(idx + 1, total, "Building lookups")
    
    finalize_progress()
    return lookup_cw, lookup_cg, lookup_es, lookup_cs  # Return 4 dicts

def classify_change(curr_row, prev_row, prev_lookup_cg, differences, col_idx):
    key_cg = (curr_row[col_idx[COL_SEQUENCE]], curr_row[col_idx[COL_STRORIGIN]])
    
    if key_cg in prev_lookup_cg and prev_lookup_cg[key_cg] != curr_row[col_idx[COL_EVENTNAME]]:
        strorigin_value = curr_row[col_idx[COL_STRORIGIN]]
        if contains_korean(strorigin_value):
            if len(differences) == 1 and differences[0] == COL_EVENTNAME:
                return "EventName Change", []
    
    existing_char_cols = [col for col in CHAR_GROUP_COLS if col in col_idx]
    char_group_diffs = [col for col in differences if col in existing_char_cols]
    
    if char_group_diffs:
        return "Character Group Change", char_group_diffs
    
    important_changes = []
    if COL_STRORIGIN in differences:
        important_changes.append("StrOrigin")
    if COL_DESC in differences:
        important_changes.append("Desc")
    if COL_STARTFRAME in differences:
        important_changes.append("TimeFrame")
    
    if important_changes:
        return "+".join(important_changes) + " Change", []
    
    return "No Relevant Change", []

def compare_rows(df_curr, prev_lookup_cw, prev_lookup_cg, prev_lookup_es, prev_lookup_cs):
    """Compare rows using 4-key system"""
    """Compare rows using 3-key system"""
    changes = []
    previous_strorigins = []
    changed_columns_map = {}
    counter = {}
    col_idx = {col: i for i, col in enumerate(df_curr.columns)}
    total_rows = len(df_curr)
    
    for idx, curr_row in enumerate(df_curr.itertuples(index=False, name=None)):
        key_cw = (curr_row[col_idx[COL_SEQUENCE]], curr_row[col_idx[COL_EVENTNAME]])
        key_cg = (curr_row[col_idx[COL_SEQUENCE]], curr_row[col_idx[COL_STRORIGIN]])
        key_es = (curr_row[col_idx[COL_EVENTNAME]], curr_row[col_idx[COL_STRORIGIN]])  # NEW
        key_cs = (curr_row[col_idx[COL_CASTINGKEY]], curr_row[col_idx[COL_SEQUENCE]])  # NEW 4th key
        
        # Stage 1: Direct match by (Sequence, EventName)
        if key_cw in prev_lookup_cw:
            prev_row = prev_lookup_cw[key_cw]
            prev_strorigin = safe_str(prev_row[col_idx[COL_STRORIGIN]])
            
            differences = [
                col for col in df_curr.columns
                if safe_str(curr_row[col_idx[col]]) != safe_str(prev_row[col_idx[col]])
            ]
            
            if not differences:
                change_label = "No Change"
                changed_char_cols = []
            else:
                existing_char_cols = [col for col in CHAR_GROUP_COLS if col in col_idx]
                char_group_diffs = [col for col in differences if col in existing_char_cols]
                
                if char_group_diffs:
                    change_label = "Character Group Change"
                    changed_char_cols = char_group_diffs
                else:
                    important_changes = []
                    if COL_STRORIGIN in differences:
                        important_changes.append("StrOrigin")
                    if COL_DESC in differences:
                        important_changes.append("Desc")
                    if COL_STARTFRAME in differences:
                        important_changes.append("TimeFrame")
                    
                    if important_changes:
                        change_label = "+".join(important_changes) + " Change"
                    else:
                        change_label = "No Relevant Change"
                    changed_char_cols = []
        
        # Stage 2: StrOrigin + Sequence match - VERIFY with Key 4 (4-Tier System)
        elif key_cg in prev_lookup_cg:
            # Check if same character (Key 4 verification)
            if key_cs in prev_lookup_cs:
                # Same character → EventName Change
                old_eventname = prev_lookup_cg[key_cg]
                prev_row = prev_lookup_cw.get((curr_row[col_idx[COL_SEQUENCE]], old_eventname))

                if prev_row:
                    prev_strorigin = safe_str(prev_row[col_idx[COL_STRORIGIN]])

                    differences = [
                        col for col in df_curr.columns
                        if safe_str(curr_row[col_idx[col]]) != safe_str(prev_row[col_idx[col]])
                    ]

                    important_changes = []
                    if COL_STRORIGIN in differences:
                        important_changes.append("StrOrigin")
                    if COL_DESC in differences:
                        important_changes.append("Desc")
                    if COL_STARTFRAME in differences:
                        important_changes.append("TimeFrame")

                    existing_char_cols = [col for col in CHAR_GROUP_COLS if col in col_idx]
                    char_group_diffs = [col for col in differences if col in existing_char_cols]

                    if char_group_diffs:
                        change_label = "Character Group Change"
                        changed_char_cols = char_group_diffs
                    elif not important_changes:
                        if contains_korean(curr_row[col_idx[COL_STRORIGIN]]):
                            change_label = "EventName Change"
                        else:
                            change_label = "No Relevant Change"
                        changed_char_cols = []
                    else:
                        important_changes.insert(0, "EventName")
                        change_label = "+".join(important_changes) + " Change"
                        changed_char_cols = []
                else:
                    change_label = "New Row"
                    prev_row = None
                    prev_strorigin = ""
                    changed_char_cols = []
            else:
                # Different character → New Row (duplicate StrOrigin case)
                change_label = "New Row"
                prev_row = None
                prev_strorigin = ""
                changed_char_cols = []
        
        # Stage 3: SequenceName changed - match by (EventName, StrOrigin) - NEW
        elif key_es in prev_lookup_es:
            prev_row = prev_lookup_es[key_es]
            prev_strorigin = safe_str(prev_row[col_idx[COL_STRORIGIN]])
            
            # Same content but different sequence
            change_label = "SequenceName Change"
            changed_char_cols = []
        
        # Stage 4: Truly new row
        else:
            change_label = "New Row"
            prev_row = None
            prev_strorigin = ""
            changed_char_cols = []
        
        changes.append(change_label)
        previous_strorigins.append(prev_strorigin)
        
        if changed_char_cols:
            changed_columns_map[idx] = changed_char_cols
        
        counter[change_label] = counter.get(change_label, 0) + 1
        
        if (idx + 1) % 500 == 0 or idx == total_rows - 1:
            print_progress(idx + 1, total_rows, "Comparing rows")
    
    finalize_progress()
    return changes, previous_strorigins, changed_columns_map, counter

def find_deleted_rows(df_prev, df_curr):
    """Find deleted rows using 3-key system"""
    # Build all 3 key types from current
    curr_keys_cw = set((row[COL_SEQUENCE], row[COL_EVENTNAME]) for row in df_curr.to_dict("records"))
    curr_keys_cg = set((row[COL_SEQUENCE], row[COL_STRORIGIN]) for row in df_curr.to_dict("records"))
    curr_keys_es = set((row[COL_EVENTNAME], row[COL_STRORIGIN]) for row in df_curr.to_dict("records"))  # NEW
    
    deleted_mask = []
    for row in df_prev.to_dict("records"):
        key_cw = (row[COL_SEQUENCE], row[COL_EVENTNAME])
        key_cg = (row[COL_SEQUENCE], row[COL_STRORIGIN])
        key_es = (row[COL_EVENTNAME], row[COL_STRORIGIN])  # NEW
        
        # Only mark as deleted if ALL 3 keys are missing
        is_deleted = (key_cw not in curr_keys_cw) and \
                     (key_cg not in curr_keys_cg) and \
                     (key_es not in curr_keys_es)  # NEW
        deleted_mask.append(is_deleted)
    
    return df_prev[deleted_mask].copy()

def create_raw_summary(counter, prev_path, curr_path, df_res):
    summary_rows = [
        ["RAW VRS CHECK SUMMARY", "", "", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", ""],
        ["Previous file", os.path.basename(prev_path), "", ""],
        ["Current file", os.path.basename(curr_path), "", ""],
        ["", "", "", ""],
        ["RESULT COUNTS", "Count", "Word Count (Korean)", "Word Count (Translation)"],
    ]
    
    sorted_keys = sorted([k for k in counter.keys() if k != "Deleted Rows"])
    for key in sorted_keys:
        word_count_korean = df_res.loc[df_res["CHANGES"] == key, COL_STRORIGIN].apply(lambda x: len(safe_str(x).split()) if safe_str(x) else 0).sum()
        word_count_translation = df_res.loc[df_res["CHANGES"] == key, COL_TEXT].apply(lambda x: len(safe_str(x).split()) if safe_str(x) else 0).sum()
        summary_rows.append([key, counter[key], word_count_korean, word_count_translation])
    
    if "Deleted Rows" in counter:
        summary_rows.append(["Deleted Rows", counter["Deleted Rows"], 0, 0])
    
    return pd.DataFrame(summary_rows, columns=["Metric", "Value", "Word Count (Korean)", "Word Count (Translation)"])

# =========================================================================== #
# WORKING VRS CHECK FUNCTIONS - WITH 3 KEY SYSTEM
# =========================================================================== #
def build_working_lookups(df, label="PREVIOUS"):
    """Build 4 lookup dictionaries for Working Process (4-Tier Key System)"""
    lookup_cw = {}
    lookup_cg = {}
    lookup_es = {}
    lookup_cs = {}  # NEW 4th key

    total = len(df)
    log(f"Building {label} lookup dictionaries...")

    for idx, row in df.iterrows():
        row_dict = row.to_dict()

        key_cw = (row[COL_SEQUENCE], row[COL_EVENTNAME])
        key_cg = (row[COL_SEQUENCE], row[COL_STRORIGIN])
        key_es = (row[COL_EVENTNAME], row[COL_STRORIGIN])
        key_cs = (row[COL_CASTINGKEY], row[COL_SEQUENCE])  # NEW 4th key

        if key_cw not in lookup_cw:
            lookup_cw[key_cw] = row_dict
        if key_cg not in lookup_cg:
            lookup_cg[key_cg] = row[COL_EVENTNAME]
        if key_es not in lookup_es:
            lookup_es[key_es] = row_dict
        if key_cs not in lookup_cs:
            lookup_cs[key_cs] = row_dict  # NEW 4th key

        if (idx + 1) % 500 == 0 or idx == total - 1:
            print_progress(idx + 1, total, f"Indexing {label}")

    finalize_progress()
    log(f"  → Indexed {len(lookup_cw):,} unique {label} rows")
    return lookup_cw, lookup_cg, lookup_es, lookup_cs  # Return 4 dicts

def classify_working_change(curr_row, prev_row, prev_lookup_cg, prev_lookup_cs):
    """Classify change type with 4-key system (4-Tier Key System)"""
    key_cg = (curr_row[COL_SEQUENCE], curr_row[COL_STRORIGIN])
    key_cs = (curr_row[COL_CASTINGKEY], curr_row[COL_SEQUENCE])  # NEW 4th key

    # Check for EventName Change with Key 4 verification
    if key_cg in prev_lookup_cg and prev_lookup_cg[key_cg] != curr_row[COL_EVENTNAME]:
        # Verify if same character (Key 4 check)
        if key_cs in prev_lookup_cs:
            # Same character → EventName Change
            strorigin_value = curr_row[COL_STRORIGIN]
            if contains_korean(strorigin_value):
                return "EventName Change"
        # Different character → Not EventName Change (will be classified as New Row elsewhere)

    differences = [col for col in curr_row.keys() if col in prev_row and curr_row[col] != prev_row[col]]

    if not differences:
        return "No Change"

    important_changes = []
    if COL_STRORIGIN in differences:
        important_changes.append("StrOrigin")
    if COL_DESC in differences:
        important_changes.append("Desc")
    if COL_STARTFRAME in differences:
        important_changes.append("TimeFrame")

    if important_changes:
        return "+".join(important_changes) + " Change"

    return "No Relevant Change"

def classify_alllang_change(curr_row, prev_row, prev_lookup_cg, prev_lookup_cs):
    """Classify change type for All Language Process with 4-key system (4-Tier Key System)"""
    key_cg = (curr_row[COL_SEQUENCE], curr_row[COL_STRORIGIN])
    key_cs = (curr_row[COL_CASTINGKEY], curr_row[COL_SEQUENCE])  # NEW 4th key

    # Check for EventName Change with Key 4 verification
    if key_cg in prev_lookup_cg and prev_lookup_cg[key_cg] != curr_row[COL_EVENTNAME]:
        # Verify if same character (Key 4 check)
        if key_cs in prev_lookup_cs:
            # Same character → EventName Change
            strorigin_value = curr_row[COL_STRORIGIN]
            if contains_korean(strorigin_value):
                return "EventName Change"
        # Different character → Not EventName Change (will be classified as New Row elsewhere)

    differences = [col for col in curr_row.keys() if col in prev_row and curr_row[col] != prev_row[col]]

    if not differences:
        return "No Change"

    important_changes = []
    if COL_STRORIGIN in differences:
        important_changes.append("StrOrigin")
    if COL_STARTFRAME in differences:
        important_changes.append("TimeFrame")

    if important_changes:
        return "+".join(important_changes) + " Change"

    return "No Relevant Change"

def apply_import_logic(curr_row, prev_row, change_type):
    result = {}
    
    if prev_row:
        result[COL_FREEMEMO] = safe_str(prev_row.get(COL_FREEMEMO, ""))
    
    if change_type == "New Row":
        return result
    
    prev_status = prev_row.get(COL_STATUS, "") if prev_row else ""
    is_after_recording = is_after_recording_status(prev_status)
    
    if "StrOrigin" in change_type:
        if is_after_recording:
            result[COL_TEXT] = safe_str(prev_row.get(COL_TEXT, "")) if prev_row else ""
            result[COL_DESC] = safe_str(prev_row.get(COL_DESC, "")) if prev_row else ""
            result[COL_STATUS] = prev_status
        else:
            result[COL_TEXT] = safe_str(curr_row.get(COL_TEXT, ""))
            result[COL_DESC] = safe_str(prev_row.get(COL_DESC, "")) if prev_row else ""
            result[COL_STATUS] = "NEED CHECK"
    
    elif change_type == "Desc Change":
        result[COL_TEXT] = safe_str(prev_row.get(COL_TEXT, "")) if prev_row else ""
        result[COL_DESC] = safe_str(curr_row.get(COL_DESC, ""))
        result[COL_STATUS] = prev_status
    
    elif "TimeFrame" in change_type:
        result[COL_TEXT] = safe_str(prev_row.get(COL_TEXT, "")) if prev_row else ""
        result[COL_DESC] = safe_str(prev_row.get(COL_DESC, "")) if prev_row else ""
        result[COL_STATUS] = prev_status
    
    elif change_type in ["No Change", "No Relevant Change", "EventName Change", "SequenceName Change"]:  # Added SequenceName Change
        result[COL_TEXT] = safe_str(prev_row.get(COL_TEXT, "")) if prev_row else ""
        result[COL_DESC] = safe_str(prev_row.get(COL_DESC, "")) if prev_row else ""
        result[COL_STATUS] = prev_status
    
    else:
        result[COL_TEXT] = safe_str(prev_row.get(COL_TEXT, "")) if prev_row else ""
        result[COL_DESC] = safe_str(prev_row.get(COL_DESC, "")) if prev_row else ""
        result[COL_STATUS] = prev_status
    
    return result

def apply_import_logic_alllang_lang(curr_row, prev_row, change_type, lang_suffix):
    result = {}
    
    text_col = f"Text_{lang_suffix}"
    status_col = f"STATUS_{lang_suffix}"
    freememo_col = f"FREEMEMO_{lang_suffix}"
    
    if prev_row:
        result[freememo_col] = safe_str(prev_row.get(COL_FREEMEMO, ""))
    
    if change_type == "New Row":
        return result
    
    prev_status = prev_row.get(COL_STATUS, "") if prev_row else ""
    is_after_recording = is_after_recording_status(prev_status)
    
    if "StrOrigin" in change_type:
        if is_after_recording:
            result[text_col] = safe_str(prev_row.get(COL_TEXT, "")) if prev_row else ""
            result[status_col] = prev_status
        else:
            result[text_col] = safe_str(curr_row.get(text_col, ""))
            result[status_col] = "NEED CHECK"
    
    elif "TimeFrame" in change_type:
        result[text_col] = safe_str(prev_row.get(COL_TEXT, "")) if prev_row else ""
        result[status_col] = prev_status
    
    elif change_type in ["No Change", "No Relevant Change", "EventName Change", "SequenceName Change"]:  # Added SequenceName Change
        result[text_col] = safe_str(prev_row.get(COL_TEXT, "")) if prev_row else ""
        result[status_col] = prev_status
    
    else:
        result[text_col] = safe_str(prev_row.get(COL_TEXT, "")) if prev_row else ""
        result[status_col] = prev_status
    
    return result

def process_working_comparison(df_curr, prev_lookup_cw, prev_lookup_cg, prev_lookup_es, prev_lookup_cs):
    """Compare with 4-key system (4-Tier Key System)"""
    log("Comparing and importing data...")

    results = []
    counter = {}
    total_rows = len(df_curr)

    for idx, curr_row in df_curr.iterrows():
        curr_dict = curr_row.to_dict()
        key_cw = (curr_row[COL_SEQUENCE], curr_row[COL_EVENTNAME])
        key_cg = (curr_row[COL_SEQUENCE], curr_row[COL_STRORIGIN])
        key_es = (curr_row[COL_EVENTNAME], curr_row[COL_STRORIGIN])
        key_cs = (curr_row[COL_CASTINGKEY], curr_row[COL_SEQUENCE])  # NEW 4th key

        mainline_translation = safe_str(curr_dict.get(COL_TEXT, ""))

        # Stage 1: Direct match (SequenceName + EventName)
        if key_cw in prev_lookup_cw:
            prev_row = prev_lookup_cw[key_cw]

            differences = [col for col in curr_dict.keys() if col in prev_row and curr_dict[col] != prev_row[col]]

            if not differences:
                change_type = "No Change"
            else:
                important_changes = []
                if COL_STRORIGIN in differences:
                    important_changes.append("StrOrigin")
                if COL_DESC in differences:
                    important_changes.append("Desc")
                if COL_STARTFRAME in differences:
                    important_changes.append("TimeFrame")

                if important_changes:
                    change_type = "+".join(important_changes) + " Change"
                else:
                    change_type = "No Relevant Change"

        # Stage 2: StrOrigin+Sequence match - VERIFY with Key 4 (4-Tier System)
        elif key_cg in prev_lookup_cg:
            # Check if same character (Key 4 verification)
            if key_cs in prev_lookup_cs:
                # Same character → EventName Change
                old_eventname = prev_lookup_cg[key_cg]
                prev_row = prev_lookup_cw.get((curr_row[COL_SEQUENCE], old_eventname))

                if prev_row:
                    differences = [col for col in curr_dict.keys() if col in prev_row and curr_dict[col] != prev_row[col]]

                    important_changes = []
                    if COL_STRORIGIN in differences:
                        important_changes.append("StrOrigin")
                    if COL_DESC in differences:
                        important_changes.append("Desc")
                    if COL_STARTFRAME in differences:
                        important_changes.append("TimeFrame")

                    if not important_changes:
                        if contains_korean(curr_row[COL_STRORIGIN]):
                            change_type = "EventName Change"
                        else:
                            change_type = "No Relevant Change"
                    else:
                        important_changes.insert(0, "EventName")
                        change_type = "+".join(important_changes) + " Change"
                else:
                    change_type = "New Row"
                    prev_row = None
            else:
                # Different character → New Row (duplicate StrOrigin case)
                change_type = "New Row"
                prev_row = None

        # Stage 3: SequenceName changed (EventName + StrOrigin match)
        elif key_es in prev_lookup_es:
            prev_row = prev_lookup_es[key_es]
            change_type = "SequenceName Change"

        # Stage 4: Truly new (no keys match)
        else:
            change_type = "New Row"
            prev_row = None
        
        import_data = apply_import_logic(curr_dict, prev_row, change_type)
        
        for col, value in import_data.items():
            curr_dict[col] = value
        
        curr_dict[COL_CASTINGKEY] = generate_casting_key(
            curr_dict.get(COL_CHARACTERKEY, ""),
            curr_dict.get(COL_DIALOGVOICE, ""),
            curr_dict.get(COL_SPEAKER_GROUPKEY, ""),
            curr_dict.get("DialogType", "")
        )
        
        curr_dict[COL_PREVIOUSDATA] = generate_previous_data(
            prev_row, COL_TEXT, COL_STATUS, COL_FREEMEMO
        )
        
        curr_dict[COL_MAINLINE_TRANSLATION] = mainline_translation
        curr_dict["CHANGES"] = change_type
        
        results.append(curr_dict)
        counter[change_type] = counter.get(change_type, 0) + 1
        
        if (idx + 1) % 500 == 0 or idx == total_rows - 1:
            print_progress(idx + 1, total_rows, "Processing rows")
    
    finalize_progress()
    return pd.DataFrame(results), counter

def find_working_deleted_rows(df_prev, df_curr):
    """Find deleted with 4-key system (4-Tier Key System)"""
    log("Finding deleted rows (PREVIOUS rows not in CURRENT)...")

    # Build all 4 key types
    curr_keys_cw = set((row[COL_SEQUENCE], row[COL_EVENTNAME]) for _, row in df_curr.iterrows())
    curr_keys_cg = set((row[COL_SEQUENCE], row[COL_STRORIGIN]) for _, row in df_curr.iterrows())
    curr_keys_es = set((row[COL_EVENTNAME], row[COL_STRORIGIN]) for _, row in df_curr.iterrows())
    curr_keys_cs = set((row[COL_CASTINGKEY], row[COL_SEQUENCE]) for _, row in df_curr.iterrows())  # NEW 4th key

    deleted_rows = []

    for idx, row in df_prev.iterrows():
        key_cw = (row[COL_SEQUENCE], row[COL_EVENTNAME])
        key_cg = (row[COL_SEQUENCE], row[COL_STRORIGIN])
        key_es = (row[COL_EVENTNAME], row[COL_STRORIGIN])
        key_cs = (row[COL_CASTINGKEY], row[COL_SEQUENCE])  # NEW 4th key

        # Only mark as deleted if ALL 4 keys are missing
        if (key_cw not in curr_keys_cw) and \
           (key_cg not in curr_keys_cg) and \
           (key_es not in curr_keys_es) and \
           (key_cs not in curr_keys_cs):  # NEW 4th key check
            deleted_rows.append(row.to_dict())
    
    log(f"  → Found {len(deleted_rows):,} deleted rows")
    return pd.DataFrame(deleted_rows) if deleted_rows else pd.DataFrame()

def create_working_summary(counter, prev_path, curr_path, df_res):
    summary_rows = [
        ["WORKING VRS CHECK SUMMARY", "", "", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", ""],
        ["Previous file", os.path.basename(prev_path), "", ""],
        ["Current file", os.path.basename(curr_path), "", ""],
        ["", "", "", ""],
        ["RESULT COUNTS", "Count", "Word Count (Korean)", "Word Count (Translation)"],
    ]
    
    sorted_keys = sorted([k for k in counter.keys() if k != "Deleted Rows"])
    for key in sorted_keys:
        word_count_korean = df_res.loc[df_res["CHANGES"] == key, COL_STRORIGIN].apply(lambda x: len(safe_str(x).split()) if safe_str(x) else 0).sum()
        word_count_translation = df_res.loc[df_res["CHANGES"] == key, COL_TEXT].apply(lambda x: len(safe_str(x).split()) if safe_str(x) else 0).sum()
        summary_rows.append([key, counter[key], word_count_korean, word_count_translation])
    
    if "Deleted Rows" in counter:
        summary_rows.append(["Deleted Rows", counter["Deleted Rows"], 0, 0])
    
    return pd.DataFrame(summary_rows, columns=["Metric", "Value", "Word Count (Korean)", "Word Count (Translation)"])

def create_working_update_history_sheet():
    history = load_update_history("working")
    
    history_data = [
        ["WORKING PROCESS - UPDATE HISTORY"],
        [""],
        ["This update history is tracked in: " + WORKING_HISTORY_FILE],
        ["Use 'View Update History' button in main GUI for complete details"],
        [""],
    ]
    
    if not history["updates"]:
        history_data.append(["No updates recorded yet"])
        history_data.append([""])
    else:
        latest = history["updates"][-1]
        history_data.append(["LATEST UPDATE"])
        history_data.append([f"Timestamp: {latest['timestamp']}"])
        history_data.append([f"Output File: {latest['output_file']}"])
        history_data.append([f"Previous: {latest['previous_file']}"])
        history_data.append([f"Current: {latest['current_file']}"])
        history_data.append([""])
        history_data.append([f"Total Rows: {latest['statistics']['total_rows']:,}"])
        history_data.append([""])
        history_data.append(["Change Breakdown:"])
        for key, value in latest['statistics'].items():
            if key != 'total_rows':
                history_data.append([f"  {key}: {value:,}"])
        history_data.append([""])
        history_data.append([f"Total Updates Recorded: {len(history['updates'])}"])
    
    history_data.append([""])
    history_data.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    
    return pd.DataFrame(history_data, columns=["Update History"])

# =========================================================================== #
# ALL LANGUAGE VRS CHECK FUNCTIONS - WITH 3 KEY SYSTEM
# =========================================================================== #
def find_alllang_files():
    script_dir = get_script_dir()
    previous_folder = os.path.join(script_dir, "Previous")
    current_folder = os.path.join(script_dir, "Current")
    
    if not os.path.exists(previous_folder):
        raise FileNotFoundError(f"Previous folder not found: {previous_folder}")
    if not os.path.exists(current_folder):
        raise FileNotFoundError(f"Current folder not found: {current_folder}")
    
    current_files = [f for f in os.listdir(current_folder) if f.endswith(('.xlsx', '.xlsm', '.xls'))]
    
    curr_kr = None
    curr_en = None
    curr_cn = None
    
    for filename in current_files:
        if '_KR' in filename:
            curr_kr = os.path.join(current_folder, filename)
        elif '_EN' in filename:
            curr_en = os.path.join(current_folder, filename)
        elif '_CN' in filename:
            curr_cn = os.path.join(current_folder, filename)
    
    if not curr_kr:
        raise FileNotFoundError("Korean current file (_KR) not found in Current folder")
    if not curr_en:
        raise FileNotFoundError("English current file (_EN) not found in Current folder")
    if not curr_cn:
        raise FileNotFoundError("Chinese current file (_CN) not found in Current folder")
    
    previous_files = [f for f in os.listdir(previous_folder) if f.endswith(('.xlsx', '.xlsm', '.xls'))]
    
    prev_kr = None
    prev_en = None
    prev_cn = None
    
    for filename in previous_files:
        if '_KR' in filename:
            prev_kr = os.path.join(previous_folder, filename)
        elif '_EN' in filename:
            prev_en = os.path.join(previous_folder, filename)
        elif '_CN' in filename:
            prev_cn = os.path.join(previous_folder, filename)
    
    return curr_kr, curr_en, curr_cn, prev_kr, prev_en, prev_cn

def merge_current_files(curr_kr_path, curr_en_path, curr_cn_path):
    log("\n" + "="*70)
    log("PHASE 1: MERGING CURRENT FILES")
    log("="*70)
    
    log(f"Reading KR Current: {os.path.basename(curr_kr_path)}")
    df_kr = safe_read_excel(curr_kr_path, header=0, dtype=str)
    df_kr = normalize_dataframe_status(df_kr)
    log(f"  → {len(df_kr):,} rows")
    
    log(f"Reading EN Current: {os.path.basename(curr_en_path)}")
    df_en = safe_read_excel(curr_en_path, header=0, dtype=str)
    df_en = normalize_dataframe_status(df_en)
    log(f"  → {len(df_en):,} rows")
    
    log(f"Reading CN Current: {os.path.basename(curr_cn_path)}")
    df_cn = safe_read_excel(curr_cn_path, header=0, dtype=str)
    df_cn = normalize_dataframe_status(df_cn)
    log(f"  → {len(df_cn):,} rows")
    
    log("Building EN/CN lookup dictionaries...")
    lookup_en = {}
    lookup_cn = {}
    
    for _, row in df_en.iterrows():
        key = (row[COL_SEQUENCE], row[COL_EVENTNAME])
        lookup_en[key] = row.to_dict()
    
    for _, row in df_cn.iterrows():
        key = (row[COL_SEQUENCE], row[COL_EVENTNAME])
        lookup_cn[key] = row.to_dict()
    
    log(f"  → EN: {len(lookup_en):,} rows indexed")
    log(f"  → CN: {len(lookup_cn):,} rows indexed")
    
    log("Merging data into unified structure...")
    merged_rows = []
    total = len(df_kr)
    
    for idx, kr_row in df_kr.iterrows():
        merged = kr_row.to_dict()
        key = (kr_row[COL_SEQUENCE], kr_row[COL_EVENTNAME])
        
        merged["Text_KR"] = safe_str(merged.pop(COL_TEXT, ""))
        merged["FREEMEMO_KR"] = safe_str(merged.pop(COL_FREEMEMO, ""))
        merged["STATUS_KR"] = safe_str(merged.pop(COL_STATUS, ""))
        merged["CharacterName_KR"] = safe_str(merged.pop(COL_CHARACTERNAME, ""))
        
        if key in lookup_en:
            en_row = lookup_en[key]
            merged["Text_EN"] = safe_str(en_row.get(COL_TEXT, ""))
            merged["FREEMEMO_EN"] = safe_str(en_row.get(COL_FREEMEMO, ""))
            merged["STATUS_EN"] = safe_str(en_row.get(COL_STATUS, ""))
            merged["CharacterName_EN"] = safe_str(en_row.get(COL_CHARACTERNAME, ""))
        else:
            merged["Text_EN"] = ""
            merged["FREEMEMO_EN"] = ""
            merged["STATUS_EN"] = ""
            merged["CharacterName_EN"] = ""
        
        if key in lookup_cn:
            cn_row = lookup_cn[key]
            merged["Text_CN"] = safe_str(cn_row.get(COL_TEXT, ""))
            merged["FREEMEMO_CN"] = safe_str(cn_row.get(COL_FREEMEMO, ""))
            merged["STATUS_CN"] = safe_str(cn_row.get(COL_STATUS, ""))
            merged["CharacterName_CN"] = safe_str(cn_row.get(COL_CHARACTERNAME, ""))
        else:
            merged["Text_CN"] = ""
            merged["FREEMEMO_CN"] = ""
            merged["STATUS_CN"] = ""
            merged["CharacterName_CN"] = ""
        
        merged_rows.append(merged)
        
        if (idx + 1) % 500 == 0 or idx == total - 1:
            print_progress(idx + 1, total, "Merging")
    
    finalize_progress()
    
    df_merged = pd.DataFrame(merged_rows)
    log(f"✓ Unified structure created: {len(df_merged):,} rows with tri-lingual columns")
    
    return df_merged

def process_alllang_comparison(df_curr, lookup_kr, lookup_en, lookup_cn,
                               lookup_cg_kr, lookup_es_kr, lookup_cs_kr,  # Added lookup_cs_kr (4th key)
                               has_kr, has_en, has_cn):
    """All language with 4-key system (4-Tier Key System)"""
    log("\n" + "="*70)
    log("PHASE 2: IMPORTING DATA FROM PREVIOUS FILES")
    log("="*70)
    log(f"Languages to update: KR={has_kr}, EN={has_en}, CN={has_cn}")

    results = []
    counter = {}
    total_rows = len(df_curr)

    for idx, curr_row in df_curr.iterrows():
        curr_dict = curr_row.to_dict()
        key_cw = (curr_row[COL_SEQUENCE], curr_row[COL_EVENTNAME])
        key_cg = (curr_row[COL_SEQUENCE], curr_row[COL_STRORIGIN])
        key_es = (curr_row[COL_EVENTNAME], curr_row[COL_STRORIGIN])
        key_cs = (curr_row[COL_CASTINGKEY], curr_row[COL_SEQUENCE])  # NEW 4th key

        mainline_kr = safe_str(curr_dict.get("Text_KR", ""))
        mainline_en = safe_str(curr_dict.get("Text_EN", ""))
        mainline_cn = safe_str(curr_dict.get("Text_CN", ""))

        if has_kr:
            # Stage 1: Direct match (SequenceName + EventName)
            if key_cw in lookup_kr:
                prev_kr = lookup_kr[key_cw]

                differences = [col for col in curr_dict.keys() if col in prev_kr and curr_dict[col] != prev_kr[col]]

                if not differences:
                    change_type = "No Change"
                else:
                    important_changes = []
                    if COL_STRORIGIN in differences:
                        important_changes.append("StrOrigin")
                    if COL_STARTFRAME in differences:
                        important_changes.append("TimeFrame")

                    if important_changes:
                        change_type = "+".join(important_changes) + " Change"
                    else:
                        change_type = "No Relevant Change"

            # Stage 2: StrOrigin+Sequence match - VERIFY with Key 4 (4-Tier System)
            elif key_cg in lookup_cg_kr:
                # Check if same character (Key 4 verification)
                if key_cs in lookup_cs_kr:
                    # Same character → EventName Change
                    old_eventname = lookup_cg_kr[key_cg]
                    prev_kr = lookup_kr.get((curr_row[COL_SEQUENCE], old_eventname))

                    if prev_kr:
                        differences = [col for col in curr_dict.keys() if col in prev_kr and curr_dict[col] != prev_kr[col]]

                        important_changes = []
                        if COL_STRORIGIN in differences:
                            important_changes.append("StrOrigin")
                        if COL_STARTFRAME in differences:
                            important_changes.append("TimeFrame")

                        if not important_changes:
                            if contains_korean(curr_row[COL_STRORIGIN]):
                                change_type = "EventName Change"
                            else:
                                change_type = "No Relevant Change"
                        else:
                            important_changes.insert(0, "EventName")
                            change_type = "+".join(important_changes) + " Change"
                    else:
                        change_type = "New Row"
                        prev_kr = None
                else:
                    # Different character → New Row (duplicate StrOrigin case)
                    change_type = "New Row"
                    prev_kr = None

            # Stage 3: SequenceName changed (EventName + StrOrigin match)
            elif key_es in lookup_es_kr:
                prev_kr = lookup_es_kr[key_es]
                change_type = "SequenceName Change"

            # Stage 4: Truly new (no keys match)
            else:
                change_type = "New Row"
                prev_kr = None
        else:
            change_type = "No Change"
            prev_kr = None
        
        if has_kr:
            import_kr = apply_import_logic_alllang_lang(curr_dict, prev_kr, change_type, "KR")
            for col, value in import_kr.items():
                curr_dict[col] = value
            
            curr_dict["PreviousData_KR"] = generate_previous_data(
                prev_kr, COL_TEXT, COL_STATUS, COL_FREEMEMO
            ) if prev_kr else ""
        else:
            curr_dict["PreviousData_KR"] = ""
        
        if has_en:
            prev_en = lookup_en.get(key_cw)
            import_en = apply_import_logic_alllang_lang(curr_dict, prev_en, change_type, "EN")
            for col, value in import_en.items():
                curr_dict[col] = value
            
            curr_dict["PreviousData_EN"] = generate_previous_data(
                prev_en, COL_TEXT, COL_STATUS, COL_FREEMEMO
            ) if prev_en else ""
        else:
            curr_dict["PreviousData_EN"] = ""
        
        if has_cn:
            prev_cn = lookup_cn.get(key_cw)
            import_cn = apply_import_logic_alllang_lang(curr_dict, prev_cn, change_type, "CN")
            for col, value in import_cn.items():
                curr_dict[col] = value
            
            curr_dict["PreviousData_CN"] = generate_previous_data(
                prev_cn, COL_TEXT, COL_STATUS, COL_FREEMEMO
            ) if prev_cn else ""
        else:
            curr_dict["PreviousData_CN"] = ""
        
        curr_dict[COL_CASTINGKEY] = generate_casting_key(
            curr_dict.get(COL_CHARACTERKEY, ""),
            curr_dict.get(COL_DIALOGVOICE, ""),
            curr_dict.get(COL_SPEAKER_GROUPKEY, ""),
            curr_dict.get("DialogType", "")
        )
        
        curr_dict["Mainline Translation_KR"] = mainline_kr
        curr_dict["Mainline Translation_EN"] = mainline_en
        curr_dict["Mainline Translation_CN"] = mainline_cn
        
        curr_dict["CHANGES"] = change_type
        
        results.append(curr_dict)
        counter[change_type] = counter.get(change_type, 0) + 1
        
        if (idx + 1) % 500 == 0 or idx == total_rows - 1:
            print_progress(idx + 1, total_rows, "Processing rows")
    
    finalize_progress()
    return pd.DataFrame(results), counter

def create_alllang_update_history_sheet():
    history = load_update_history("alllang")
    
    history_data = [
        ["ALL LANGUAGE PROCESS - UPDATE HISTORY"],
        [""],
        ["This update history is tracked in: " + ALLLANG_HISTORY_FILE],
        ["Use 'View Update History' button in main GUI for complete details"],
        [""],
    ]
    
    if not history["updates"]:
        history_data.append(["No updates recorded yet"])
        history_data.append([""])
    else:
        latest = history["updates"][-1]
        history_data.append(["LATEST UPDATE"])
        history_data.append([f"Timestamp: {latest['timestamp']}"])
        history_data.append([f"Output File: {latest['output_file']}"])
        history_data.append([""])
        history_data.append(["LANGUAGES UPDATED:"])
        for lang in ["KR", "EN", "CN"]:
            status = "✓ UPDATED" if latest["languages_updated"][lang] else "○ Preserved"
            previous = latest["previous_files"][lang] if latest["previous_files"][lang] else "N/A"
            history_data.append([f"  {lang}: {status} | Previous: {previous}"])
        history_data.append([""])
        history_data.append(["CURRENT FILES (Complete Base):"])
        for lang in ["KR", "EN", "CN"]:
            history_data.append([f"  {lang}: {latest['current_files'][lang]}"])
        history_data.append([""])
        history_data.append([f"Total Rows: {latest['statistics']['total_rows']:,}"])
        history_data.append([""])
        history_data.append(["Change Breakdown:"])
        for key, value in latest['statistics'].items():
            if key != 'total_rows':
                history_data.append([f"  {key}: {value:,}"])
        history_data.append([""])
        history_data.append([f"Total Updates Recorded: {len(history['updates'])}"])
    
    history_data.append([""])
    history_data.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    
    return pd.DataFrame(history_data, columns=["Update History"])

def create_alllang_summary(counter, prev_kr, prev_en, prev_cn, curr_kr, df_res, has_kr, has_en, has_cn):
    summary_rows = [
        ["ALL LANGUAGE VRS CHECK SUMMARY", "", "", "", "", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", "", ""],
        ["", "", "", "", "", ""],
        ["CURRENT FILES (Complete Base)", "", "", "", "", ""],
        ["  Korean", os.path.basename(curr_kr), "", "", "", ""],
        ["", "", "", "", "", ""],
        ["PREVIOUS FILES (Updated)", "", "", "", "", ""],
        ["  Korean", os.path.basename(prev_kr) if prev_kr else "Not updated", "", "", "", ""],
        ["  English", os.path.basename(prev_en) if prev_en else "Not updated", "", "", "", ""],
        ["  Chinese", os.path.basename(prev_cn) if prev_cn else "Not updated", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["RESULT COUNTS", "Count", "Word Count (Korean)", "Word Count KR", "Word Count EN", "Word Count CN"],
    ]
    
    sorted_keys = sorted([k for k in counter.keys() if k != "Deleted Rows"])
    for key in sorted_keys:
        word_count_korean = df_res.loc[df_res["CHANGES"] == key, COL_STRORIGIN].apply(
            lambda x: len(safe_str(x).split()) if safe_str(x) else 0
        ).sum()
        
        word_count_kr = df_res.loc[df_res["CHANGES"] == key, "Text_KR"].apply(
            lambda x: len(safe_str(x).split()) if safe_str(x) else 0
        ).sum() if has_kr else 0
        
        word_count_en = df_res.loc[df_res["CHANGES"] == key, "Text_EN"].apply(
            lambda x: len(safe_str(x).split()) if safe_str(x) else 0
        ).sum() if has_en else 0
        
        word_count_cn = df_res.loc[df_res["CHANGES"] == key, "Text_CN"].apply(
            lambda x: len(safe_str(x).split()) if safe_str(x) else 0
        ).sum() if has_cn else 0
        
        summary_rows.append([key, counter[key], word_count_korean, word_count_kr, word_count_en, word_count_cn])
    
    if "Deleted Rows" in counter:
        summary_rows.append(["Deleted Rows", counter["Deleted Rows"], 0, 0, 0, 0])
    
    return pd.DataFrame(summary_rows, columns=["Metric", "Value", "Word Count (Korean)", "Word Count KR", "Word Count EN", "Word Count CN"])

# =========================================================================== #
# MASTER FILE UPDATE FUNCTIONS - VERSION 1114 WITH 3 KEY SYSTEM
# =========================================================================== #
def process_master_file_update():
    log("\n" + "="*70)
    log("PROCESS MASTER FILE UPDATE")
    log("="*70)
    
    root = tk.Tk()
    root.withdraw()
    
    source = filedialog.askopenfilename(
        title="MASTER FILE UPDATE: Select SOURCE (Working Process output)",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    if not source:
        log("User cancelled – exiting.")
        root.destroy()
        return
    
    target = filedialog.askopenfilename(
        title="MASTER FILE UPDATE: Select TARGET (Master File to update)",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    if not target:
        log("User cancelled – exiting.")
        root.destroy()
        return
    
    root.destroy()
    
    try:
        log(f"Reading SOURCE: {os.path.basename(source)}")
        df_source = safe_read_excel(source, header=0, dtype=str)
        log(f"  → {len(df_source):,} rows, {df_source.shape[1]} columns")
        
        log(f"Reading TARGET: {os.path.basename(target)}")
        df_target = safe_read_excel(target, header=0, dtype=str)
        log(f"  → {len(df_target):,} rows, {df_target.shape[1]} columns")
        
        # STEP 1: IDENTIFY TARGET STRUCTURE
        target_columns = list(df_target.columns)
        log(f"\nTARGET structure: {len(target_columns)} columns")
        
        # Define output structure (TARGET + CHANGES + Importance)
        output_structure = target_columns.copy()
        if "CHANGES" not in output_structure:
            output_structure.append("CHANGES")
        if COL_IMPORTANCE not in output_structure:
            output_structure.append(COL_IMPORTANCE)
        
        log(f"Output structure: {len(output_structure)} columns (all sheets will use this)")
        
        log("\nNormalizing STATUS columns...")
        df_source = normalize_dataframe_status(df_source)
        df_target = normalize_dataframe_status(df_target)
        
        if COL_IMPORTANCE not in df_source.columns:
            log("Warning: SOURCE has no 'Importance' column - treating all rows as High")
            df_source[COL_IMPORTANCE] = "High"
        
        # STEP 2: SEPARATE HIGH AND LOW IMPORTANCE
        log("\nSeparating SOURCE by Importance...")
        df_high = df_source[df_source[COL_IMPORTANCE].str.lower() == "high"].copy()
        df_low = df_source[df_source[COL_IMPORTANCE].str.lower() != "high"].copy()
        log(f"  → High importance: {len(df_high):,} rows")
        log(f"  → Low importance: {len(df_low):,} rows")
        
        # STEP 3: BUILD LOOKUPS (4-TIER KEY SYSTEM)
        log("\nBuilding lookups with 4-tier key system...")
        
        # SOURCE High lookup - 3 keys
        source_high_lookup = {}
        source_high_lookup_cg = {}
        source_high_lookup_es = {}  # NEW
        source_high_lookup_cs = {}  # NEW 4th key
        for _, row in df_high.iterrows():
            key_cw = (row[COL_SEQUENCE], row[COL_EVENTNAME])
            key_cg = (row[COL_SEQUENCE], row[COL_STRORIGIN])
            key_es = (row[COL_EVENTNAME], row[COL_STRORIGIN])
            key_cs = (row[COL_CASTINGKEY], row[COL_SEQUENCE])  # NEW 4th key
            source_high_lookup[key_cw] = row.to_dict()
            if key_cg not in source_high_lookup_cg:
                source_high_lookup_cg[key_cg] = row[COL_EVENTNAME]
            if key_es not in source_high_lookup_es:
                source_high_lookup_es[key_es] = row.to_dict()
            if key_cs not in source_high_lookup_cs:
                source_high_lookup_cs[key_cs] = row.to_dict()  # NEW 4th key
        
        # SOURCE Low lookup - 3 keys
        source_low_lookup = {}
        source_low_lookup_cg = {}
        source_low_lookup_es = {}  # NEW
        source_low_lookup_cs = {}  # NEW 4th key
        for _, row in df_low.iterrows():
            key_cw = (row[COL_SEQUENCE], row[COL_EVENTNAME])
            key_cg = (row[COL_SEQUENCE], row[COL_STRORIGIN])
            key_es = (row[COL_EVENTNAME], row[COL_STRORIGIN])
            key_cs = (row[COL_CASTINGKEY], row[COL_SEQUENCE])  # NEW 4th key
            source_low_lookup[key_cw] = row.to_dict()
            if key_cg not in source_low_lookup_cg:
                source_low_lookup_cg[key_cg] = row[COL_EVENTNAME]
            if key_es not in source_low_lookup_es:
                source_low_lookup_es[key_es] = row.to_dict()
            if key_cs not in source_low_lookup_cs:
                source_low_lookup_cs[key_cs] = row.to_dict()  # NEW 4th key
        
        # TARGET lookup - 3 keys
        target_lookup = {}
        target_lookup_cg = {}
        target_lookup_es = {}  # NEW
        target_lookup_cs = {}  # NEW 4th key
        for _, row in df_target.iterrows():
            key_cw = (row[COL_SEQUENCE], row[COL_EVENTNAME])
            key_cg = (row[COL_SEQUENCE], row[COL_STRORIGIN])
            key_es = (row[COL_EVENTNAME], row[COL_STRORIGIN])
            key_cs = (row[COL_CASTINGKEY], row[COL_SEQUENCE])  # NEW 4th key
            target_lookup[key_cw] = row.to_dict()
            if key_cg not in target_lookup_cg:
                target_lookup_cg[key_cg] = row[COL_EVENTNAME]
            if key_es not in target_lookup_es:
                target_lookup_es[key_es] = row.to_dict()
            if key_cs not in target_lookup_cs:
                target_lookup_cs[key_cs] = row.to_dict()  # NEW 4th key
        
        log(f"  → SOURCE High indexed: {len(source_high_lookup):,} rows")
        log(f"  → SOURCE Low indexed: {len(source_low_lookup):,} rows")
        log(f"  → TARGET indexed: {len(target_lookup):,} rows")
        
        # STEP 4: PROCESS HIGH IMPORTANCE ROWS
        log("\nProcessing HIGH importance rows...")
        high_rows = []
        high_counter = {}
        
        for idx, source_row in df_high.iterrows():
            key_cw = (source_row[COL_SEQUENCE], source_row[COL_EVENTNAME])
            key_cg = (source_row[COL_SEQUENCE], source_row[COL_STRORIGIN])
            key_es = (source_row[COL_EVENTNAME], source_row[COL_STRORIGIN])
            key_cs = (source_row[COL_CASTINGKEY], source_row[COL_SEQUENCE])  # NEW 4th key

            # Copy SOURCE values to output structure
            output_row = {}
            for col in output_structure:
                if col in source_row.index:
                    output_row[col] = safe_str(source_row[col])
                else:
                    output_row[col] = ""

            # Determine change type using 4-tier key system
            target_row = None

            # Stage 1: Direct match
            if key_cw in target_lookup:
                change_type = safe_str(source_row.get("CHANGES", "Edited"))
                target_row = target_lookup[key_cw]

            # Stage 2: StrOrigin+Sequence match - VERIFY with Key 4
            elif key_cg in target_lookup_cg:
                # Check if same character (Key 4 verification)
                if key_cs in target_lookup_cs:
                    # Same character → EventName Change
                    change_type = safe_str(source_row.get("CHANGES", "Edited"))
                    old_eventname = target_lookup_cg[key_cg]
                    target_row = target_lookup.get((source_row[COL_SEQUENCE], old_eventname))
                else:
                    # Different character → New Row (duplicate StrOrigin)
                    change_type = "New Row"
                    target_row = None

            # Stage 3: SequenceName changed
            elif key_es in target_lookup_es:
                change_type = "SequenceName Change"
                target_row = target_lookup_es[key_es]

            # Stage 4: New row
            else:
                change_type = "New Row"
                target_row = None
            
            # EXCEPTION: ONLY for these 2 specific cases, preserve TARGET TimeFrame
            if target_row and change_type in ["TimeFrame Change", "EventName+TimeFrame Change", "TimeFrame+EventName Change"]:
                if COL_STARTFRAME in target_row:
                    output_row[COL_STARTFRAME] = safe_str(target_row[COL_STARTFRAME])
                if COL_ENDFRAME in target_row:
                    output_row[COL_ENDFRAME] = safe_str(target_row[COL_ENDFRAME])
            
            output_row["CHANGES"] = change_type
            output_row[COL_IMPORTANCE] = "High"
            
            high_rows.append(output_row)
            high_counter[change_type] = high_counter.get(change_type, 0) + 1
        
        df_high_output = pd.DataFrame(high_rows, columns=output_structure)
        log(f"  → Processed {len(df_high_output):,} HIGH rows")
        
        # STEP 5: PROCESS LOW IMPORTANCE ROWS
        log("\nProcessing LOW importance rows...")
        low_rows = []
        low_counter = {}

        for idx, source_row in df_low.iterrows():
            key_cw = (source_row[COL_SEQUENCE], source_row[COL_EVENTNAME])
            key_cg = (source_row[COL_SEQUENCE], source_row[COL_STRORIGIN])
            key_es = (source_row[COL_EVENTNAME], source_row[COL_STRORIGIN])
            key_cs = (source_row[COL_CASTINGKEY], source_row[COL_SEQUENCE])  # NEW 4th key

            # Determine change type using 4-key lookup
            target_row = None

            # Stage 1: Direct match (SequenceName + EventName)
            if key_cw in target_lookup:
                change_type = safe_str(source_row.get("CHANGES", "Edited"))
                target_row = target_lookup[key_cw]

            # Stage 2: StrOrigin+Sequence match - VERIFY with Key 4 (4-Tier System)
            elif key_cg in target_lookup_cg:
                # Check if same character (Key 4 verification)
                if key_cs in target_lookup_cs:
                    # Same character → EventName Change
                    change_type = safe_str(source_row.get("CHANGES", "Edited"))
                    old_eventname = target_lookup_cg[key_cg]
                    target_row = target_lookup.get((source_row[COL_SEQUENCE], old_eventname))
                else:
                    # Different character → New Row (duplicate StrOrigin case)
                    change_type = "New Row"
                    target_row = None

            # Stage 3: SequenceName changed (EventName + StrOrigin match)
            elif key_es in target_lookup_es:
                change_type = "SequenceName Change"
                target_row = target_lookup_es[key_es]

            # Stage 4: New row (no keys match)
            else:
                change_type = "New Row"
                target_row = None

            # LOW IMPORTANCE LOGIC: For existing rows, preserve TARGET data
            output_row = {}
            if change_type != "New Row" and target_row is not None:
                # Existing row: Use TARGET data (preserve original)
                for col in output_structure:
                    if col in target_row:
                        output_row[col] = safe_str(target_row[col])
                    elif col in source_row.index:
                        output_row[col] = safe_str(source_row[col])
                    else:
                        output_row[col] = ""
            else:
                # New row: Use SOURCE data (will be deleted in post-process)
                for col in output_structure:
                    if col in source_row.index:
                        output_row[col] = safe_str(source_row[col])
                    else:
                        output_row[col] = ""

            # Always set CHANGES and Importance from classification
            output_row["CHANGES"] = change_type
            output_row[COL_IMPORTANCE] = "Low"

            low_rows.append(output_row)
            low_counter[change_type] = low_counter.get(change_type, 0) + 1

        df_low_output = pd.DataFrame(low_rows, columns=output_structure) if low_rows else pd.DataFrame(columns=output_structure)
        log(f"  → Processed {len(df_low_output):,} LOW rows")

        # STEP 5b: FILTER OUT LOW + NEW ROW (Post-process deletion)
        log("\nFiltering LOW importance NEW rows...")
        original_low_count = len(df_low_output)
        df_low_output = df_low_output[df_low_output["CHANGES"] != "New Row"]
        filtered_count = original_low_count - len(df_low_output)
        if filtered_count > 0:
            log(f"  → Removed {filtered_count:,} LOW importance NEW rows")
            low_counter["Deleted (LOW+New)"] = filtered_count
        log(f"  → Final LOW rows: {len(df_low_output):,}")
        
        # STEP 6: FIND DELETED ROWS (4-KEY SYSTEM)
        log("\nFinding deleted rows with 4-key system...")
        deleted_rows = []

        # Build all current keys (4 types)
        source_all_keys_cw = set(source_high_lookup.keys()) | set(source_low_lookup.keys())
        source_all_keys_cg = set(source_high_lookup_cg.keys()) | set(source_low_lookup_cg.keys())
        source_all_keys_es = set(source_high_lookup_es.keys()) | set(source_low_lookup_es.keys())
        source_all_keys_cs = set(source_high_lookup_cs.keys()) | set(source_low_lookup_cs.keys())  # NEW 4th key

        for _, target_row in df_target.iterrows():
            key_cw = (target_row[COL_SEQUENCE], target_row[COL_EVENTNAME])
            key_cg = (target_row[COL_SEQUENCE], target_row[COL_STRORIGIN])
            key_es = (target_row[COL_EVENTNAME], target_row[COL_STRORIGIN])
            key_cs = (target_row[COL_CASTINGKEY], target_row[COL_SEQUENCE])  # NEW 4th key

            # Only mark as deleted if ALL 4 keys are missing
            if (key_cw not in source_all_keys_cw) and \
               (key_cg not in source_all_keys_cg) and \
               (key_es not in source_all_keys_es) and \
               (key_cs not in source_all_keys_cs):  # NEW 4th key check
                output_row = {}
                for col in output_structure:
                    if col in target_row.index:
                        output_row[col] = safe_str(target_row[col])
                    else:
                        output_row[col] = ""
                
                output_row["CHANGES"] = ""
                output_row[COL_IMPORTANCE] = ""
                
                deleted_rows.append(output_row)
        
        df_deleted = pd.DataFrame(deleted_rows, columns=output_structure) if deleted_rows else pd.DataFrame(columns=output_structure)
        log(f"  → Found {len(df_deleted):,} deleted rows")
        
        # STEP 7: CREATE SUMMARY
        log("\nCreating summary report...")
        
        total_counter = {}
        for k, v in high_counter.items():
            total_counter[k] = total_counter.get(k, 0) + v
        for k, v in low_counter.items():
            total_counter[k] = total_counter.get(k, 0) + v
        if len(df_deleted) > 0:
            total_counter["Deleted Rows"] = len(df_deleted)
        
        summary_rows = [
            ["MASTER FILE UPDATE SUMMARY", "", "", ""],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", ""],
            ["Source file", os.path.basename(source), "", ""],
            ["Target file", os.path.basename(target), "", ""],
            ["", "", "", ""],
            ["RESULT COUNTS", "Count", "", ""],
            ["Main Sheet (High)", len(df_high_output), "", ""],
            ["Low Importance Sheet", len(df_low_output), "", ""],
            ["Deleted Rows Sheet", len(df_deleted), "", ""],
            ["", "", "", ""],
            ["CHANGE BREAKDOWN", "Count", "", ""],
        ]
        
        for change_type in sorted(total_counter.keys()):
            summary_rows.append([f"  {change_type}", total_counter[change_type], "", ""])
        
        df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value", "Col3", "Col4"])
        
        df_history = create_master_file_update_history_sheet()
        
        # STEP 8: WRITE OUTPUT
        script_dir = get_script_dir()
        out_filename = "MasterFile_Updated_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        out_path = os.path.join(script_dir, out_filename)
        log(f"\nWriting results to: {out_filename}")
        
        # Load TARGET for formatting
        log("Loading TARGET workbook for formatting...")
        wb_target = load_workbook(target)
        ws_target = wb_target.active
        
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df_high_output.to_excel(writer, sheet_name="Main Sheet (High)", index=False)
            df_low_output.to_excel(writer, sheet_name="Low Importance", index=False)
            df_deleted.to_excel(writer, sheet_name="Deleted Rows", index=False)
            df_history.to_excel(writer, sheet_name="📅 Update History", index=False, header=False)
            df_summary.to_excel(writer, sheet_name="Summary Report", index=False, header=True)
            
            wb = writer.book
            
            # FAST FORMATTING: Only copy column widths and header row
            for sheet_name in ["Main Sheet (High)", "Low Importance", "Deleted Rows"]:
                ws = wb[sheet_name]
                
                log(f"Applying formatting to {sheet_name}...")
                
                # Copy column widths
                for col_idx in range(1, min(ws_target.max_column + 1, ws.max_column + 1)):
                    col_letter = get_column_letter(col_idx)
                    if col_letter in ws_target.column_dimensions:
                        ws.column_dimensions[col_letter].width = ws_target.column_dimensions[col_letter].width
                
                # Copy header row formatting
                for col_idx in range(1, min(ws_target.max_column + 1, ws.max_column + 1)):
                    source_cell = ws_target.cell(row=1, column=col_idx)
                    target_cell = ws.cell(row=1, column=col_idx)
                    
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.alignment = copy(source_cell.alignment)
                
                # Apply CHANGES column coloring
                apply_direct_coloring(ws, is_master=False)
            
            format_update_history_sheet(wb["📅 Update History"])
            widen_summary_columns(wb["Summary Report"])
        
        wb_target.close()
        
        add_master_file_update_record(out_filename, source, target, total_counter, len(df_high_output))
        
        log(f"✓ File saved: {out_path}")
        log("="*70)
        
        summary_msg = "Master File Update completed successfully!\n\n"
        summary_msg += f"Output file:\n{out_path}\n\n"
        summary_msg += "Results:\n"
        summary_msg += f"  Main Sheet (High): {len(df_high_output):,} rows\n"
        summary_msg += f"  Low Importance: {len(df_low_output):,} rows\n"
        summary_msg += f"  Deleted Rows: {len(df_deleted):,} rows\n\n"
        summary_msg += "Change Summary:\n"
        for change_type, count in sorted(total_counter.items()):
            summary_msg += f"  {change_type}: {count:,}\n"
        
        messagebox.showinfo("MASTER FILE UPDATE Complete", summary_msg)
        
    except Exception as exc:
        log(f"FATAL ERROR: {exc}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Error", f"Something went wrong:\n\n{exc}")

def create_master_file_update_history_sheet():
    history = load_update_history("master")
    
    history_data = [
        ["MASTER FILE UPDATE - UPDATE HISTORY"],
        [""],
        ["This update history is tracked in: " + MASTER_HISTORY_FILE],
        ["Use 'View Update History' button in main GUI for complete details"],
        [""],
    ]
    
    if not history["updates"]:
        history_data.append(["No updates recorded yet"])
        history_data.append([""])
    else:
        latest = history["updates"][-1]
        history_data.append(["LATEST UPDATE"])
        history_data.append([f"Timestamp: {latest['timestamp']}"])
        history_data.append([f"Output File: {latest['output_file']}"])
        history_data.append([f"Source: {latest['source_file']}"])
        history_data.append([f"Target: {latest['target_file']}"])
        history_data.append([""])
        history_data.append([f"Total Rows: {latest['statistics']['total_rows']:,}"])
        history_data.append([""])
        history_data.append(["Change Breakdown:"])
        for key, value in latest['statistics'].items():
            if key != 'total_rows':
                history_data.append([f"  {key}: {value:,}"])
        history_data.append([""])
        history_data.append([f"Total Updates Recorded: {len(history['updates'])}"])
    
    history_data.append([""])
    history_data.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    
    return pd.DataFrame(history_data, columns=["Update History"])

# =========================================================================== #
# EXCEL FORMATTING FUNCTIONS
# =========================================================================== #
def apply_direct_coloring(ws, is_master=False, changed_columns_map=None):
    changes_col_idx = None
    status_col_indices = {}
    char_group_col_indices = {}
    
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "CHANGES":
            changes_col_idx = idx
        elif is_master:
            if cell.value in ["STATUS_KR", "STATUS_EN", "STATUS_CN"]:
                status_col_indices[cell.value] = idx
        else:
            if cell.value and cell.value.upper() == "STATUS":
                status_col_indices["STATUS"] = idx
        
        if cell.value in CHAR_GROUP_COLS:
            char_group_col_indices[cell.value] = idx
    
    changes_fills = {
        "StrOrigin Change": PatternFill(start_color="FFD580", fill_type="solid"),
        "Desc Change": PatternFill(start_color="E1D5FF", fill_type="solid"),
        "TimeFrame Change": PatternFill(start_color="FF9999", fill_type="solid"),
        "EventName Change": PatternFill(start_color="FFFF99", fill_type="solid"),
        "SequenceName Change": PatternFill(start_color="B3E5FC", fill_type="solid"),  # NEW - Light Blue
        "StrOrigin+Desc Change": PatternFill(start_color="FFA07A", fill_type="solid"),
        "StrOrigin+TimeFrame Change": PatternFill(start_color="FFB6C1", fill_type="solid"),
        "Desc+TimeFrame Change": PatternFill(start_color="DDA0DD", fill_type="solid"),
        "StrOrigin+Desc+TimeFrame Change": PatternFill(start_color="F08080", fill_type="solid"),
        "EventName+Desc Change": PatternFill(start_color="F0E68C", fill_type="solid"),
        "EventName+TimeFrame Change": PatternFill(start_color="FFDAB9", fill_type="solid"),
        "EventName+Desc+TimeFrame Change": PatternFill(start_color="FFD8A8", fill_type="solid"),
        "Character Group Change": PatternFill(start_color="87CEFA", fill_type="solid"),
        "New Row": PatternFill(start_color="90EE90", fill_type="solid"),
        "No Relevant Change": PatternFill(start_color="D3D3D3", fill_type="solid"),
        "No Change": PatternFill(start_color="E8E8E8", fill_type="solid"),
    }
    
    char_group_change_fill = PatternFill(start_color="FFD700", fill_type="solid")
    
    status_fills = {
        "RECORDED": PatternFill(start_color="90EE90", fill_type="solid"),
        "POLISHED": PatternFill(start_color="E6D5FF", fill_type="solid"),
        "RE-RECORD": PatternFill(start_color="FFB3B3", fill_type="solid"),
        "RERECORD": PatternFill(start_color="FFB3B3", fill_type="solid"),
        "RE-RECORDED": PatternFill(start_color="C5E8C5", fill_type="solid"),
        "RERECORDED": PatternFill(start_color="C5E8C5", fill_type="solid"),
        "PREVIOUSLY RECORDED": PatternFill(start_color="FFFFE0", fill_type="solid"),
        "FINAL": PatternFill(start_color="87CEEB", fill_type="solid"),
        "SHIPPED": PatternFill(start_color="87CEEB", fill_type="solid"),
        "SPEC-OUT": PatternFill(start_color="FFC0CB", fill_type="solid"),
        "CHECK": PatternFill(start_color="FFE4B5", fill_type="solid"),
        "NEED CHECK": PatternFill(start_color="FFB3B3", fill_type="solid"),
        "전달 완료": PatternFill(start_color="FFFFE0", fill_type="solid"),
        "녹음 완료": PatternFill(start_color="90EE90", fill_type="solid"),
        "재녹음 필요": PatternFill(start_color="FFB3B3", fill_type="solid"),
        "재녹음 완료": PatternFill(start_color="C5E8C5", fill_type="solid"),
        "준비 중": PatternFill(start_color="E6D5FF", fill_type="solid"),
        "확인 필요": PatternFill(start_color="FFE4B5", fill_type="solid"),
        "已传达": PatternFill(start_color="FFFFE0", fill_type="solid"),
        "已录音": PatternFill(start_color="90EE90", fill_type="solid"),
        "需补录": PatternFill(start_color="FFB3B3", fill_type="solid"),
        "已补录": PatternFill(start_color="C5E8C5", fill_type="solid"),
        "准备中": PatternFill(start_color="E6D5FF", fill_type="solid"),
        "需要确认": PatternFill(start_color="FFE4B5", fill_type="solid"),
    }
    
    header_fill = PatternFill(start_color="ADD8E6", fill_type="solid")
    
    for cell in ws[1]:
        cell.fill = header_fill
    
    if changes_col_idx:
        colored_count = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, 
                                min_col=changes_col_idx, max_col=changes_col_idx), start=0):
            cell = row[0]
            cell_value = cell.value
            if cell_value in changes_fills:
                cell.fill = changes_fills[cell_value]
                colored_count += 1
                
                if cell_value == "Character Group Change" and changed_columns_map and row_idx in changed_columns_map:
                    changed_cols = changed_columns_map[row_idx]
                    for col_name in changed_cols:
                        if col_name in char_group_col_indices:
                            col_idx = char_group_col_indices[col_name]
                            char_cell = ws.cell(row=row_idx + 2, column=col_idx)
                            char_cell.fill = char_group_change_fill
        
        ws.column_dimensions[get_column_letter(changes_col_idx)].width = 40
    
    for status_col_name, status_col_idx in status_col_indices.items():
        colored_count = 0
        unknown_statuses = {}
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, 
                                min_col=status_col_idx, max_col=status_col_idx):
            cell = row[0]
            cell_value = cell.value
            
            if cell_value and str(cell_value).strip():
                if cell_value in status_fills:
                    cell.fill = status_fills[cell_value]
                    colored_count += 1
                else:
                    if cell_value not in unknown_statuses:
                        unknown_statuses[cell_value] = PatternFill(
                            start_color=generate_color_for_value(cell_value), 
                            fill_type="solid"
                        )
                    cell.fill = unknown_statuses[cell_value]
                    colored_count += 1
        
        ws.column_dimensions[get_column_letter(status_col_idx)].width = 25
    
    ws.sheet_view.showGridLines = True
    ws.auto_filter.ref = ws.dimensions

def widen_summary_columns(ws):
    ws.column_dimensions[get_column_letter(1)].width = 50
    ws.column_dimensions[get_column_letter(2)].width = 25
    ws.column_dimensions[get_column_letter(3)].width = 25
    ws.column_dimensions[get_column_letter(4)].width = 25
    ws.column_dimensions[get_column_letter(5)].width = 25
    ws.column_dimensions[get_column_letter(6)].width = 25

def format_update_history_sheet(ws):
    ws.column_dimensions[get_column_letter(1)].width = 80
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            if cell.value:
                content = str(cell.value)
                
                if "UPDATE HISTORY" in content or "LATEST UPDATE" in content:
                    cell.font = Font(bold=True, size=14, color="000080")
                    cell.fill = PatternFill(start_color="FFFFCC", fill_type="solid")
                
                elif "✓ UPDATED" in content:
                    cell.font = Font(bold=True, color="006600")
                    cell.fill = PatternFill(start_color="CCFFCC", fill_type="solid")
                
                elif "○ Preserved" in content:
                    cell.font = Font(italic=True, color="666666")
                    cell.fill = PatternFill(start_color="F0F0F0", fill_type="solid")
    
    ws.sheet_view.showGridLines = False

# =========================================================================== #
# MAIN PROCESS FUNCTIONS
# =========================================================================== #
def process_raw_vrs_check():
    log("\n" + "="*70)
    log("PROCESS RAW VRS CHECK")
    log("="*70)
    
    root = tk.Tk()
    root.withdraw()
    prev = filedialog.askopenfilename(
        title="RAW CHECK: Select PREVIOUS Excel file",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    if not prev:
        log("User cancelled – exiting.")
        root.destroy()
        return
    
    curr = filedialog.askopenfilename(
        title="RAW CHECK: Select CURRENT Excel file",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    if not curr:
        log("User cancelled – exiting.")
        root.destroy()
        return
    
    root.destroy()
    
    try:
        log(f"Reading PREVIOUS: {os.path.basename(prev)}")
        df_prev = safe_read_excel(prev, header=0, dtype=str)
        log(f"  → {len(df_prev):,} rows, {df_prev.shape[1]} columns")
        
        log(f"Reading CURRENT: {os.path.basename(curr)}")
        df_curr = safe_read_excel(curr, header=0, dtype=str)
        log(f"  → {len(df_curr):,} rows, {df_curr.shape[1]} columns")
        
        log("Normalizing STATUS columns...")
        df_prev = normalize_dataframe_status(df_prev)
        df_curr = normalize_dataframe_status(df_curr)
        
        log("Building lookup dictionaries with 4-key system...")
        prev_lookup_cw, prev_lookup_cg, prev_lookup_es, prev_lookup_cs = build_lookups(df_prev)
        log(f"  → Indexed {len(prev_lookup_cw):,} unique previous rows")

        log("Comparing rows...")
        changes, previous_strorigins, changed_columns_map, counter = compare_rows(
            df_curr, prev_lookup_cw, prev_lookup_cg, prev_lookup_es, prev_lookup_cs
        )
        
        log("Finding deleted rows...")
        df_deleted = find_deleted_rows(df_prev, df_curr)
        counter["Deleted Rows"] = len(df_deleted)
        log(f"  → Found {len(df_deleted):,} deleted rows")
        
        df_res = df_curr.copy()
        df_res["CHANGES"] = changes
        df_res[COL_PREVIOUS_STRORIGIN] = previous_strorigins
        
        log("Generating CastingKey column...")
        casting_keys = []
        for idx, row in df_res.iterrows():
            casting_key = generate_casting_key(
                row.get(COL_CHARACTERKEY, ""),
                row.get(COL_DIALOGVOICE, ""),
                row.get(COL_SPEAKER_GROUPKEY, ""),
                row.get("DialogType", "")
            )
            casting_keys.append(casting_key)
        df_res[COL_CASTINGKEY] = casting_keys
        log(f"  → Generated CastingKey for {len(casting_keys):,} rows")
        
        log("Filtering output columns...")
        df_res = filter_output_columns(df_res, OUTPUT_COLUMNS_RAW)
        
        df_summary = create_raw_summary(counter, prev, curr, df_res)
        
        script_dir = get_script_dir()
        out_filename = os.path.splitext(os.path.basename(curr))[0] + "_diff.xlsx"
        out_path = os.path.join(script_dir, out_filename)
        log(f"Writing results to: {os.path.basename(out_path)}")
        
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df_res.to_excel(writer, sheet_name="Comparison", index=False)
            if not df_deleted.empty:
                df_deleted_filtered = filter_output_columns(df_deleted, OUTPUT_COLUMNS_RAW)
                df_deleted_filtered.to_excel(writer, sheet_name="Deleted Rows", index=False)
            df_summary.to_excel(writer, sheet_name="Summary Report", index=False, header=True)
            
            wb = writer.book
            apply_direct_coloring(wb["Comparison"], is_master=False, changed_columns_map=changed_columns_map)
            widen_summary_columns(wb["Summary Report"])
        
        log(f"✓ File saved: {out_path}")
        log("="*70)
        
        messagebox.showinfo(
            "RAW VRS CHECK Complete",
            f"Process completed successfully!\n\nOutput file:\n{out_path}"
        )
        
    except Exception as exc:
        log(f"FATAL ERROR: {exc}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Error", f"Something went wrong:\n\n{exc}")

def process_working_vrs_check():
    log("\n" + "="*70)
    log("PROCESS WORKING VRS CHECK")
    log("="*70)
    
    root = tk.Tk()
    root.withdraw()
    prev = filedialog.askopenfilename(
        title="WORKING CHECK: Select PREVIOUS Excel file",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    if not prev:
        log("User cancelled – exiting.")
        root.destroy()
        return
    
    curr = filedialog.askopenfilename(
        title="WORKING CHECK: Select CURRENT Excel file",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    if not curr:
        log("User cancelled – exiting.")
        root.destroy()
        return
    
    root.destroy()
    
    try:
        log(f"Reading PREVIOUS: {os.path.basename(prev)}")
        df_prev = safe_read_excel(prev, header=0, dtype=str)
        log(f"  → {len(df_prev):,} rows, {df_prev.shape[1]} columns")
        
        log(f"Reading CURRENT: {os.path.basename(curr)}")
        df_curr = safe_read_excel(curr, header=0, dtype=str)
        log(f"  → {len(df_curr):,} rows, {df_curr.shape[1]} columns")
        
        log("Normalizing STATUS columns...")
        df_prev = normalize_dataframe_status(df_prev)
        df_curr = normalize_dataframe_status(df_curr)
        log("  → STATUS columns normalized")
        
        prev_lookup_cw, prev_lookup_cg, prev_lookup_es, prev_lookup_cs = build_working_lookups(df_prev, "PREVIOUS")

        df_result, counter = process_working_comparison(df_curr, prev_lookup_cw, prev_lookup_cg, prev_lookup_es, prev_lookup_cs)
        
        df_deleted = find_working_deleted_rows(df_prev, df_curr)
        if not df_deleted.empty:
            counter["Deleted Rows"] = len(df_deleted)
        
        log("Filtering output columns...")
        df_result = filter_output_columns(df_result)
        log(f"  → Output contains {len(df_result.columns)} columns")
        
        log("Creating summary report...")
        df_summary = create_working_summary(counter, prev, curr, df_result)
        df_history = create_working_update_history_sheet()
        
        script_dir = get_script_dir()
        out_filename = os.path.splitext(os.path.basename(curr))[0] + "_WorkTransform.xlsx"
        out_path = os.path.join(script_dir, out_filename)
        log(f"Writing results to: {out_filename}")
        
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df_result.to_excel(writer, sheet_name="Work Transform", index=False)
            
            df_history.to_excel(writer, sheet_name="📅 Update History", index=False, header=False)
            
            if not df_deleted.empty:
                df_deleted_filtered = filter_output_columns(df_deleted)
                df_deleted_filtered.to_excel(writer, sheet_name="Deleted Rows", index=False)
                log(f"  → Created 'Deleted Rows' sheet with {len(df_deleted)} rows")
            
            df_summary.to_excel(writer, sheet_name="Summary Report", index=False, header=True)
            
            wb = writer.book
            apply_direct_coloring(wb["Work Transform"], is_master=False)
            format_update_history_sheet(wb["📅 Update History"])
            widen_summary_columns(wb["Summary Report"])
        
        add_working_update_record(out_filename, prev, curr, counter, len(df_result))
        
        log(f"✓ File saved: {out_path}")
        log("="*70)
        
        summary_msg = "Process completed successfully!\n\n"
        summary_msg += f"Output file:\n{out_path}\n\n"
        summary_msg += "Change Summary:\n"
        for change_type, count in sorted(counter.items()):
            summary_msg += f"  {change_type}: {count:,}\n"
        
        messagebox.showinfo("WORKING VRS CHECK Complete", summary_msg)
        
    except Exception as exc:
        log(f"FATAL ERROR: {exc}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Error", f"Something went wrong:\n\n{exc}")

def process_all_language_check():
    log("\n" + "="*70)
    log("PROCESS ALL LANGUAGE CHECK")
    log("="*70)
    
    try:
        log("Auto-detecting files from Previous/ and Current/ folders...")
        curr_kr, curr_en, curr_cn, prev_kr, prev_en, prev_cn = find_alllang_files()
        
        log("\nCURRENT FILES (complete base - required):")
        log(f"  ✓ KR: {os.path.basename(curr_kr)}")
        log(f"  ✓ EN: {os.path.basename(curr_en)}")
        log(f"  ✓ CN: {os.path.basename(curr_cn)}")
        
        log("\nPREVIOUS FILES (selective updates - flexible):")
        log(f"  {'✓' if prev_kr else '○'} KR: {os.path.basename(prev_kr) if prev_kr else 'Not present - will preserve'}")
        log(f"  {'✓' if prev_en else '○'} EN: {os.path.basename(prev_en) if prev_en else 'Not present - will preserve'}")
        log(f"  {'✓' if prev_cn else '○'} CN: {os.path.basename(prev_cn) if prev_cn else 'Not present - will preserve'}")
        
        df_curr = merge_current_files(curr_kr, curr_en, curr_cn)
        
        lookup_kr = {}
        lookup_en = {}
        lookup_cn = {}
        lookup_cg_kr = {}
        lookup_es_kr = {}
        lookup_cs_kr = {}  # NEW 4th key

        has_kr = prev_kr is not None
        has_en = prev_en is not None
        has_cn = prev_cn is not None

        if has_kr:
            log(f"\nReading KR Previous: {os.path.basename(prev_kr)}")
            df_kr = safe_read_excel(prev_kr, header=0, dtype=str)
            df_kr = normalize_dataframe_status(df_kr)
            log(f"  → {len(df_kr):,} rows")
            lookup_kr, lookup_cg_kr, lookup_es_kr, lookup_cs_kr = build_working_lookups(df_kr, "KR PREVIOUS")

        if has_en:
            log(f"\nReading EN Previous: {os.path.basename(prev_en)}")
            df_en = safe_read_excel(prev_en, header=0, dtype=str)
            df_en = normalize_dataframe_status(df_en)
            log(f"  → {len(df_en):,} rows")
            lookup_en, _, _, _ = build_working_lookups(df_en, "EN PREVIOUS")

        if has_cn:
            log(f"\nReading CN Previous: {os.path.basename(prev_cn)}")
            df_cn = safe_read_excel(prev_cn, header=0, dtype=str)
            df_cn = normalize_dataframe_status(df_cn)
            log(f"  → {len(df_cn):,} rows")
            lookup_cn, _, _, _ = build_working_lookups(df_cn, "CN PREVIOUS")

        df_result, counter = process_alllang_comparison(
            df_curr, lookup_kr, lookup_en, lookup_cn, lookup_cg_kr, lookup_es_kr, lookup_cs_kr,
            has_kr, has_en, has_cn
        )
        
        if has_kr:
            df_kr_full = safe_read_excel(prev_kr, header=0, dtype=str)
            df_deleted = find_working_deleted_rows(df_kr_full, df_curr)
            if not df_deleted.empty:
                counter["Deleted Rows"] = len(df_deleted)
        else:
            df_deleted = pd.DataFrame()
        
        log("\nFiltering output columns...")
        df_result = filter_output_columns(df_result, OUTPUT_COLUMNS_MASTER)
        log(f"  → Output contains {len(df_result.columns)} columns")
        
        log("Creating summary report...")
        df_summary = create_alllang_summary(counter, prev_kr, prev_en, prev_cn, curr_kr, df_result, has_kr, has_en, has_cn)
        df_history = create_alllang_update_history_sheet()
        
        script_dir = get_script_dir()
        out_filename = "AllLanguage_VRS_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        out_path = os.path.join(script_dir, out_filename)
        log(f"\nWriting results to: {out_filename}")
        
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df_result.to_excel(writer, sheet_name="All Language Transform", index=False)
            
            df_history.to_excel(writer, sheet_name="📅 Update History", index=False, header=False)
            
            if not df_deleted.empty:
                df_deleted_filtered = filter_output_columns(df_deleted, OUTPUT_COLUMNS_MASTER)
                df_deleted_filtered.to_excel(writer, sheet_name="Deleted Rows", index=False)
                log(f"  → Created 'Deleted Rows' sheet with {len(df_deleted)} rows")
            
            df_summary.to_excel(writer, sheet_name="Summary Report", index=False, header=True)
            
            wb = writer.book
            apply_direct_coloring(wb["All Language Transform"], is_master=True)
            format_update_history_sheet(wb["📅 Update History"])
            widen_summary_columns(wb["Summary Report"])
        
        add_alllang_update_record(out_filename, prev_kr, prev_en, prev_cn, curr_kr, curr_en, curr_cn, 
                                 counter, len(df_result))
        
        log(f"✓ File saved: {out_path}")
        log("="*70)
        
        summary_msg = "All Language Check completed successfully!\n\n"
        summary_msg += f"Output file:\n{out_path}\n\n"
        summary_msg += "Languages Updated:\n"
        summary_msg += f"  KR: {'✓ UPDATED' if has_kr else '○ Preserved'}\n"
        summary_msg += f"  EN: {'✓ UPDATED' if has_en else '○ Preserved'}\n"
        summary_msg += f"  CN: {'✓ UPDATED' if has_cn else '○ Preserved'}\n\n"
        summary_msg += "Change Summary:\n"
        for change_type, count in sorted(counter.items()):
            summary_msg += f"  {change_type}: {count:,}\n"
        
        messagebox.showinfo("ALL LANGUAGE CHECK Complete", summary_msg)
        
    except Exception as exc:
        log(f"FATAL ERROR: {exc}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Error", f"Something went wrong:\n\n{exc}")

# =========================================================================== #
# THREADING WRAPPERS FOR GUI
# =========================================================================== #
def run_raw_process_thread(btn_raw, btn_working, btn_alllang, btn_master, btn_history, status_label):
    def run():
        btn_raw.config(state=tk.DISABLED)
        btn_working.config(state=tk.DISABLED)
        btn_alllang.config(state=tk.DISABLED)
        btn_master.config(state=tk.DISABLED)
        btn_history.config(state=tk.DISABLED)
        status_label.config(text="⏳ Processing Raw VRS Check...")
        
        try:
            process_raw_vrs_check()
        finally:
            btn_raw.config(state=tk.NORMAL)
            btn_working.config(state=tk.NORMAL)
            btn_alllang.config(state=tk.NORMAL)
            btn_master.config(state=tk.NORMAL)
            btn_history.config(state=tk.NORMAL)
            status_label.config(text="✓ Ready")
    
    thread = threading.Thread(target=run, daemon=True)
    thread.start()

def run_working_process_thread(btn_raw, btn_working, btn_alllang, btn_master, btn_history, status_label):
    def run():
        btn_raw.config(state=tk.DISABLED)
        btn_working.config(state=tk.DISABLED)
        btn_alllang.config(state=tk.DISABLED)
        btn_master.config(state=tk.DISABLED)
        btn_history.config(state=tk.DISABLED)
        status_label.config(text="⏳ Processing Working VRS Check...")
        
        try:
            process_working_vrs_check()
        finally:
            btn_raw.config(state=tk.NORMAL)
            btn_working.config(state=tk.NORMAL)
            btn_alllang.config(state=tk.NORMAL)
            btn_master.config(state=tk.NORMAL)
            btn_history.config(state=tk.NORMAL)
            status_label.config(text="✓ Ready")
    
    thread = threading.Thread(target=run, daemon=True)
    thread.start()

def run_alllang_process_thread(btn_raw, btn_working, btn_alllang, btn_master, btn_history, status_label):
    def run():
        btn_raw.config(state=tk.DISABLED)
        btn_working.config(state=tk.DISABLED)
        btn_alllang.config(state=tk.DISABLED)
        btn_master.config(state=tk.DISABLED)
        btn_history.config(state=tk.DISABLED)
        status_label.config(text="⏳ Processing All Language Check...")
        
        try:
            process_all_language_check()
        finally:
            btn_raw.config(state=tk.NORMAL)
            btn_working.config(state=tk.NORMAL)
            btn_alllang.config(state=tk.NORMAL)
            btn_master.config(state=tk.NORMAL)
            btn_history.config(state=tk.NORMAL)
            status_label.config(text="✓ Ready")
    
    thread = threading.Thread(target=run, daemon=True)
    thread.start()

def run_master_file_update_thread(btn_raw, btn_working, btn_alllang, btn_master, btn_history, status_label):
    def run():
        btn_raw.config(state=tk.DISABLED)
        btn_working.config(state=tk.DISABLED)
        btn_alllang.config(state=tk.DISABLED)
        btn_master.config(state=tk.DISABLED)
        btn_history.config(state=tk.DISABLED)
        status_label.config(text="⏳ Processing Master File Update...")
        
        try:
            process_master_file_update()
        finally:
            btn_raw.config(state=tk.NORMAL)
            btn_working.config(state=tk.NORMAL)
            btn_alllang.config(state=tk.NORMAL)
            btn_master.config(state=tk.NORMAL)
            btn_history.config(state=tk.NORMAL)
            status_label.config(text="✓ Ready")
    
    thread = threading.Thread(target=run, daemon=True)
    thread.start()

# =========================================================================== #
# GUI
# =========================================================================== #
def create_gui():
    window = tk.Tk()
    window.title("VRS Manager by Neil Schmitt (ver. 1114v3)")
    window.geometry("480x800")
    window.resizable(False, False)
    
    bg_color = "#f0f0f0"
    window.configure(bg=bg_color)
    
    def on_closing():
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            window.quit()
            window.destroy()
            sys.exit(0)
    
    window.protocol("WM_DELETE_WINDOW", on_closing)
    
    title_frame = tk.Frame(window, bg=bg_color)
    title_frame.pack(pady=20)
    
    title_label = tk.Label(
        title_frame,
        text="VRS Manager",
        font=("Arial", 18, "bold"),
        bg=bg_color,
        fg="#333333"
    )
    title_label.pack()
    
    subtitle_label = tk.Label(
        title_frame,
        text="Select a process to begin",
        font=("Arial", 10),
        bg=bg_color,
        fg="#666666"
    )
    subtitle_label.pack()
    
    status_label = tk.Label(
        window,
        text="✓ Ready",
        font=("Arial", 10, "italic"),
        bg=bg_color,
        fg="#006600"
    )
    status_label.pack(pady=5)
    
    button_frame = tk.Frame(window, bg=bg_color)
    button_frame.pack(pady=20)
    
    btn_raw = tk.Button(
        button_frame,
        text="Raw Process",
        font=("Arial", 11, "bold"),
        bg="#4CAF50",
        fg="white",
        width=42,
        height=2,
        relief=tk.RAISED,
        bd=3,
        cursor="hand2"
    )
    btn_raw.pack(pady=8)
    
    desc_raw = tk.Label(
        button_frame,
        text="Compare PREVIOUS ↔ CURRENT and detect changes",
        font=("Arial", 9, "italic"),
        bg=bg_color,
        fg="#666666"
    )
    desc_raw.pack()
    
    btn_working = tk.Button(
        button_frame,
        text="Working Process",
        font=("Arial", 11, "bold"),
        bg="#2196F3",
        fg="white",
        width=42,
        height=2,
        relief=tk.RAISED,
        bd=3,
        cursor="hand2"
    )
    btn_working.pack(pady=8)
    
    desc_working = tk.Label(
        button_frame,
        text="Import with intelligent logic (Previous → Current)",
        font=("Arial", 9, "italic"),
        bg=bg_color,
        fg="#666666"
    )
    desc_working.pack()
    
    btn_alllang = tk.Button(
        button_frame,
        text="All Language Process",
        font=("Arial", 11, "bold"),
        bg="#FF9800",
        fg="white",
        width=42,
        height=2,
        relief=tk.RAISED,
        bd=3,
        cursor="hand2"
    )
    btn_alllang.pack(pady=8)
    
    desc_alllang = tk.Label(
        button_frame,
        text="Tri-lingual import (KR/EN/CN flexible updates)",
        font=("Arial", 9, "italic"),
        bg=bg_color,
        fg="#666666"
    )
    desc_alllang.pack()
    
    btn_master = tk.Button(
        button_frame,
        text="Master File Update",
        font=("Arial", 11, "bold"),
        bg="#9C27B0",
        fg="white",
        width=42,
        height=2,
        relief=tk.RAISED,
        bd=3,
        cursor="hand2"
    )
    btn_master.pack(pady=8)
    
    desc_master = tk.Label(
        button_frame,
        text="Update Master File with Working Process output (3-Key Copy-Paste)",
        font=("Arial", 9, "italic"),
        bg=bg_color,
        fg="#666666"
    )
    desc_master.pack()
    
    separator = tk.Frame(window, height=2, bd=1, relief=tk.SUNKEN, bg="#cccccc")
    separator.pack(fill=tk.X, padx=20, pady=10)
    
    btn_history = tk.Button(
        button_frame,
        text="📊 View Update History",
        font=("Arial", 11, "bold"),
        bg="#607D8B",
        fg="white",
        width=42,
        height=2,
        relief=tk.RAISED,
        bd=3,
        cursor="hand2",
        command=show_update_history_viewer
    )
    btn_history.pack(pady=8)
    
    desc_history = tk.Label(
        button_frame,
        text="View complete update history (All processes)",
        font=("Arial", 9, "italic"),
        bg=bg_color,
        fg="#666666"
    )
    desc_history.pack()
    
    btn_raw.config(command=lambda: run_raw_process_thread(btn_raw, btn_working, btn_alllang, btn_master, btn_history, status_label))
    btn_working.config(command=lambda: run_working_process_thread(btn_raw, btn_working, btn_alllang, btn_master, btn_history, status_label))
    btn_alllang.config(command=lambda: run_alllang_process_thread(btn_raw, btn_working, btn_alllang, btn_master, btn_history, status_label))
    btn_master.config(command=lambda: run_master_file_update_thread(btn_raw, btn_working, btn_alllang, btn_master, btn_history, status_label))
    
    footer_label = tk.Label(
        window,
        text="ver. 1114v3 | 4-Tier Key System | Duplicate StrOrigin Fix",
        font=("Arial", 8),
        bg=bg_color,
        fg="#999999"
    )
    footer_label.pack(side=tk.BOTTOM, pady=10)
    
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')
    
    window.mainloop()

# =========================================================================== #
# MAIN
# =========================================================================== #
if __name__ == "__main__":
    create_gui()